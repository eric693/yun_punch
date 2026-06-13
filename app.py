import hashlib
import math
import os
import secrets
import threading
import time
import traceback
import urllib.request
from datetime import date
from functools import wraps

import psycopg
from psycopg.rows import dict_row
from flask import (
    Flask, request, jsonify, render_template,
    session, redirect, url_for, abort
)
from linebot import LineBotApi
from linebot.models import TextSendMessage

from config import (
    SECRET_KEY, ADMIN_PASSWORD, ANTHROPIC_API_KEY,
    DATABASE_URL, RENDER_EXTERNAL_URL, TW_TZ, WEEKDAY_ZH,
)
from utils import (
    _month_ts_range, _month_date_range, _hash_pw, _gps_distance,
    _parse_tw_datetime, _calc_service_years, _eval_formula,
    _roc_date, _roc_year, _month_last_day, _b64url_encode, _b64url_decode,
)
from db import get_db, _init_pool
from serializers import (
    _admin_row, punch_staff_row, punch_record_row, loc_row, punch_req_row,
    ot_req_row, shift_type_row, shift_assign_row, sched_req_row, leave_type_row,
    leave_req_row, leave_balance_row, salary_item_row, salary_record_row, ann_row,
    holiday_row, _finance_cat_row, _finance_rec_row, _recurring_row, _bank_row,
    _payable_row, _expense_row, _perf_template_row, _perf_review_row,
)
from auth import login_required, require_module, require_super
from services import (
    get_line_punch_config, _send_line_punch, _send_line_with_quick_reply,
    _call_line_api, _notify_staff_line, _notify_review_result,
    _broadcast_announcement_line, _calc_punch_hours, _auto_generate_salary,
    _trigger_salary_regen_for_leave, _calc_annual_leave_days,
    _calc_annual_leave_schedule, _calc_leave_days, _get_staff_scheduled_dates,
    _update_leave_balance, _calc_ot_pay, _is_holiday,
    _get_grade_config, _grade_labels, _score_to_grade,
)

app = Flask(__name__)
app.secret_key = SECRET_KEY

# ─── gzip 壓縮（縮小首屏與 JSON 傳輸量）────────────────────────────────────────
try:
    from flask_compress import Compress
    app.config['COMPRESS_MIMETYPES'] = [
        'text/html', 'text/css', 'text/javascript', 'application/javascript',
        'application/json', 'image/svg+xml',
    ]
    app.config['COMPRESS_LEVEL'] = 6
    app.config['COMPRESS_MIN_SIZE'] = 1024
    Compress(app)
    print("[OK] gzip compression enabled")
except ImportError:
    print("[WARN] flask_compress not installed - responses will not be gzipped")


@app.after_request
def _set_cache_headers(resp):
    """靜態資源給長快取；其餘預設不快取，避免後台資料被瀏覽器留存。"""
    if request.path.startswith('/static/'):
        resp.headers.setdefault('Cache-Control', 'public, max-age=86400')
    return resp

print(f"[startup] DATABASE_URL prefix: {DATABASE_URL[:20] if DATABASE_URL else 'NOT SET'}")

# LINE 推播一律用各設定自帶的 channel_access_token 即時建立 LineBotApi(...),
# 故此處不再建立模組級 client(原 line_bot_api/handler 為未使用的死碼)。

# ─── Imports ──────────────────────────────────────────────────────────────────
import json as _json
from datetime import datetime as _dt, timedelta as _td, timezone as _tz

import calendar as _calendar
from datetime import date as _date_cls

# 月份/時間區間、GPS、雜湊、公式等純工具函式已抽至 utils.py

# ─── PostgreSQL ───────────────────────────────────────────────────────────────
# 連線池與 get_db 已抽至 db.py

def init_db():
    if not DATABASE_URL:
        print("[WARNING] DATABASE_URL not set - skipping init_db()")
        return
    try:
        with get_db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS punch_staff (
                    id              SERIAL PRIMARY KEY,
                    name            TEXT NOT NULL UNIQUE,
                    username        TEXT UNIQUE,
                    password_hash   TEXT DEFAULT '',
                    role            TEXT DEFAULT '',
                    active          BOOLEAN DEFAULT TRUE,
                    employee_code   TEXT DEFAULT '',
                    department      TEXT DEFAULT '',
                    position_title  TEXT DEFAULT '',
                    hire_date       DATE,
                    birth_date      DATE,
                    base_salary     NUMERIC(12,2) DEFAULT 0,
                    insured_salary  NUMERIC(12,2) DEFAULT 0,
                    daily_hours     NUMERIC(4,1) DEFAULT 8,
                    ot_rate1        NUMERIC(4,2) DEFAULT 1.33,
                    ot_rate2        NUMERIC(4,2) DEFAULT 1.67,
                    salary_type     TEXT DEFAULT 'monthly',
                    hourly_rate     NUMERIC(12,2) DEFAULT 0,
                    vacation_quota  INT DEFAULT NULL,
                    salary_notes    TEXT DEFAULT '',
                    line_user_id    TEXT,
                    bind_code       TEXT,
                    created_at      TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS punch_records (
                    id            SERIAL PRIMARY KEY,
                    staff_id      INT REFERENCES punch_staff(id) ON DELETE CASCADE,
                    punch_type    TEXT NOT NULL,
                    punched_at    TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    note          TEXT DEFAULT '',
                    is_manual     BOOLEAN DEFAULT FALSE,
                    manual_by     TEXT DEFAULT '',
                    latitude      NUMERIC(10,6),
                    longitude     NUMERIC(10,6),
                    gps_distance  INT,
                    location_name TEXT DEFAULT '',
                    created_at    TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS punch_locations (
                    id            SERIAL PRIMARY KEY,
                    location_name TEXT NOT NULL DEFAULT '打卡地點',
                    lat           NUMERIC(10,6) NOT NULL,
                    lng           NUMERIC(10,6) NOT NULL,
                    radius_m      INT DEFAULT 100,
                    active        BOOLEAN DEFAULT TRUE,
                    created_at    TIMESTAMPTZ DEFAULT NOW(),
                    updated_at    TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS punch_config (
                    id           INT PRIMARY KEY DEFAULT 1,
                    gps_required BOOLEAN DEFAULT FALSE,
                    updated_at   TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            conn.execute("""
                INSERT INTO punch_config (id, gps_required)
                VALUES (1, FALSE)
                ON CONFLICT (id) DO NOTHING
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS line_punch_config (
                    id                   INT PRIMARY KEY DEFAULT 1,
                    channel_access_token TEXT DEFAULT '',
                    channel_secret       TEXT DEFAULT '',
                    enabled              BOOLEAN DEFAULT FALSE,
                    updated_at           TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            conn.execute("""
                INSERT INTO line_punch_config (id)
                VALUES (1)
                ON CONFLICT (id) DO NOTHING
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS schedule_config (
                    month           TEXT PRIMARY KEY,
                    max_off_per_day INT DEFAULT 2,
                    vacation_quota  INT DEFAULT 8,
                    notes           TEXT DEFAULT '',
                    updated_at      TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS schedule_requests (
                    id           SERIAL PRIMARY KEY,
                    staff_id     INT REFERENCES punch_staff(id) ON DELETE CASCADE,
                    month        TEXT NOT NULL,
                    dates        JSONB NOT NULL DEFAULT '[]',
                    status       TEXT DEFAULT 'pending',
                    submit_note  TEXT DEFAULT '',
                    reviewed_by  TEXT DEFAULT '',
                    reviewed_at  TIMESTAMPTZ,
                    review_note  TEXT DEFAULT '',
                    created_at   TIMESTAMPTZ DEFAULT NOW(),
                    updated_at   TIMESTAMPTZ DEFAULT NOW(),
                    UNIQUE(staff_id, month)
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS punch_requests (
                    id            SERIAL PRIMARY KEY,
                    staff_id      INT REFERENCES punch_staff(id) ON DELETE CASCADE,
                    punch_type    TEXT NOT NULL,
                    requested_at  TIMESTAMPTZ NOT NULL,
                    reason        TEXT DEFAULT '',
                    status        TEXT DEFAULT 'pending',
                    reviewed_by   TEXT DEFAULT '',
                    review_note   TEXT DEFAULT '',
                    reviewed_at   TIMESTAMPTZ,
                    created_at    TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS shift_types (
                    id          SERIAL PRIMARY KEY,
                    name        TEXT NOT NULL,
                    start_time  TIME NOT NULL,
                    end_time    TIME NOT NULL,
                    color       TEXT DEFAULT '#4a7bda',
                    departments TEXT DEFAULT '',
                    active      BOOLEAN DEFAULT TRUE,
                    sort_order  INT DEFAULT 0,
                    created_at  TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS shift_assignments (
                    id            SERIAL PRIMARY KEY,
                    staff_id      INT REFERENCES punch_staff(id) ON DELETE CASCADE,
                    shift_type_id INT REFERENCES shift_types(id) ON DELETE CASCADE,
                    shift_date    DATE NOT NULL,
                    note          TEXT DEFAULT '',
                    created_at    TIMESTAMPTZ DEFAULT NOW(),
                    UNIQUE(staff_id, shift_date)
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS overtime_requests (
                    id              SERIAL PRIMARY KEY,
                    staff_id        INT REFERENCES punch_staff(id) ON DELETE CASCADE,
                    request_date    DATE NOT NULL,
                    start_time      TIME NOT NULL,
                    end_time        TIME NOT NULL,
                    ot_hours        NUMERIC(5,2),
                    reason          TEXT DEFAULT '',
                    status          TEXT DEFAULT 'pending',
                    reviewed_by     TEXT DEFAULT '',
                    review_note     TEXT DEFAULT '',
                    ot_pay          NUMERIC(12,2) DEFAULT 0,
                    day_type        TEXT DEFAULT 'weekday',
                    reviewed_at     TIMESTAMPTZ,
                    created_at      TIMESTAMPTZ DEFAULT NOW()
                )
            """)

            # Seed default shifts if empty
            existing_shifts = conn.execute("SELECT COUNT(*) as cnt FROM shift_types").fetchone()
            if existing_shifts['cnt'] == 0:
                defaults = [
                    ('吧台班',  '08:00', '16:00', '#8b5cf6', '吧台', 1),
                    ('外場A班', '09:00', '17:00', '#2e9e6b', '外場', 2),
                    ('外場B班', '14:00', '22:00', '#0ea5e9', '外場', 3),
                    ('廚房A班', '08:00', '16:00', '#e07b2a', '廚房', 4),
                    ('廚房B班', '12:00', '20:00', '#d64242', '廚房', 5),
                ]
                for name, st, et, color, dept, sort in defaults:
                    conn.execute(
                        "INSERT INTO shift_types (name,start_time,end_time,color,departments,sort_order) VALUES (%s,%s,%s,%s,%s,%s)",
                        (name, st, et, color, dept, sort)
                    )

        print("[OK] Database tables created")
    except Exception as e:
        print(f"[ERROR] init_db failed: {e}")
        raise

    # Schema migrations (each in its own connection to avoid transaction abort)
    migrations = [
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS username TEXT",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS password_hash TEXT DEFAULT ''",
        "ALTER TABLE punch_records ADD COLUMN IF NOT EXISTS latitude NUMERIC(10,6)",
        "ALTER TABLE punch_records ADD COLUMN IF NOT EXISTS longitude NUMERIC(10,6)",
        "ALTER TABLE punch_records ADD COLUMN IF NOT EXISTS gps_distance INT",
        "ALTER TABLE punch_records ADD COLUMN IF NOT EXISTS location_name TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS line_user_id TEXT",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS bind_code TEXT",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS employee_code TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS department TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS position_title TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS hire_date DATE",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS birth_date DATE",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS base_salary NUMERIC(12,2) DEFAULT 0",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS insured_salary NUMERIC(12,2) DEFAULT 0",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS salary_notes TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS daily_hours NUMERIC(4,1) DEFAULT 8",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS ot_rate1 NUMERIC(4,2) DEFAULT 1.33",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS ot_rate2 NUMERIC(4,2) DEFAULT 1.67",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS ot_rate3 NUMERIC(4,2) DEFAULT 2.0",
        "ALTER TABLE leave_requests ADD COLUMN IF NOT EXISTS document_id INT REFERENCES finance_documents(id) ON DELETE SET NULL",
        "ALTER TABLE leave_requests ADD COLUMN IF NOT EXISTS start_time TEXT DEFAULT ''",
        "ALTER TABLE leave_requests ADD COLUMN IF NOT EXISTS end_time TEXT DEFAULT ''",
        "ALTER TABLE finance_documents ADD COLUMN IF NOT EXISTS image_data TEXT",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS salary_type TEXT DEFAULT 'monthly'",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS hourly_rate NUMERIC(12,2) DEFAULT 0",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS vacation_quota INT DEFAULT NULL",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS bank_code TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS bank_name TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS bank_branch TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS bank_account TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS account_holder TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS password_plain TEXT DEFAULT ''",
        "ALTER TABLE overtime_requests ADD COLUMN IF NOT EXISTS day_type TEXT DEFAULT 'weekday'",
        "ALTER TABLE overtime_requests ALTER COLUMN start_time DROP NOT NULL",
        "ALTER TABLE overtime_requests ALTER COLUMN end_time DROP NOT NULL",
        # 員工個人/保險欄位
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS national_id TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS gender TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS insurance_type TEXT DEFAULT 'regular'",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS address TEXT DEFAULT ''",
        # 多店
        """CREATE TABLE IF NOT EXISTS stores (
            id         SERIAL PRIMARY KEY,
            name       TEXT NOT NULL,
            code       TEXT UNIQUE,
            address    TEXT DEFAULT '',
            active     BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMPTZ DEFAULT NOW()
        )""",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS store_id INT REFERENCES stores(id) ON DELETE SET NULL",
        "ALTER TABLE punch_locations ADD COLUMN IF NOT EXISTS store_id INT REFERENCES stores(id) ON DELETE SET NULL",
        "ALTER TABLE admin_accounts ADD COLUMN IF NOT EXISTS store_ids JSONB DEFAULT '[]'",
        "ALTER TABLE schedule_requests ADD COLUMN IF NOT EXISTS updated_at TIMESTAMPTZ DEFAULT NOW()",
        """CREATE TABLE IF NOT EXISTS shift_staffing_requirements (
            id            SERIAL PRIMARY KEY,
            shift_type_id INT REFERENCES shift_types(id) ON DELETE CASCADE,
            day_of_week   SMALLINT NOT NULL,
            required_count INT NOT NULL DEFAULT 1,
            created_at    TIMESTAMPTZ DEFAULT NOW(),
            updated_at    TIMESTAMPTZ DEFAULT NOW(),
            UNIQUE(shift_type_id, day_of_week)
        )""",
        """CREATE TABLE IF NOT EXISTS admin_accounts (
            id              SERIAL PRIMARY KEY,
            username        TEXT NOT NULL UNIQUE,
            password_hash   TEXT NOT NULL,
            display_name    TEXT DEFAULT '',
            permissions     JSONB DEFAULT '[]',
            is_super        BOOLEAN DEFAULT FALSE,
            active          BOOLEAN DEFAULT TRUE,
            created_at      TIMESTAMPTZ DEFAULT NOW(),
            last_login_at   TIMESTAMPTZ
        )""",
        # ── 效能索引 ──────────────────────────────────────────────────────────
        # punch_records: 最常被查詢的資料表,依 staff_id + punched_at 過濾
        "CREATE INDEX IF NOT EXISTS idx_pr_staff_at   ON punch_records(staff_id, punched_at DESC)",
        "CREATE INDEX IF NOT EXISTS idx_pr_at          ON punch_records(punched_at DESC)",
        # leave_requests: 依員工 + 日期 + 狀態過濾
        "CREATE INDEX IF NOT EXISTS idx_lr_staff_date  ON leave_requests(staff_id, start_date)",
        "CREATE INDEX IF NOT EXISTS idx_lr_status      ON leave_requests(status)",
        # overtime_requests
        "CREATE INDEX IF NOT EXISTS idx_ot_staff_date  ON overtime_requests(staff_id, request_date)",
        "CREATE INDEX IF NOT EXISTS idx_ot_status      ON overtime_requests(status)",
        # salary_records: 月份批次查詢
        "CREATE INDEX IF NOT EXISTS idx_sr_month       ON salary_records(month)",
        # punch_staff: LINE bot 依 line_user_id 查詢
        "CREATE INDEX IF NOT EXISTS idx_ps_line_uid    ON punch_staff(line_user_id) WHERE line_user_id IS NOT NULL",
        # punch_requests
        "CREATE INDEX IF NOT EXISTS idx_pq_staff_status ON punch_requests(staff_id, status)",
        # shift_assignments: 月份整表查詢用 shift_date;薪資/請假的每員工查詢用複合索引
        "CREATE INDEX IF NOT EXISTS idx_sa_date        ON shift_assignments(shift_date)",
        "CREATE INDEX IF NOT EXISTS idx_sa_staff_date  ON shift_assignments(staff_id, shift_date)",
        # leave_requests: dashboard/異常偵測常以 status + 日期區間過濾(無 staff_id)
        "CREATE INDEX IF NOT EXISTS idx_lr_status_date ON leave_requests(status, start_date)",
        # finance_records
        "CREATE INDEX IF NOT EXISTS idx_fr_date        ON finance_records(record_date)",
        # schedule_requests: 月份查詢
        "CREATE INDEX IF NOT EXISTS idx_schedr_month   ON schedule_requests(month)",
        # leave_balances
        "CREATE INDEX IF NOT EXISTS idx_lb_staff_year  ON leave_balances(staff_id, year)",
    ]
    for sql in migrations:
        try:
            with get_db() as mc:
                mc.execute(sql)
        except Exception as me:
            print(f"[MIGRATION SKIP] {sql[:70]}: {me}")

    # Seed default super admin; always sync password from ADMIN_PASSWORD env var
    try:
        all_modules = _json.dumps(['punch','sched','leave','salary','ann','holiday','finance'])
        pw_hash = _hash_pw(ADMIN_PASSWORD)
        with get_db() as conn:
            existing = conn.execute(
                "SELECT id FROM admin_accounts WHERE username='admin'"
            ).fetchone()
            if existing:
                conn.execute(
                    "UPDATE admin_accounts SET password_hash=%s, is_super=TRUE WHERE username='admin'",
                    (pw_hash,)
                )
                print("[OK] admin password synced from ADMIN_PASSWORD env var")
            else:
                conn.execute("""
                    INSERT INTO admin_accounts (username, password_hash, display_name, permissions, is_super)
                    VALUES (%s,%s,'超級管理員',%s,TRUE)
                """, ('admin', pw_hash, all_modules))
                print("[OK] Default super admin seeded (username: admin)")
    except Exception as e:
        print(f"[WARN] admin seed: {e}")

    # 確保預設店家存在,並補齊舊資料
    try:
        with get_db() as conn:
            conn.execute("INSERT INTO stores (name, code) VALUES ('主店','main') ON CONFLICT (code) DO NOTHING")
            conn.execute("UPDATE punch_staff     SET store_id=(SELECT id FROM stores WHERE code='main') WHERE store_id IS NULL")
            conn.execute("UPDATE punch_locations SET store_id=(SELECT id FROM stores WHERE code='main') WHERE store_id IS NULL")
    except Exception as e:
        print(f"[WARN] store seed: {e}")

    print("[OK] Database initialised")


_init_pool()
init_db()

# ─── Keep-Alive ───────────────────────────────────────────────────────────────

def keep_alive():
    time.sleep(10)
    while True:
        try:
            base = RENDER_EXTERNAL_URL.rstrip('/') if RENDER_EXTERNAL_URL else 'http://localhost:5000'
            urllib.request.urlopen(
                urllib.request.Request(f'{base}/health', headers={'User-Agent': 'KeepAlive/1.0'}),
                timeout=10
            )
        except Exception as e:
            print(f"[keep-alive] ping failed: {e}")
        time.sleep(14 * 60)

threading.Thread(target=keep_alive, daemon=True).start()


# ─── 特休自動同步 ─────────────────────────────────────────────────────────────

def _run_annual_leave_sync():
    """依勞基法第38條,依到職日計算特休天數,寫入 leave_balances.每日午夜自動執行."""
    from datetime import date as _d_sync
    year = str(_d_sync.today().year)
    try:
        with get_db() as conn:
            # 多 worker 防重:取得 advisory lock,拿不到就跳過
            got_lock = conn.execute("SELECT pg_try_advisory_lock(1001)").fetchone()[0]
            if not got_lock:
                return
            try:
                staff_list = conn.execute(
                    "SELECT id, name, hire_date FROM punch_staff WHERE active=TRUE AND hire_date IS NOT NULL"
                ).fetchall()
                lt = conn.execute("SELECT id FROM leave_types WHERE code='annual'").fetchone()
                if not lt:
                    return
                lt_id = lt['id']
                for s in staff_list:
                    days = _calc_annual_leave_days(s['hire_date'])
                    conn.execute("""
                        INSERT INTO leave_balances (staff_id, leave_type_id, year, total_days, used_days)
                        VALUES (%s,%s,%s,%s,0)
                        ON CONFLICT (staff_id, leave_type_id, year) DO UPDATE
                          SET total_days=EXCLUDED.total_days, updated_at=NOW()
                    """, (s['id'], lt_id, int(year), days))
            finally:
                conn.execute("SELECT pg_advisory_unlock(1001)")
    except Exception as e:
        print(f"[annual_leave_sync] {e}")


def _annual_leave_sync_loop():
    import time as _time_sync
    # 啟動時立即執行一次
    _run_annual_leave_sync()
    while True:
        # 計算距離明天 00:05 台北時間的秒數
        now = _dt.now(TW_TZ)
        tmr = (now + _td(days=1)).date()
        tomorrow_05 = _dt(tmr.year, tmr.month, tmr.day, 0, 5, tzinfo=TW_TZ)
        sleep_secs = (tomorrow_05 - now).total_seconds()
        if sleep_secs < 0:
            sleep_secs = 3600
        _time_sync.sleep(sleep_secs)
        _run_annual_leave_sync()


threading.Thread(target=_annual_leave_sync_loop, daemon=True).start()


# ─── 每月1日自動產生上月薪資 ──────────────────────────────────────────────────

def _run_monthly_salary_auto_generate():
    """每月1日 00:10(台北時間)自動產生所有在職員工上月薪資(draft)."""
    from datetime import date as _d_sal
    import json as _json_sal

    # 計算上月月份字串,例如今天是 2026-05-01 -> 上月為 2026-04
    today = _d_sal.today()
    if today.month == 1:
        last_month = f"{today.year - 1}-12"
    else:
        last_month = f"{today.year}-{today.month - 1:02d}"

    print(f"[salary_auto] 開始自動產生 {last_month} 薪資...")
    try:
        with get_db() as conn:
            # 多 worker 防重:取得 advisory lock,拿不到就跳過
            got_lock = conn.execute("SELECT pg_try_advisory_lock(1002)").fetchone()[0]
            if not got_lock:
                print(f"[salary_auto] 另一 worker 正在執行,略過.")
                return
            try:
              staff_list = conn.execute(
                "SELECT * FROM punch_staff WHERE active=TRUE"
              ).fetchall()
              generated = 0
              for staff in staff_list:
                data = _auto_generate_salary(conn, dict(staff), last_month)
                items_json = _json_sal.dumps(data['items'], ensure_ascii=False)
                conn.execute("""
                    INSERT INTO salary_records
                      (staff_id, month, base_salary, insured_salary, work_days, actual_days,
                       leave_days, unpaid_days, ot_pay, allowance_total, deduction_total,
                       net_pay, items, status, updated_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s::jsonb,'draft',NOW())
                    ON CONFLICT (staff_id, month) DO NOTHING
                """, (
                    data['staff_id'], last_month, data['base_salary'], data['insured_salary'],
                    data['work_days'], data['actual_days'], data['leave_days'], data['unpaid_days'],
                    data['ot_pay'], data['allowance_total'], data['deduction_total'],
                    data['net_pay'], items_json,
                ))
                generated += 1
              print(f"[salary_auto] {last_month} 薪資自動產生完成,共 {generated} 筆.")
            finally:
                conn.execute("SELECT pg_advisory_unlock(1002)")
    except Exception as e:
        print(f"[salary_auto] 產生失敗:{e}")


def _monthly_salary_auto_loop():
    import time as _time_sal
    # 等待 app 完全啟動後再進入迴圈
    _time_sal.sleep(30)
    while True:
        now = _dt.now(TW_TZ)
        # 只在每月1日執行
        if now.day == 1:
            target = _dt(now.year, now.month, 1, 0, 10, tzinfo=TW_TZ)
            if now >= target:
                _run_monthly_salary_auto_generate()
                # 執行完後等待 25 小時,避免同一天重複執行
                _time_sal.sleep(25 * 3600)
                continue
        # 計算下次執行時間:下個月1日 00:10
        if now.month == 12:
            next_run = _dt(now.year + 1, 1, 1, 0, 10, tzinfo=TW_TZ)
        else:
            next_run = _dt(now.year, now.month + 1, 1, 0, 10, tzinfo=TW_TZ)
        sleep_secs = (next_run - now).total_seconds()
        print(f"[salary_auto] 下次執行時間:{next_run.strftime('%Y-%m-%d %H:%M')}(台北時間),約 {sleep_secs/3600:.1f} 小時後")
        _time_sal.sleep(max(sleep_secs, 60))


threading.Thread(target=_monthly_salary_auto_loop, daemon=True).start()


# ─── Health ───────────────────────────────────────────────────────────────────

@app.route('/health')
def health():
    try:
        with get_db() as conn:
            conn.execute('SELECT 1')
        return jsonify({'status': 'ok', 'db': 'connected'}), 200
    except Exception as e:
        return jsonify({'status': 'error', 'detail': str(e)}), 500

# ─── Admin Auth ───────────────────────────────────────────────────────────────

# login_required / require_module / require_super 已抽至 auth.py

@app.route('/')
def index():
    return redirect(url_for('admin_login'))

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if not username or not password:
            error = '請輸入帳號與密碼'
        else:
            with get_db() as conn:
                row = conn.execute(
                    "SELECT * FROM admin_accounts WHERE username=%s AND active=TRUE",
                    (username,)
                ).fetchone()
            if row and row['password_hash'] == _hash_pw(password):
                perms = row['permissions']
                if isinstance(perms, str):
                    try: perms = _json.loads(perms)
                    except: perms = []
                session['logged_in']          = True
                session['admin_id']           = row['id']
                session['admin_username']     = row['username']
                session['admin_display_name'] = row['display_name'] or row['username']
                session['admin_permissions']  = perms
                session['admin_is_super']     = bool(row['is_super'])
                with get_db() as conn:
                    conn.execute("UPDATE admin_accounts SET last_login_at=NOW() WHERE id=%s", (row['id'],))
                return redirect(url_for('admin_dashboard'))
            error = '帳號或密碼錯誤'
    return render_template('login.html', error=error)

@app.route('/admin/logout')
def admin_logout():
    session.clear()
    return redirect(url_for('admin_login'))

@app.route('/admin')
@app.route('/admin/')
@login_required
def admin_dashboard():
    perms    = session.get('admin_permissions') or []
    is_super = bool(session.get('admin_is_super'))
    return render_template('admin.html',
        admin_display_name=session.get('admin_display_name',''),
        admin_permissions=perms,
        admin_is_super=is_super,
    )

# ── Admin Accounts API ────────────────────────────────────────────────────────


@app.route('/api/admin/me', methods=['GET'])
@login_required
def api_admin_me():
    return jsonify({
        'id':           session.get('admin_id'),
        'username':     session.get('admin_username'),
        'display_name': session.get('admin_display_name'),
        'permissions':  session.get('admin_permissions') or [],
        'is_super':     bool(session.get('admin_is_super')),
    })

@app.route('/api/admin/accounts', methods=['GET'])
@require_super
def api_admin_accounts_list():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM admin_accounts ORDER BY id").fetchall()
    return jsonify([_admin_row(r) for r in rows])

@app.route('/api/admin/accounts', methods=['POST'])
@require_super
def api_admin_account_create():
    b = request.get_json(force=True)
    username = b.get('username','').strip()
    password = b.get('password','').strip()
    if not username: return jsonify({'error': '帳號為必填'}), 400
    if not password or len(password) < 4: return jsonify({'error': '密碼至少 4 個字元'}), 400
    perms = b.get('permissions', [])
    with get_db() as conn:
        try:
            row = conn.execute("""
                INSERT INTO admin_accounts (username, password_hash, display_name, permissions, is_super, active)
                VALUES (%s,%s,%s,%s,%s,%s) RETURNING *
            """, (username, _hash_pw(password), b.get('display_name','').strip(),
                  _json.dumps(perms), bool(b.get('is_super', False)), True)).fetchone()
        except Exception as e:
            if 'unique' in str(e).lower(): return jsonify({'error': '帳號已存在'}), 409
            return jsonify({'error': str(e)}), 500
    return jsonify(_admin_row(row)), 201

@app.route('/api/admin/accounts/<int:aid>', methods=['PUT'])
@require_super
def api_admin_account_update(aid):
    b = request.get_json(force=True)
    username = b.get('username','').strip()
    if not username: return jsonify({'error': '帳號為必填'}), 400
    password = b.get('password','').strip()
    perms = b.get('permissions', [])
    with get_db() as conn:
        if password:
            if len(password) < 4: return jsonify({'error': '密碼至少 4 個字元'}), 400
            row = conn.execute("""
                UPDATE admin_accounts SET username=%s, password_hash=%s, display_name=%s,
                  permissions=%s, is_super=%s, active=%s WHERE id=%s RETURNING *
            """, (username, _hash_pw(password), b.get('display_name','').strip(),
                  _json.dumps(perms), bool(b.get('is_super', False)),
                  bool(b.get('active', True)), aid)).fetchone()
        else:
            row = conn.execute("""
                UPDATE admin_accounts SET username=%s, display_name=%s,
                  permissions=%s, is_super=%s, active=%s WHERE id=%s RETURNING *
            """, (username, b.get('display_name','').strip(),
                  _json.dumps(perms), bool(b.get('is_super', False)),
                  bool(b.get('active', True)), aid)).fetchone()
    return jsonify(_admin_row(row)) if row else ('', 404)

@app.route('/api/admin/accounts/<int:aid>', methods=['DELETE'])
@require_super
def api_admin_account_delete(aid):
    if aid == session.get('admin_id'):
        return jsonify({'error': '不能刪除自己的帳號'}), 400
    with get_db() as conn:
        conn.execute("DELETE FROM admin_accounts WHERE id=%s", (aid,))
    return jsonify({'deleted': aid})

# ─── Shared Helpers ───────────────────────────────────────────────────────────











# ═══════════════════════════════════════════════════════════════════
# Employee Punch Page
# ═══════════════════════════════════════════════════════════════════

@app.route('/punch')
@app.route('/staff')
def punch_page():
    return render_template('staff.html')

# ── Employee Session ──────────────────────────────────────────────

@app.route('/api/punch/login', methods=['POST'])
def api_punch_login():
    b = request.get_json(force=True)
    username = b.get('username', '').strip()
    password = b.get('password', '').strip()
    if not username or not password:
        return jsonify({'error': '請輸入帳號及密碼'}), 400
    with get_db() as conn:
        staff = conn.execute(
            "SELECT * FROM punch_staff WHERE username=%s AND active=TRUE", (username,)
        ).fetchone()
    if not staff or staff['password_hash'] != _hash_pw(password):
        return jsonify({'error': '帳號或密碼錯誤'}), 401
    session['punch_staff_id']   = staff['id']
    session['punch_staff_name'] = staff['name']
    return jsonify({'id': staff['id'], 'name': staff['name'], 'role': staff['role']})

@app.route('/api/punch/logout', methods=['POST'])
def api_punch_logout():
    session.pop('punch_staff_id', None)
    session.pop('punch_staff_name', None)
    return jsonify({'ok': True})

@app.route('/api/punch/me', methods=['GET'])
def api_punch_me():
    sid = session.get('punch_staff_id')
    if not sid:
        return jsonify({'error': 'not logged in'}), 401
    with get_db() as conn:
        staff = conn.execute(
            "SELECT id,name,role FROM punch_staff WHERE id=%s AND active=TRUE", (sid,)
        ).fetchone()
    if not staff:
        session.pop('punch_staff_id', None)
        return jsonify({'error': 'not logged in'}), 401
    return jsonify(dict(staff))

# ── GPS Settings ──────────────────────────────────────────────────

@app.route('/api/punch/settings', methods=['GET'])
def api_punch_settings_get():
    """Public: GPS config + active locations for the punch page."""
    with get_db() as conn:
        cfg  = conn.execute("SELECT * FROM punch_config WHERE id=1").fetchone()
        locs = conn.execute(
            "SELECT * FROM punch_locations WHERE active=TRUE ORDER BY id"
        ).fetchall()
    return jsonify({
        'gps_required': cfg['gps_required'] if cfg else False,
        'locations': [loc_row(r) for r in locs]
    })

@app.route('/api/punch/config', methods=['PUT'])
@login_required
def api_punch_config_update():
    b = request.get_json(force=True)
    gps_required = bool(b.get('gps_required', False))
    with get_db() as conn:
        conn.execute(
            "UPDATE punch_config SET gps_required=%s, updated_at=NOW() WHERE id=1",
            (gps_required,)
        )
    return jsonify({'gps_required': gps_required})

@app.route('/api/punch/locations', methods=['GET'])
@login_required
def api_punch_locations_list():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM punch_locations ORDER BY id").fetchall()
    return jsonify([loc_row(r) for r in rows])

@app.route('/api/punch/locations', methods=['POST'])
@login_required
def api_punch_locations_create():
    b = request.get_json(force=True)
    name = b.get('location_name', '').strip() or '打卡地點'
    try:
        lat = float(b['lat']); lng = float(b['lng'])
    except Exception:
        return jsonify({'error': '請填入有效的緯度和經度'}), 400
    radius_m = int(b.get('radius_m') or 100)
    with get_db() as conn:
        row = conn.execute(
            "INSERT INTO punch_locations (location_name, lat, lng, radius_m) VALUES (%s,%s,%s,%s) RETURNING *",
            (name, lat, lng, radius_m)
        ).fetchone()
    return jsonify(loc_row(row)), 201

@app.route('/api/punch/locations/<int:lid>', methods=['PUT'])
@login_required
def api_punch_locations_update(lid):
    b = request.get_json(force=True)
    name = b.get('location_name', '').strip() or '打卡地點'
    try:
        lat = float(b['lat']); lng = float(b['lng'])
    except Exception:
        return jsonify({'error': '請填入有效的緯度和經度'}), 400
    radius_m = int(b.get('radius_m') or 100)
    active   = bool(b.get('active', True))
    with get_db() as conn:
        row = conn.execute(
            "UPDATE punch_locations SET location_name=%s,lat=%s,lng=%s,radius_m=%s,active=%s,updated_at=NOW() WHERE id=%s RETURNING *",
            (name, lat, lng, radius_m, active, lid)
        ).fetchone()
    return jsonify(loc_row(row)) if row else ('', 404)

@app.route('/api/punch/locations/<int:lid>', methods=['DELETE'])
@login_required
def api_punch_locations_delete(lid):
    with get_db() as conn:
        conn.execute("DELETE FROM punch_locations WHERE id=%s", (lid,))
    return jsonify({'deleted': lid})

# ── Clock In/Out ──────────────────────────────────────────────────

@app.route('/api/punch/clock', methods=['POST'])
def api_punch_clock():
    sid = session.get('punch_staff_id')
    if not sid:
        return jsonify({'error': '請先登入'}), 401

    b          = request.get_json(force=True)
    punch_type = b.get('punch_type')
    lat        = b.get('lat')
    lng        = b.get('lng')

    if punch_type not in ('in', 'out', 'break_out', 'break_in'):
        return jsonify({'error': '無效的打卡類型'}), 400

    with get_db() as conn:
        staff = conn.execute(
            "SELECT * FROM punch_staff WHERE id=%s AND active=TRUE", (sid,)
        ).fetchone()
        if not staff:
            return jsonify({'error': '員工不存在'}), 404
        cfg  = conn.execute("SELECT * FROM punch_config WHERE id=1").fetchone()
        locs = conn.execute("SELECT * FROM punch_locations WHERE active=TRUE").fetchall()

    gps_required = cfg['gps_required'] if cfg else False
    gps_distance = None
    matched_loc  = None

    if lat is not None and lng is not None and locs:
        for loc in locs:
            d = _gps_distance(lat, lng, float(loc['lat']), float(loc['lng']))
            if gps_distance is None or d < gps_distance:
                gps_distance = d
                matched_loc  = loc

    if gps_required:
        if lat is None or lng is None:
            return jsonify({'error': '無法取得 GPS,請允許定位權限後重試'}), 403
        if not locs:
            return jsonify({'error': '管理員尚未設定任何打卡地點'}), 403
        if gps_distance is None or gps_distance > int(matched_loc['radius_m']):
            return jsonify({
                'error': f'距離最近地點「{matched_loc["location_name"]}」{gps_distance} 公尺,超出允許範圍({matched_loc["radius_m"]} 公尺)',
                'distance': gps_distance,
                'radius': int(matched_loc['radius_m'])
            }), 403

    with get_db() as conn:
        recent = conn.execute("""
            SELECT id FROM punch_records
            WHERE staff_id=%s AND punch_type=%s
              AND punched_at > NOW() - INTERVAL '1 minute'
        """, (sid, punch_type)).fetchone()
        if recent:
            return jsonify({'error': '1 分鐘內已打過卡'}), 429

        matched_name = matched_loc['location_name'] if matched_loc else ''
        row = conn.execute("""
            INSERT INTO punch_records
              (staff_id, punch_type, latitude, longitude, gps_distance, location_name)
            VALUES (%s,%s,%s,%s,%s,%s) RETURNING *
        """, (sid, punch_type, lat, lng, gps_distance, matched_name)).fetchone()

    d = punch_record_row(row)
    d['staff_name']   = staff['name']
    d['gps_distance'] = gps_distance
    return jsonify(d), 201

@app.route('/api/punch/today', methods=['GET'])
def api_punch_today():
    sid = session.get('punch_staff_id')
    if not sid:
        return jsonify([])
    with get_db() as conn:
        rows = conn.execute("""
            SELECT pr.*, ps.name as staff_name
            FROM punch_records pr JOIN punch_staff ps ON ps.id=pr.staff_id
            WHERE pr.staff_id=%s
              AND (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
                = (NOW() AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY pr.punched_at ASC
        """, (sid,)).fetchall()
    return jsonify([punch_record_row(r) for r in rows])

@app.route('/api/punch/my-records', methods=['GET'])
def api_punch_my_records():
    """Employee self-service: own punch records for a month."""
    sid = session.get('punch_staff_id')
    if not sid:
        return jsonify({'error': 'not logged in'}), 401
    month = request.args.get('month', '')
    if not month:
        from datetime import timezone as _tz, timedelta as _tda
        month = _dt.now(_tz(_tda(hours=8))).strftime('%Y-%m')
    _ts_s, _ts_e = _month_ts_range(month)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT punch_type, punched_at, gps_distance, location_name, is_manual
            FROM punch_records
            WHERE staff_id=%s
              AND punched_at >= %s AND punched_at < %s
            ORDER BY punched_at ASC
        """, (sid, _ts_s, _ts_e)).fetchall()
    from datetime import timezone as _tz2, timedelta as _tdb
    TW = _tz2(_tdb(hours=8))
    LABEL = {'in': '上班', 'out': '下班', 'break_out': '休息開始', 'break_in': '休息結束'}
    result = {}
    for r in rows:
        pa = r['punched_at']
        if pa.tzinfo is None:
            from datetime import timezone as _utz
            pa = pa.replace(tzinfo=_utz.utc)
        pa_tw    = pa.astimezone(TW)
        date_str = pa_tw.strftime('%Y-%m-%d')
        time_str = pa_tw.strftime('%H:%M')
        if date_str not in result:
            result[date_str] = []
        result[date_str].append({
            'type':          r['punch_type'],
            'label':         LABEL.get(r['punch_type'], r['punch_type']),
            'time':          time_str,
            'gps_distance':  r['gps_distance'],
            'location_name': r['location_name'] or '',
            'is_manual':     bool(r['is_manual']),
        })
    return jsonify({'month': month, 'records': result})

# ── Admin: Staff CRUD ─────────────────────────────────────────────

@app.route('/api/punch/staff', methods=['GET'])
@login_required
def api_punch_staff_list():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM punch_staff ORDER BY name").fetchall()
    return jsonify([punch_staff_row(r) for r in rows])

@app.route('/api/punch/staff', methods=['POST'])
@login_required
def api_punch_staff_create():
    b        = request.get_json(force=True)
    name     = b.get('name', '').strip()
    username = b.get('username', '').strip()
    password = b.get('password', '').strip()
    if not name:     return jsonify({'error': '姓名為必填'}), 400
    if not username: return jsonify({'error': '帳號為必填'}), 400
    if not password or len(password) < 4:
        return jsonify({'error': '密碼至少 4 個字元'}), 400
    employee_code = b.get('employee_code', '') or None
    if employee_code: employee_code = employee_code.strip() or None
    department     = (b.get('department') or '').strip()
    hire_date      = b.get('hire_date') or None
    birth_date     = b.get('birth_date') or None
    bank_code      = (b.get('bank_code') or '').strip()
    bank_name      = (b.get('bank_name') or '').strip()
    bank_branch    = (b.get('bank_branch') or '').strip()
    bank_account   = (b.get('bank_account') or '').strip()
    account_holder = (b.get('account_holder') or '').strip()
    try:
        with get_db() as conn:
            default_store = conn.execute(
                "SELECT id FROM stores ORDER BY id LIMIT 1"
            ).fetchone()
            default_store_id = default_store['id'] if default_store else None
            row = conn.execute("""
                INSERT INTO punch_staff
                  (name, username, password_hash, password_plain, role, employee_code,
                   department, hire_date, birth_date,
                   bank_code, bank_name, bank_branch, bank_account, account_holder, store_id)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *
            """, (name, username, _hash_pw(password), password, b.get('role', '').strip(), employee_code,
                  department, hire_date, birth_date,
                  bank_code, bank_name, bank_branch, bank_account, account_holder,
                  default_store_id)).fetchone()
        return jsonify(punch_staff_row(row)), 201
    except psycopg.errors.UniqueViolation:
        return jsonify({'error': '姓名或帳號已存在,請換一個'}), 409
    except Exception as e:
        print(f"[punch_staff_create] error: {e}")
        # Check if it's a unique constraint in the error message
        if 'unique' in str(e).lower() or 'duplicate' in str(e).lower():
            return jsonify({'error': '姓名或帳號已存在,請換一個'}), 409
        return jsonify({'error': f'新增失敗:{str(e)}'}), 500

@app.route('/api/punch/staff/<int:sid>', methods=['PUT'])
@login_required
def api_punch_staff_update(sid):
    b             = request.get_json(force=True)
    name          = b.get('name', '').strip()
    username      = b.get('username', '').strip()
    password      = b.get('password', '').strip()
    role          = b.get('role', '').strip()
    active        = bool(b.get('active', True))
    employee_code = b.get('employee_code', '') or None
    if employee_code: employee_code = employee_code.strip() or None
    bank_code      = (b.get('bank_code') or '').strip()
    bank_name      = (b.get('bank_name') or '').strip()
    bank_branch    = (b.get('bank_branch') or '').strip()
    bank_account   = (b.get('bank_account') or '').strip()
    account_holder = (b.get('account_holder') or '').strip()
    department     = (b.get('department') or '').strip()
    hire_date      = b.get('hire_date') or None
    birth_date     = b.get('birth_date') or None
    if not name or not username:
        return jsonify({'error': '姓名和帳號為必填'}), 400
    with get_db() as conn:
        if password:
            if len(password) < 4:
                return jsonify({'error': '密碼至少 4 個字元'}), 400
            row = conn.execute("""
                UPDATE punch_staff
                SET name=%s,username=%s,password_hash=%s,password_plain=%s,role=%s,active=%s,employee_code=%s,
                    department=%s,hire_date=%s,birth_date=%s,
                    bank_code=%s,bank_name=%s,bank_branch=%s,bank_account=%s,account_holder=%s
                WHERE id=%s RETURNING *
            """, (name, username, _hash_pw(password), password, role, active, employee_code,
                  department, hire_date, birth_date,
                  bank_code, bank_name, bank_branch, bank_account, account_holder, sid)).fetchone()
        else:
            row = conn.execute("""
                UPDATE punch_staff
                SET name=%s,username=%s,role=%s,active=%s,employee_code=%s,
                    department=%s,hire_date=%s,birth_date=%s,
                    bank_code=%s,bank_name=%s,bank_branch=%s,bank_account=%s,account_holder=%s
                WHERE id=%s RETURNING *
            """, (name, username, role, active, employee_code,
                  department, hire_date, birth_date,
                  bank_code, bank_name, bank_branch, bank_account, account_holder, sid)).fetchone()
    return jsonify(punch_staff_row(row)) if row else ('', 404)

@app.route('/api/punch/staff/<int:sid>', methods=['DELETE'])
@login_required
def api_punch_staff_delete(sid):
    with get_db() as conn:
        punch_cnt = conn.execute(
            "SELECT COUNT(*) AS n FROM punch_records WHERE staff_id=%s", (sid,)
        ).fetchone()['n']
        leave_cnt = conn.execute(
            "SELECT COUNT(*) AS n FROM leave_requests WHERE staff_id=%s", (sid,)
        ).fetchone()['n']
        salary_cnt = conn.execute(
            "SELECT COUNT(*) AS n FROM salary_records WHERE staff_id=%s", (sid,)
        ).fetchone()['n']
        if punch_cnt + leave_cnt + salary_cnt > 0:
            return jsonify({
                'error': f'此員工有歷史資料(打卡 {punch_cnt} 筆、假單 {leave_cnt} 筆、薪資 {salary_cnt} 筆),'
                         '請改用「離職」功能停用帳號以保留記錄.'
            }), 409
        conn.execute("DELETE FROM punch_staff WHERE id=%s", (sid,))
    return jsonify({'deleted': sid})

# ── Admin: Punch Records ──────────────────────────────────────────

@app.route('/api/punch/records', methods=['GET'])
@login_required
def api_punch_records():
    staff_id  = request.args.get('staff_id')
    date_from = request.args.get('date_from')
    date_to   = request.args.get('date_to')
    month     = request.args.get('month')

    conds, params = ["TRUE"], []
    if staff_id: conds.append("pr.staff_id=%s"); params.append(int(staff_id))
    if month:
        _ts_s, _ts_e = _month_ts_range(month)
        conds.append("pr.punched_at >= %s AND pr.punched_at < %s"); params += [_ts_s, _ts_e]
    elif date_from:
        conds.append("(pr.punched_at AT TIME ZONE 'Asia/Taipei')::date>=%s"); params.append(date_from)
        if date_to:
            conds.append("(pr.punched_at AT TIME ZONE 'Asia/Taipei')::date<=%s"); params.append(date_to)

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT pr.*, ps.name as staff_name, ps.role as staff_role
            FROM punch_records pr JOIN punch_staff ps ON ps.id=pr.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY pr.punched_at DESC LIMIT 500
        """, params).fetchall()
    return jsonify([punch_record_row(r) for r in rows])

@app.route('/api/punch/records', methods=['POST'])
@login_required
def api_punch_record_manual():
    b          = request.get_json(force=True)
    staff_id   = b.get('staff_id')
    punch_type = b.get('punch_type')
    punched_at = b.get('punched_at')
    note       = b.get('note', '').strip()
    manual_by  = b.get('manual_by', '').strip()
    if not all([staff_id, punch_type, punched_at]):
        return jsonify({'error': '缺少必要欄位'}), 400
    if punch_type not in ('in', 'out', 'break_out', 'break_in'):
        return jsonify({'error': '無效的打卡類型'}), 400
    punched_at_parsed = _parse_tw_datetime(punched_at)
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO punch_records
              (staff_id, punch_type, punched_at, note, is_manual, manual_by)
            VALUES (%s,%s,%s,%s,TRUE,%s) RETURNING *
        """, (staff_id, punch_type, punched_at_parsed, note, manual_by)).fetchone()
        staff = conn.execute("SELECT name FROM punch_staff WHERE id=%s", (staff_id,)).fetchone()
        punch_month = str(punched_at_parsed)[:7]
        _trigger_salary_regen_for_leave(conn, int(staff_id), punch_month)
    d = punch_record_row(row)
    if staff: d['staff_name'] = staff['name']
    return jsonify(d), 201

@app.route('/api/punch/records/<int:rid>', methods=['PUT'])
@login_required
def api_punch_record_update(rid):
    b = request.get_json(force=True)
    punch_type = b.get('punch_type')
    if punch_type not in ('in', 'out', 'break_out', 'break_in'):
        return jsonify({'error': '無效的打卡類型'}), 400
    punched_at_parsed = _parse_tw_datetime(b.get('punched_at'))
    if not punched_at_parsed:
        return jsonify({'error': '打卡時間格式錯誤'}), 400
    with get_db() as conn:
        row = conn.execute("""
            UPDATE punch_records
            SET punch_type=%s, punched_at=%s, note=%s, is_manual=TRUE, manual_by=%s
            WHERE id=%s RETURNING *
        """, (punch_type, punched_at_parsed,
              b.get('note', ''), b.get('manual_by', ''), rid)).fetchone()
        if row:
            _trigger_salary_regen_for_leave(conn, row['staff_id'],
                                            str(punched_at_parsed)[:7])
    return jsonify(punch_record_row(row)) if row else ('', 404)

@app.route('/api/punch/records/<int:rid>', methods=['DELETE'])
@login_required
def api_punch_record_delete(rid):
    with get_db() as conn:
        old = conn.execute("SELECT staff_id, punched_at FROM punch_records WHERE id=%s",
                           (rid,)).fetchone()
        conn.execute("DELETE FROM punch_records WHERE id=%s", (rid,))
        if old:
            _trigger_salary_regen_for_leave(conn, old['staff_id'],
                                            str(old['punched_at'])[:7])
    return jsonify({'deleted': rid})

@app.route('/api/punch/summary', methods=['GET'])
@login_required
def api_punch_summary():
    month = request.args.get('month') or _dt.now(TW_TZ).strftime('%Y-%m')
    _ts_s, _ts_e = _month_ts_range(month)
    with get_db() as conn:
        rows = [dict(r) for r in conn.execute("""
            SELECT ps.id as staff_id, ps.name as staff_name,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   MIN(CASE WHEN pr.punch_type='in'  THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as clock_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as clock_out,
                   COUNT(*) as punch_count,
                   BOOL_OR(pr.is_manual) as has_manual
            FROM punch_records pr JOIN punch_staff ps ON ps.id=pr.staff_id
            WHERE pr.punched_at >= %s AND pr.punched_at < %s
            GROUP BY ps.id, ps.name, (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date ASC, ps.name
        """, (_ts_s, _ts_e)).fetchall()]

    # 跨日班次合併:上班在 day N、下班在 day N+1
    from datetime import date as _dcm, timedelta as _tdcm
    _row_map = {(r['staff_id'], str(r['work_date'])): r for r in rows}
    _skip_keys = set()
    for r in rows:
        sid = r['staff_id']
        ds  = str(r['work_date'])
        if r['clock_in'] and not r['clock_out']:
            next_ds = (_dcm.fromisoformat(ds) + _tdcm(days=1)).isoformat()
            nr = _row_map.get((sid, next_ds))
            if nr and nr['clock_out'] and not nr['clock_in']:
                r['clock_out']   = nr['clock_out']
                r['punch_count'] = r['punch_count'] + nr['punch_count']
                _skip_keys.add((sid, next_ds))
    rows = sorted(
        [r for r in rows if (r['staff_id'], str(r['work_date'])) not in _skip_keys],
        key=lambda x: (str(x['work_date']), x['staff_name']),
        reverse=True,
    )

    result = []
    for r in rows:
        d = r
        d['work_date']  = d['work_date'].isoformat()  if d['work_date']  else None
        d['clock_in']   = d['clock_in'].isoformat()   if d['clock_in']   else None
        d['clock_out']  = d['clock_out'].isoformat()  if d['clock_out']  else None
        if d['clock_in'] and d['clock_out']:
            from datetime import datetime as _dt2
            ci = _dt2.fromisoformat(d['clock_in'].replace('Z', ''))
            co = _dt2.fromisoformat(d['clock_out'].replace('Z', ''))
            d['duration_min'] = max(0, int((co - ci).total_seconds() / 60))
        else:
            d['duration_min'] = None
        result.append(d)
    return jsonify(result)


# ── Punch Requests (補打卡申請) ───────────────────────────────────

@app.route('/api/punch/request', methods=['POST'])
def api_punch_req_submit():
    sid = session.get('punch_staff_id')
    if not sid: return jsonify({'error': 'not logged in'}), 401
    b            = request.get_json(force=True)
    punch_type   = b.get('punch_type')
    requested_at = b.get('requested_at')
    reason       = b.get('reason', '').strip()
    if punch_type not in ('in', 'out', 'break_out', 'break_in'):
        return jsonify({'error': '無效的打卡類型'}), 400
    if not requested_at:
        return jsonify({'error': '請選擇補打時間'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO punch_requests (staff_id, punch_type, requested_at, reason)
            VALUES (%s,%s,%s,%s) RETURNING *
        """, (sid, punch_type, requested_at, reason)).fetchone()
    return jsonify(punch_req_row(row)), 201

@app.route('/api/punch/request/my', methods=['GET'])
def api_punch_req_my():
    sid = session.get('punch_staff_id')
    if not sid: return jsonify({'error': 'not logged in'}), 401
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM punch_requests WHERE staff_id=%s ORDER BY requested_at DESC LIMIT 20",
            (sid,)
        ).fetchall()
    return jsonify([punch_req_row(r) for r in rows])

@app.route('/api/punch/requests', methods=['GET'])
@login_required
def api_punch_reqs_list():
    status = request.args.get('status', '')
    conds, params = ['TRUE'], []
    if status: conds.append('pr.status=%s'); params.append(status)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT pr.*, ps.name as staff_name, ps.role as staff_role
            FROM punch_requests pr JOIN punch_staff ps ON ps.id=pr.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY pr.created_at DESC LIMIT 200
        """, params).fetchall()
    return jsonify([punch_req_row(r) for r in rows])

@app.route('/api/punch/requests/<int:rid>', methods=['DELETE'])
@login_required
def api_punch_req_delete(rid):
    with get_db() as conn:
        conn.execute("DELETE FROM punch_requests WHERE id=%s", (rid,))
    return jsonify({'deleted': rid})

# ═══════════════════════════════════════════════════════════════════
# LINE Punch Clock
# ═══════════════════════════════════════════════════════════════════

CUSTOM_RICHMENU_IMAGE_PATH = '/tmp/custom_richmenu.png'
_pending_line_punches = {}   # {line_user_id: punch_type}








@app.route('/line-punch/webhook', methods=['GET', 'POST'])
def line_punch_webhook():
    if request.method == 'GET':
        return 'OK', 200
    cfg = get_line_punch_config()
    if not cfg or not cfg.get('enabled') or not cfg.get('channel_secret'):
        return 'disabled', 200

    signature = request.headers.get('X-Line-Signature', '')
    body      = request.get_data(as_text=True)

    import hmac, hashlib as _hl, base64 as _b64
    secret   = cfg['channel_secret'].encode('utf-8')
    computed = _b64.b64encode(
        hmac.new(secret, body.encode('utf-8'), _hl.sha256).digest()
    ).decode('utf-8')
    if not hmac.compare_digest(computed, signature):
        return 'Invalid signature', 400

    events = _json.loads(body).get('events', [])
    for event in events:
        try:
            _handle_line_punch_event(event, cfg)
        except Exception as e:
            print(f"[LINE PUNCH] event handler error: {e}\n{traceback.format_exc()}")
    return 'OK', 200


def _handle_line_punch_event(event, cfg):
    source   = event.get('source', {})
    user_id  = source.get('userId')
    evt_type = event.get('type')
    if not user_id: return

    msg      = event.get('message', {})
    msg_type = msg.get('type', '')

    if evt_type == 'follow':
        _send_line_punch(user_id,
            '歡迎使用員工打卡系統!👋\n\n'
            '請輸入您的登入帳號完成綁定.\n\n'
            '✏️ 輸入範例:\n  綁定 mary123\n'
            '(請將 mary123 換成您自己的帳號)\n\n'
            '不知道帳號?請詢問管理員.')
        return

    if evt_type != 'message': return

    with get_db() as conn:
        staff = conn.execute(
            "SELECT * FROM punch_staff WHERE line_user_id=%s AND active=TRUE", (user_id,)
        ).fetchone()

    # ── Not bound yet ─────────────────────────────────────────
    if not staff:
        if msg_type == 'text':
            text = msg.get('text', '').strip()
            if text.startswith('綁定 ') or text.startswith('绑定 '):
                username = text.split(' ', 1)[1].strip()
                if username in ('帳號', '您的帳號', '[您的帳號]', 'username', '帳號名稱'):
                    _send_line_punch(user_id,
                        '請輸入您「實際的」登入帳號,而非說明文字.\n\n'
                        '範例:綁定 mary123')
                    return
                with get_db() as conn:
                    candidate = conn.execute(
                        "SELECT * FROM punch_staff WHERE username=%s AND active=TRUE",
                        (username,)
                    ).fetchone()
                if not candidate:
                    _send_line_punch(user_id,
                        f'找不到帳號「{username}」\n\n'
                        '請確認帳號是否正確,或詢問管理員您的登入帳號.')
                    return
                if candidate['line_user_id']:
                    _send_line_punch(user_id, '此帳號已綁定其他 LINE 帳號,請聯絡管理員.')
                    return
                with get_db() as conn:
                    conn.execute(
                        "UPDATE punch_staff SET line_user_id=%s WHERE id=%s",
                        (user_id, candidate['id'])
                    )
                _send_line_punch(user_id,
                    f'✅ 綁定成功!\n歡迎 {candidate["name"]}!\n\n'
                    '打卡方式:\n📍 傳送位置訊息 -> 自動打卡\n'
                    '💬 或輸入:上班 / 下班 / 休息 / 回來\n\n'
                    '輸入「狀態」可查看今日打卡記錄.')
            else:
                _send_line_punch(user_id,
                    '您尚未綁定打卡帳號.\n\n'
                    '請輸入您的登入帳號:\n  綁定 [您的帳號]\n\n'
                    '範例:綁定 mary123')
        return

    # ── Bound staff ───────────────────────────────────────────
    PUNCH_CMDS = {
        '上班': 'in', '上班打卡': 'in',
        '下班': 'out', '下班打卡': 'out',
        '休息': 'break_out', '休息開始': 'break_out',
        '回來': 'break_in', '休息結束': 'break_in',
    }
    PUNCH_LABEL = {
        'in': '上班打卡', 'out': '下班打卡',
        'break_out': '休息開始', 'break_in': '休息結束',
    }

    if msg_type == 'location':
        lat = msg.get('latitude'); lng = msg.get('longitude')
        _do_line_punch(staff, user_id, lat, lng, None, PUNCH_LABEL)

    elif msg_type == 'text':
        text = msg.get('text', '').strip()

        if text in ('狀態', '打卡記錄'):
            _send_status(staff, user_id); return

        if text == '解除綁定':
            with get_db() as conn:
                conn.execute("UPDATE punch_staff SET line_user_id=NULL WHERE id=%s", (staff['id'],))
            _send_line_punch(user_id, '已解除 LINE 帳號綁定.'); return

        punch_type = PUNCH_CMDS.get(text)
        if punch_type:
            with get_db() as conn:
                pcfg = conn.execute("SELECT * FROM punch_config WHERE id=1").fetchone()
                locs = conn.execute("SELECT * FROM punch_locations WHERE active=TRUE").fetchall()
            gps_required = pcfg['gps_required'] if pcfg else False
            if gps_required and locs:
                from linebot.models import QuickReply, QuickReplyButton, LocationAction
                cfg_lp = get_line_punch_config()
                if cfg_lp and cfg_lp.get('enabled') and cfg_lp.get('channel_access_token'):
                    qr = QuickReply(items=[QuickReplyButton(action=LocationAction(label='📍 傳送位置'))])
                    msg = TextSendMessage(
                        text=f'請傳送您的位置來完成{PUNCH_LABEL[punch_type]}\n點下方「傳送位置」按鈕即可打卡',
                        quick_reply=qr)
                    try:
                        LineBotApi(cfg_lp['channel_access_token']).push_message(user_id, msg)
                    except Exception as _e:
                        print(f"[LINE PUNCH] location qr error: {_e}")
                _pending_line_punches[user_id] = punch_type
            else:
                _do_line_punch(staff, user_id, None, None, punch_type, PUNCH_LABEL)
        elif text in ('查餘假', '餘假', '假期', '查假', '特休'):
            _line_query_leave_balance(staff, user_id)
        elif text in ('查薪資', '薪資', '薪水', '薪資單', '查薪水'):
            _line_query_salary(staff, user_id)
        elif text.startswith('請假'):
            _line_submit_leave(staff, user_id, text)
        elif text in ('績效', '考核', '我的考核', '查績效'):
            _line_query_performance(staff, user_id)
        elif text in ('假別', '假別清單', '假別列表'):
            _line_show_leave_types(staff, user_id)
        elif (text in ('出勤紀錄', '出勤記錄', '月出勤', '打卡紀錄', '打卡記錄', '出勤查詢')
              or text.startswith('出勤紀錄 ') or text.startswith('出勤記錄 ')
              or text.startswith('打卡紀錄 ') or text.startswith('打卡記錄 ')):
            _line_query_monthly_records(staff, user_id, text)
        elif text == '加班':
            _line_overtime_start(staff, user_id)
        elif text.startswith('申請加班'):
            _line_submit_overtime(staff, user_id, text)
        elif text in ('選單', '功能', '菜單', '?', '?', 'help', 'Help', 'HELP'):
            _line_show_help(staff, user_id)
        else:
            _line_show_help(staff, user_id)


def _do_line_punch(staff, user_id, lat, lng, forced_type, PUNCH_LABEL):
    from datetime import datetime as _dt3, timezone as _tz3, timedelta as _td3
    TW = _tz3(_td3(hours=8))

    # Determine punch type
    if forced_type:
        punch_type = forced_type
    elif user_id in _pending_line_punches:
        punch_type = _pending_line_punches.pop(user_id)
    else:
        with get_db() as conn:
            last = conn.execute("""
                SELECT punch_type, punched_at FROM punch_records
                WHERE staff_id=%s
                  AND punched_at >= NOW() - INTERVAL '30 hours'
                ORDER BY punched_at DESC LIMIT 1
            """, (staff['id'],)).fetchone()
        if not last:
            punch_type = 'in'
        elif last['punch_type'] == 'in':
            punch_type = 'out'
        elif last['punch_type'] == 'break_out':
            punch_type = 'break_in'
        else:
            # 上次為 out 或 break_in -> 預期下次 in
            # 若剛打卡下班不到 10 分鐘,先確認避免誤操作
            if last['punch_type'] == 'out' and last.get('punched_at'):
                last_time = last['punched_at']
                now_tw = _dt3.now(TW)
                if not last_time.tzinfo:
                    last_time = last_time.replace(tzinfo=TW)
                diff_min = (now_tw - last_time).total_seconds() / 60
                if diff_min < 10:
                    last_str = last_time.strftime('%H:%M')
                    _send_line_with_quick_reply(user_id,
                        f'⚠️ 您剛於 {last_str} 打卡下班({diff_min:.0f} 分鐘前)\n\n確定要打卡上班嗎?',
                        [{'label': '✅ 確認上班打卡', 'text': '上班打卡'},
                         {'label': '❌ 取消',        'text': '狀態'}])
                    return
            punch_type = 'in'

    label = PUNCH_LABEL.get(punch_type, punch_type)

    # GPS check
    gps_distance = None; matched_name = ''
    if lat is not None and lng is not None:
        with get_db() as conn:
            pcfg = conn.execute("SELECT * FROM punch_config WHERE id=1").fetchone()
            locs = conn.execute("SELECT * FROM punch_locations WHERE active=TRUE").fetchall()
        gps_required = pcfg['gps_required'] if pcfg else False
        if locs:
            min_dist = None; min_loc = None
            for loc in locs:
                d = _gps_distance(lat, lng, float(loc['lat']), float(loc['lng']))
                if min_dist is None or d < min_dist:
                    min_dist = d; min_loc = loc
            gps_distance = min_dist
            matched_name = min_loc['location_name'] if min_loc else ''
            if gps_required and min_dist > int(min_loc['radius_m']):
                _send_line_punch(user_id,
                    f'❌ {label}失敗\n'
                    f'您距離「{min_loc["location_name"]}」{min_dist} 公尺\n'
                    f'超出允許範圍 {min_loc["radius_m"]} 公尺\n\n'
                    '請確認您在正確地點後重試.')
                return

    # Duplicate guard
    with get_db() as conn:
        recent = conn.execute("""
            SELECT id FROM punch_records
            WHERE staff_id=%s AND punch_type=%s
              AND punched_at > NOW() - INTERVAL '1 minute'
        """, (staff['id'], punch_type)).fetchone()
        if recent:
            _send_line_punch(user_id, f'⚠️ 1 分鐘內已打過{label},請勿重複打卡.'); return

        conn.execute("""
            INSERT INTO punch_records
              (staff_id, punch_type, latitude, longitude, gps_distance, location_name)
            VALUES (%s,%s,%s,%s,%s,%s)
        """, (staff['id'], punch_type, lat, lng, gps_distance, matched_name))

    now      = _dt3.now(TW)
    gps_info = f'\n📍 {matched_name} ({gps_distance}m)' if gps_distance is not None else ''
    _send_line_punch(user_id,
        f'✅ {label}成功\n'
        f'👤 {staff["name"]}\n'
        f'🕐 {now.strftime("%Y/%m/%d %H:%M")}'
        f'{gps_info}')


def _send_status(staff, user_id):
    from datetime import timezone as _tz4, timedelta as _td4
    TW = _tz4(_td4(hours=8))
    with get_db() as conn:
        rows = conn.execute("""
            SELECT punch_type, punched_at, gps_distance, location_name, is_manual
            FROM punch_records
            WHERE staff_id=%s
              AND (punched_at AT TIME ZONE 'Asia/Taipei')::date
                = (NOW() AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY punched_at ASC
        """, (staff['id'],)).fetchall()
    LABEL = {'in': '上班', 'out': '下班', 'break_out': '休息開始', 'break_in': '休息結束'}
    if not rows:
        _send_line_punch(user_id, f'📋 {staff["name"]} 今日尚無打卡記錄.'); return
    lines = [f'📋 {staff["name"]} 今日打卡記錄']
    for r in rows:
        pa = r['punched_at']
        if pa.tzinfo is None:
            from datetime import timezone as _utz2
            pa = pa.replace(tzinfo=_utz2.utc)
        t    = pa.astimezone(TW).strftime('%H:%M')
        dist = f' ({r["gps_distance"]}m)' if r['gps_distance'] is not None else ''
        man  = ' [補打]' if r['is_manual'] else ''
        lines.append(f'• {LABEL.get(r["punch_type"], r["punch_type"])} {t}{dist}{man}')
    _send_line_punch(user_id, '\n'.join(lines))

# ── Admin LINE Punch Config API ────────────────────────────────────

@app.route('/api/line-punch/config', methods=['GET'])
@login_required
def api_line_punch_config_get():
    with get_db() as conn:
        row = conn.execute("SELECT * FROM line_punch_config WHERE id=1").fetchone()
    if not row:
        return jsonify({'enabled': False, 'channel_access_token': '', 'channel_secret': ''})
    d = dict(row)
    if d.get('updated_at'): d['updated_at'] = d['updated_at'].isoformat()
    return jsonify(d)

@app.route('/api/line-punch/config', methods=['PUT'])
@login_required
def api_line_punch_config_put():
    b       = request.get_json(force=True)
    token   = b.get('channel_access_token', '').strip()
    secret  = b.get('channel_secret', '').strip()
    enabled = bool(b.get('enabled', False))
    with get_db() as conn:
        conn.execute("""
            UPDATE line_punch_config
            SET channel_access_token=%s, channel_secret=%s, enabled=%s, updated_at=NOW()
            WHERE id=1
        """, (token, secret, enabled))
    return jsonify({'ok': True, 'enabled': enabled})

@app.route('/api/line-punch/staff', methods=['GET'])
@login_required
def api_line_punch_staff():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id,name,username,role,active,line_user_id FROM punch_staff ORDER BY name"
        ).fetchall()
    return jsonify([{
        'id': r['id'], 'name': r['name'], 'username': r['username'],
        'role': r['role'], 'active': r['active'],
        'line_bound': bool(r['line_user_id']),
        'line_user_id': r['line_user_id'] or ''
    } for r in rows])

@app.route('/api/line-punch/staff/<int:sid>/unbind', methods=['POST'])
@login_required
def api_line_punch_unbind(sid):
    with get_db() as conn:
        conn.execute("UPDATE punch_staff SET line_user_id=NULL WHERE id=%s", (sid,))
    return jsonify({'ok': True})

# ── Rich Menu ──────────────────────────────────────────────────────



def _make_richmenu_png():
    """Generate a simple 2500×1686 PNG with 4 colored quadrants."""
    import struct, zlib
    W, H = 2500, 1686
    colors = [(0x2e,0x9e,0x6b), (0xd6,0x42,0x42), (0xe0,0x7b,0x2a), (0x4a,0x7b,0xda)]
    rows = []
    for y in range(H):
        row = bytearray()
        for x in range(W):
            p = (0 if y < 843 else 1) * 2 + (0 if x < 1250 else 1)
            r, g, b = colors[p]
            if x in (1249, 1250) or y in (842, 843):
                r, g, b = 0x0f, 0x1c, 0x3a
            row += bytes([r, g, b])
        rows.append(bytes([0]) + bytes(row))
    compressed = zlib.compress(b''.join(rows), 1)

    def chunk(name, data):
        c = struct.pack('>I', len(data)) + name + data
        return c + struct.pack('>I', zlib.crc32(c[4:]) & 0xffffffff)

    return (b'\x89PNG\r\n\x1a\n'
            + chunk(b'IHDR', struct.pack('>IIBBBBB', W, H, 8, 2, 0, 0, 0))
            + chunk(b'IDAT', compressed)
            + chunk(b'IEND', b''))


@app.route('/api/line-punch/richmenu/create', methods=['POST'])
@login_required
def api_richmenu_create():
    cfg = get_line_punch_config()
    if not cfg or not cfg.get('channel_access_token'):
        return jsonify({'error': '請先設定 Channel Access Token'}), 400

    b = request.get_json(force=True) or {}
    gdrive_url = b.get('gdrive_url', '').strip()
    btn_texts  = b.get('button_texts') or []
    defaults   = ['上班', '下班', '請假', '加班']
    btn_texts  = [(btn_texts[i].strip() if i < len(btn_texts) and btn_texts[i].strip() else defaults[i]) for i in range(4)]

    body = {
        "size": {"width": 2500, "height": 1686},
        "selected": True,
        "name": "打卡選單",
        "chatBarText": "打卡",
        "areas": [
            {"bounds": {"x": 0,    "y": 0,   "width": 1250, "height": 843}, "action": {"type": "message", "text": btn_texts[0]}},
            {"bounds": {"x": 1250, "y": 0,   "width": 1250, "height": 843}, "action": {"type": "message", "text": btn_texts[1]}},
            {"bounds": {"x": 0,    "y": 843, "width": 1250, "height": 843}, "action": {"type": "message", "text": btn_texts[2]}},
            {"bounds": {"x": 1250, "y": 843, "width": 1250, "height": 843}, "action": {"type": "message", "text": btn_texts[3]}},
        ]
    }

    status, data = _call_line_api(cfg, 'POST', '/richmenu', body)
    if status != 200:
        return jsonify({'error': f'建立失敗 ({status}): {data.get("error","")}'}), 500

    rich_menu_id = data.get('richMenuId', '')

    # Upload image - 1) Google Drive  2) custom local file  3) auto-generate
    png_bytes = None

    if gdrive_url:
        try:
            import re as _re
            file_id = None
            m = _re.search(r'/file/d/([^/?]+)', gdrive_url)
            if m:
                file_id = m.group(1)
            else:
                m = _re.search(r'[?&]id=([^&]+)', gdrive_url)
                if m:
                    file_id = m.group(1)
            if file_id:
                dl_url = f'https://drive.google.com/uc?export=download&id={file_id}'
                req = urllib.request.Request(dl_url, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=30) as resp:
                    png_bytes = resp.read()
                # Google Drive may return an HTML warning page for large files
                if png_bytes and png_bytes[:1] not in (b'\x89', b'\xff', b'\x47', b'BM'):
                    print(f"[LINE PUNCH] gdrive returned non-image content, ignoring")
                    png_bytes = None
        except Exception as e:
            print(f"[LINE PUNCH] gdrive download error: {e}")

    if not png_bytes:
        try:
            import os
            for _cp in [CUSTOM_RICHMENU_IMAGE_PATH,
                        CUSTOM_RICHMENU_IMAGE_PATH.replace('.png', '.jpg')]:
                if os.path.exists(_cp):
                    with open(_cp, 'rb') as f:
                        png_bytes = f.read()
                    break
        except Exception:
            pass

    if not png_bytes:
        try:
            png_bytes = _make_richmenu_png()
        except Exception:
            pass

    img_ok = False
    if png_bytes:
        content_type = 'image/jpeg' if png_bytes[:2] == b'\xff\xd8' else 'image/png'
        upload_url = f'https://api-data.line.me/v2/bot/richmenu/{rich_menu_id}/content'
        req = urllib.request.Request(
            upload_url, data=png_bytes, method='POST',
            headers={'Content-Type': content_type, 'Authorization': f'Bearer {cfg["channel_access_token"]}'}
        )
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                img_ok = resp.status in (200, 204)
        except Exception:
            pass

    # Set as default for all users
    _call_line_api(cfg, 'POST', f'/user/all/richmenu/{rich_menu_id}')

    return jsonify({'ok': True, 'rich_menu_id': rich_menu_id, 'image_uploaded': img_ok})


@app.route('/api/line-punch/richmenu/list', methods=['GET'])
@login_required
def api_richmenu_list():
    cfg = get_line_punch_config()
    if not cfg or not cfg.get('channel_access_token'):
        return jsonify({'menus': []})
    status, data = _call_line_api(cfg, 'GET', '/richmenu/list')
    if status != 200:
        return jsonify({'menus': [], 'error': data.get('error', '')})
    return jsonify({'menus': data.get('richmenus', [])})


@app.route('/api/line-punch/richmenu/<rich_menu_id>', methods=['DELETE'])
@login_required
def api_richmenu_delete(rich_menu_id):
    cfg = get_line_punch_config()
    if not cfg or not cfg.get('channel_access_token'):
        return jsonify({'error': '未設定 Token'}), 400
    _call_line_api(cfg, 'DELETE', '/user/all/richmenu')
    status, _ = _call_line_api(cfg, 'DELETE', f'/richmenu/{rich_menu_id}')
    return jsonify({'ok': status in (200, 204), 'status': status})


@app.route('/api/line-punch/richmenu/default', methods=['DELETE'])
@login_required
def api_richmenu_unset_default():
    cfg = get_line_punch_config()
    if not cfg or not cfg.get('channel_access_token'):
        return jsonify({'error': '未設定 Token'}), 400
    status, _ = _call_line_api(cfg, 'DELETE', '/user/all/richmenu')
    return jsonify({'ok': status in (200, 204)})

# ═══════════════════════════════════════════════════════════════════
# Schedule / Shift API
# ═══════════════════════════════════════════════════════════════════

# ── Employee: schedule config + my request ────────────────────────






# ── Admin: schedule config ────────────────────────────────────────




# ── Admin: schedule requests ──────────────────────────────────────










# ── Shift Types CRUD ──────────────────────────────────────────────






# ── Shift Assignments ─────────────────────────────────────────────














# ── Overtime Requests ─────────────────────────────────────────────

















# ═══════════════════════════════════════════════════════════════════
# Leave Management (請假管理)
# 2026 勞基法:
#   特休:到職1年10天、2年15天、3~5年每年+1、滿5年20天(上限)
#   病假:每年30天(半薪),超過住院病假 365 天內 30 天(全薪)
#   事假:每年14天(無薪)
#   生理假:每月1天(含病假計算,前3天半薪)
#   婚假:8天全薪
#   喪假:父母/配偶/子女8天;祖父母/孫子女/兄弟姐妹6天;曾祖父母3天
#   產假:8週全薪;陪產假:7天全薪
#   公假:全薪
# ═══════════════════════════════════════════════════════════════════

# ── Leave Tables ─────────────────────────────────────────────────────────────

def init_leave_db():
    migrations = [
        """CREATE TABLE IF NOT EXISTS leave_types (
            id          SERIAL PRIMARY KEY,
            name        TEXT NOT NULL UNIQUE,
            code        TEXT NOT NULL UNIQUE,
            pay_rate    NUMERIC(4,2) DEFAULT 1.0,
            max_days    NUMERIC(5,1),
            description TEXT DEFAULT '',
            color       TEXT DEFAULT '#4a7bda',
            active      BOOLEAN DEFAULT TRUE,
            sort_order  INT DEFAULT 0,
            created_at  TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS leave_requests (
            id              SERIAL PRIMARY KEY,
            staff_id        INT REFERENCES punch_staff(id) ON DELETE CASCADE,
            leave_type_id   INT REFERENCES leave_types(id),
            start_date      DATE NOT NULL,
            end_date        DATE NOT NULL,
            start_half      BOOLEAN DEFAULT FALSE,
            end_half        BOOLEAN DEFAULT FALSE,
            total_days      NUMERIC(5,1) NOT NULL DEFAULT 0,
            reason          TEXT DEFAULT '',
            status          TEXT DEFAULT 'pending',
            reviewed_by     TEXT DEFAULT '',
            review_note     TEXT DEFAULT '',
            reviewed_at     TIMESTAMPTZ,
            substitute_name TEXT DEFAULT '',
            created_at      TIMESTAMPTZ DEFAULT NOW(),
            updated_at      TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS leave_balances (
            id          SERIAL PRIMARY KEY,
            staff_id    INT REFERENCES punch_staff(id) ON DELETE CASCADE,
            leave_type_id INT REFERENCES leave_types(id),
            year        INT NOT NULL,
            total_days  NUMERIC(5,1) DEFAULT 0,
            used_days   NUMERIC(5,1) DEFAULT 0,
            note        TEXT DEFAULT '',
            updated_at  TIMESTAMPTZ DEFAULT NOW(),
            UNIQUE(staff_id, leave_type_id, year)
        )""",
    ]
    for sql in migrations:
        try:
            with get_db() as conn:
                conn.execute(sql)
        except Exception as e:
            print(f"[leave_init] {str(e)[:80]}")

    # Seed default leave types
    defaults = [
        ('特休假',   'annual',      1.0,  30,  '#2e9e6b', 1),
        ('病假',     'sick',        0.5,  30,  '#e07b2a', 2),
        ('住院病假', 'hospitalize', 1.0,  30,  '#d64242', 3),
        ('事假',     'personal',    0.0,  14,  '#8892a4', 4),
        ('生理假',   'menstrual',   0.5,  12,  '#c45cb8', 5),
        ('婚假',     'marriage',    1.0,   8,  '#c8a96e', 6),
        ('喪假',     'funeral',     1.0,   8,  '#4a7bda', 7),
        ('產假',     'maternity',   1.0,  56,  '#e05c8a', 8),
        ('陪產假',   'paternity',   1.0,   7,  '#5cb8c4', 9),
        ('公假',     'official',    1.0, None, '#243d6e', 10),
        ('補休',     'compensatory',1.0, None, '#8b5cf6', 11),
    ]
    try:
        with get_db() as conn:
            cnt = conn.execute("SELECT COUNT(*) as c FROM leave_types").fetchone()['c']
            if cnt == 0:
                for name, code, pay, maxd, color, sort in defaults:
                    conn.execute(
                        "INSERT INTO leave_types (name,code,pay_rate,max_days,color,sort_order) VALUES (%s,%s,%s,%s,%s,%s)",
                        (name, code, pay, maxd, color, sort)
                    )
    except Exception as e:
        print(f"[leave_seed] {e}")

init_leave_db()










# ── Leave Type CRUD ──────────────────────────────────────────────






# ── Leave Requests ────────────────────────────────────────────────







# ── Employee: submit leave request ────────────────────────────────



# ── Leave Balance ─────────────────────────────────────────────────









# ── Leave Summary (for salary integration) ───────────────────────


# ═══════════════════════════════════════════════════════════════════
# Salary Management (薪資管理)
# 2026 勞基法:
#   勞保費率 10.5%(員工負擔 20%=2.1%,含就業保險)
#   健保費率 5.17%(員工負擔 30%=1.551%)
#   勞退提撥 6%(雇主強制提撥,員工自願另計)
#   最低工資 2026年 NT$28,590(月薪)
# ═══════════════════════════════════════════════════════════════════

def init_salary_db():
    migrations = [
        """CREATE TABLE IF NOT EXISTS salary_items (
            id          SERIAL PRIMARY KEY,
            name        TEXT NOT NULL,
            item_type   TEXT NOT NULL DEFAULT 'allowance',
            formula     TEXT DEFAULT '',
            amount      NUMERIC(12,2) DEFAULT 0,
            description TEXT DEFAULT '',
            color       TEXT DEFAULT '#4a7bda',
            active      BOOLEAN DEFAULT TRUE,
            sort_order  INT DEFAULT 0,
            created_at  TIMESTAMPTZ DEFAULT NOW()
        )""",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS salary_item_ids JSONB DEFAULT NULL",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS salary_item_overrides JSONB DEFAULT NULL",
        "ALTER TABLE salary_records ADD COLUMN IF NOT EXISTS income_tax_withheld NUMERIC(12,2) DEFAULT 0",
        """CREATE TABLE IF NOT EXISTS salary_records (
            id              SERIAL PRIMARY KEY,
            staff_id        INT REFERENCES punch_staff(id) ON DELETE CASCADE,
            month           TEXT NOT NULL,
            base_salary     NUMERIC(12,2) DEFAULT 0,
            insured_salary  NUMERIC(12,2) DEFAULT 0,
            work_days       NUMERIC(5,1)  DEFAULT 0,
            actual_days     NUMERIC(5,1)  DEFAULT 0,
            leave_days      NUMERIC(5,1)  DEFAULT 0,
            unpaid_days     NUMERIC(5,1)  DEFAULT 0,
            ot_pay          NUMERIC(12,2) DEFAULT 0,
            allowance_total NUMERIC(12,2) DEFAULT 0,
            deduction_total NUMERIC(12,2) DEFAULT 0,
            net_pay         NUMERIC(12,2) DEFAULT 0,
            items           JSONB         DEFAULT '[]',
            status          TEXT          DEFAULT 'draft',
            note            TEXT          DEFAULT '',
            confirmed_by    TEXT          DEFAULT '',
            confirmed_at    TIMESTAMPTZ,
            created_at      TIMESTAMPTZ   DEFAULT NOW(),
            updated_at      TIMESTAMPTZ   DEFAULT NOW(),
            UNIQUE(staff_id, month)
        )""",
    ]
    for sql in migrations:
        try:
            with get_db() as conn:
                conn.execute(sql)
        except Exception as e:
            print(f"[salary_init] {str(e)[:80]}")

    # Seed default salary items
    defaults = [
        ('本薪',        'allowance', 'base_salary+service_years*1000', 0,    '#2e9e6b', 1),
        ('職務加給',    'allowance', '',                                0,    '#0ea5e9', 2),
        ('全勤獎金',    'allowance', '',                                0,    '#c8a96e', 3),
        ('獎金',        'allowance', '',                                0,    '#8b5cf6', 4),
        ('生日禮金',    'allowance', '',                                1000, '#e05c8a', 5),
        ('勞退6%',      'allowance', 'base_salary*0.06+service_years*1000*0.06', 0, '#4a7bda', 6),
        ('病/事/假',    'deduction', '',                                0,    '#8892a4', 7),
        ('勞保費',      'deduction', 'insured_salary*0.125*0.2',       0,    '#d64242', 8),
        ('健保費',      'deduction', 'insured_salary*0.0517*0.3',      0,    '#e07b2a', 9),
        ('勞退提撥6%',  'deduction', 'base_salary*0.06+service_years*1000*0.06', 0, '#4a7bda', 10),
    ]
    try:
        with get_db() as conn:
            cnt = conn.execute("SELECT COUNT(*) as c FROM salary_items").fetchone()['c']
            if cnt == 0:
                for name, itype, formula, amount, color, sort in defaults:
                    conn.execute("""
                        INSERT INTO salary_items (name,item_type,formula,amount,color,sort_order)
                        VALUES (%s,%s,%s,%s,%s,%s)
                    """, (name, itype, formula, amount, color, sort))
    except Exception as e:
        print(f"[salary_seed] {e}")

init_salary_db()






# ── Employee: view own payslip ────────────────────────────────────


# ── Salary Items CRUD ─────────────────────────────────────────────





# ── Salary Records ─────────────────────────────────────────────────








# ── Salary Staff Settings ─────────────────────────────────────────




# ═══════════════════════════════════════════════════════════════════
# Announcement Module (公告管理)
# ═══════════════════════════════════════════════════════════════════

def init_announcement_db():
    try:
        with get_db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS announcements (
                    id          SERIAL PRIMARY KEY,
                    title       TEXT NOT NULL,
                    content     TEXT NOT NULL,
                    category    TEXT DEFAULT 'general',
                    priority    TEXT DEFAULT 'normal',
                    is_pinned   BOOLEAN DEFAULT FALSE,
                    visible_to  TEXT DEFAULT 'all',
                    published_at TIMESTAMPTZ DEFAULT NOW(),
                    expires_at  TIMESTAMPTZ,
                    author      TEXT DEFAULT '管理員',
                    active      BOOLEAN DEFAULT TRUE,
                    view_count  INT DEFAULT 0,
                    created_at  TIMESTAMPTZ DEFAULT NOW(),
                    updated_at  TIMESTAMPTZ DEFAULT NOW()
                )
            """)
    except Exception as e:
        print(f"[announcement_init] {e}")

init_announcement_db()


# ── Admin: CRUD ───────────────────────────────────────────────────






# ── Public: employee reads ────────────────────────────────────────




# ═══════════════════════════════════════════════════════════════════
# Feature 1: Taiwan Public Holidays (國定假日)
# ═══════════════════════════════════════════════════════════════════

def init_holiday_db():
    try:
        with get_db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS public_holidays (
                    id          SERIAL PRIMARY KEY,
                    date        DATE NOT NULL UNIQUE,
                    name        TEXT NOT NULL,
                    holiday_type TEXT DEFAULT 'national',
                    note        TEXT DEFAULT '',
                    created_at  TIMESTAMPTZ DEFAULT NOW()
                )
            """)
        # Seed 2025 & 2026 Taiwan holidays
        _seed_holidays()
    except Exception as e:
        print(f"[holiday_init] {e}")

def _seed_holidays():
    """台灣2025-2026國定假日"""
    holidays_2025 = [
        ('2025-01-01', '元旦'),
        ('2025-01-27', '農曆除夕'),
        ('2025-01-28', '春節'),
        ('2025-01-29', '春節'),
        ('2025-01-30', '春節'),
        ('2025-01-31', '春節補假'),
        ('2025-02-28', '和平紀念日'),
        ('2025-04-03', '兒童節補假'),
        ('2025-04-04', '兒童節/清明節'),
        ('2025-05-01', '勞動節'),
        ('2025-05-30', '端午節補假'),
        ('2025-06-02', '端午節'),
        ('2025-10-06', '中秋節補假'),
        ('2025-10-07', '中秋節'),
        ('2025-10-10', '國慶日'),
    ]
    holidays_2026 = [
        ('2026-01-01', '元旦'),
        ('2026-01-28', '農曆除夕'),
        ('2026-01-29', '春節'),
        ('2026-01-30', '春節'),
        ('2026-01-31', '春節'),
        ('2026-02-02', '春節補假'),
        ('2026-02-28', '和平紀念日'),
        ('2026-03-02', '和平紀念日補假'),
        ('2026-04-03', '兒童節'),
        ('2026-04-04', '清明節'),
        ('2026-04-05', '清明節補假'),
        ('2026-05-01', '勞動節'),
        ('2026-06-19', '端午節'),
        ('2026-09-25', '中秋節'),
        ('2026-10-09', '國慶日補假'),
        ('2026-10-10', '國慶日'),
    ]
    all_holidays = holidays_2025 + holidays_2026
    try:
        with get_db() as conn:
            existing = conn.execute("SELECT COUNT(*) as c FROM public_holidays").fetchone()['c']
            if existing == 0:
                for date_str, name in all_holidays:
                    try:
                        conn.execute(
                            "INSERT INTO public_holidays (date, name) VALUES (%s,%s) ON CONFLICT (date) DO NOTHING",
                            (date_str, name)
                        )
                    except Exception:
                        pass
    except Exception as e:
        print(f"[holiday_seed] {e}")

init_holiday_db()



# ── Holiday CRUD API ─────────────────────────────────────────────







# ═══════════════════════════════════════════════════════════════════
# Feature 2: LINE Notification Helper
# ═══════════════════════════════════════════════════════════════════





# ═══════════════════════════════════════════════════════════════════
# Feature 3: Export Reports (出勤/薪資報表匯出)
# ═══════════════════════════════════════════════════════════════════

import csv
import io

@app.route('/api/export/attendance', methods=['GET'])
@login_required
def api_export_attendance():
    """匯出月度出勤明細 CSV"""
    month    = request.args.get('month', '')
    staff_id = request.args.get('staff_id', '')
    if not month:
        from datetime import date as _de
        month = _de.today().strftime('%Y-%m')

    _ts_s, _ts_e = _month_ts_range(month)
    conds, params = ["pr.punched_at >= %s AND pr.punched_at < %s"], [_ts_s, _ts_e]
    if staff_id:
        conds.append("pr.staff_id=%s"); params.append(int(staff_id))

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT
                ps.employee_code,
                ps.name as staff_name,
                ps.department,
                ps.role,
                (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                pr.punch_type,
                to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei', 'HH24:MI') as punch_time,
                pr.is_manual,
                pr.manual_by,
                pr.gps_distance,
                pr.location_name,
                pr.note
            FROM punch_records pr
            JOIN punch_staff ps ON ps.id = pr.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY ps.name, pr.punched_at
        """, params).fetchall()

    PUNCH_LABEL = {'in':'上班打卡','out':'下班打卡','break_out':'休息開始','break_in':'休息結束'}

    output = io.StringIO()
    output.write('\ufeff')  # UTF-8 BOM for Excel
    writer = csv.writer(output)
    writer.writerow(['員工代碼','姓名','部門','職稱','日期','打卡類型','時間','補打卡','操作人','GPS距離(m)','地點','備註'])

    for r in rows:
        writer.writerow([
            r['employee_code'] or '',
            r['staff_name'],
            r['department']    or '',
            r['role']          or '',
            str(r['work_date']),
            PUNCH_LABEL.get(r['punch_type'], r['punch_type']),
            r['punch_time'],
            '是' if r['is_manual'] else '',
            r['manual_by']     or '',
            r['gps_distance']  if r['gps_distance'] is not None else '',
            r['location_name'] or '',
            r['note']          or '',
        ])

    csv_content = output.getvalue()
    from flask import Response
    return Response(
        csv_content.encode('utf-8-sig'),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename=attendance_{month}.csv'}
    )


@app.route('/api/export/attendance-summary', methods=['GET'])
@login_required
def api_export_attendance_summary():
    """匯出月度出勤摘要 CSV(每人每天工時)"""
    month = request.args.get('month', '')
    if not month:
        from datetime import date as _df
        month = _df.today().strftime('%Y-%m')

    _ts_s, _ts_e = _month_ts_range(month)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT
                ps.employee_code,
                ps.name,
                ps.department,
                ps.role,
                (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                MIN(CASE WHEN pr.punch_type='in'  THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_in,
                MAX(CASE WHEN pr.punch_type='out' THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_out,
                MIN(CASE WHEN pr.punch_type='in'  THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as ci_ts,
                MAX(CASE WHEN pr.punch_type='out' THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as co_ts,
                BOOL_OR(pr.is_manual) as has_manual,
                COUNT(*) as punch_count
            FROM punch_records pr
            JOIN punch_staff ps ON ps.id = pr.staff_id
            WHERE pr.punched_at >= %s AND pr.punched_at < %s
            GROUP BY ps.employee_code, ps.name, ps.department, ps.role,
                     (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY ps.name, (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
        """, (_ts_s, _ts_e)).fetchall()

    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output)
    writer.writerow(['員工代碼','姓名','部門','職稱','日期','上班','下班','工時(h)','打卡次數','含補打'])

    for r in rows:
        dur_h = ''
        if r['ci_ts'] and r['co_ts']:
            from datetime import datetime as _dtx
            try:
                ci = r['ci_ts'] if hasattr(r['ci_ts'], 'timestamp') else _dtx.fromisoformat(str(r['ci_ts']))
                co = r['co_ts'] if hasattr(r['co_ts'], 'timestamp') else _dtx.fromisoformat(str(r['co_ts']))
                dur_h = round((co - ci).total_seconds() / 3600, 2)
            except Exception:
                pass
        writer.writerow([
            r['employee_code'] or '',
            r['name'], r['department'] or '', r['role'] or '',
            str(r['work_date']),
            r['clock_in'] or '', r['clock_out'] or '',
            dur_h,
            r['punch_count'],
            '是' if r['has_manual'] else '',
        ])

    from flask import Response
    return Response(
        output.getvalue().encode('utf-8-sig'),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename=attendance_summary_{month}.csv'}
    )




@app.route('/api/export/salary', methods=['GET'])
@login_required
def api_export_salary():
    """匯出月度薪資明細 CSV"""
    month = request.args.get('month', '')
    if not month:
        from datetime import date as _dg
        month = _dg.today().strftime('%Y-%m')

    with get_db() as conn:
        rows = conn.execute("""
            SELECT sr.*, ps.name as staff_name, ps.employee_code,
                   ps.department, ps.role, ps.salary_type
            FROM salary_records sr
            JOIN punch_staff ps ON ps.id = sr.staff_id
            WHERE sr.month = %s
            ORDER BY ps.name
        """, (month,)).fetchall()

    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output)
    writer.writerow([
        '員工代碼','姓名','部門','職稱','薪資制度',
        '工作日','出勤天數','請假天數','無薪假天數',
        '津貼合計','扣除合計','加班費','實領金額','狀態','備註'
    ])

    for r in rows:
        items = r['items'] if isinstance(r['items'], list) else _json.loads(r['items'] or '[]')
        sal_type = r['salary_type'] or 'monthly'
        writer.writerow([
            r['employee_code'] or '', r['staff_name'],
            r['department'] or '', r['role'] or '',
            '時薪制' if sal_type == 'hourly' else '月薪制',
            float(r['work_days'] or 0), float(r['actual_days'] or 0),
            float(r['leave_days'] or 0), float(r['unpaid_days'] or 0),
            float(r['allowance_total'] or 0), float(r['deduction_total'] or 0),
            float(r['ot_pay'] or 0), float(r['net_pay'] or 0),
            '已確認' if r['status'] == 'confirmed' else '草稿',
            r['note'] or '',
        ])

    from flask import Response
    return Response(
        output.getvalue().encode('utf-8-sig'),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename=salary_{month}.csv'}
    )


@app.route('/api/export/leave', methods=['GET'])
@login_required
def api_export_leave():
    """匯出請假記錄 CSV"""
    month    = request.args.get('month', '')
    year     = request.args.get('year',  '')
    staff_id = request.args.get('staff_id', '')

    conds, params = ['lr.status=%s'], ['approved']
    if month:
        _d_s, _d_e = _month_date_range(month)
        conds.append('lr.start_date >= %s AND lr.start_date < %s'); params.extend([_d_s, _d_e])
    if year:  conds.append("EXTRACT(YEAR FROM lr.start_date)=%s"); params.append(int(year))
    if staff_id: conds.append("lr.staff_id=%s"); params.append(int(staff_id))

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT lr.*, ps.name as staff_name, ps.employee_code,
                   ps.department, lt.name as leave_type_name, lt.pay_rate
            FROM leave_requests lr
            JOIN punch_staff ps ON ps.id = lr.staff_id
            JOIN leave_types  lt ON lt.id = lr.leave_type_id
            WHERE {' AND '.join(conds)}
            ORDER BY lr.start_date, ps.name
        """, params).fetchall()

    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output)
    writer.writerow(['員工代碼','姓名','部門','假別','薪資倍率','開始日期','結束日期','天數','原因','代理人','狀態'])

    PAY_LABEL = {1.0:'全薪', 0.5:'半薪', 0.0:'無薪'}
    for r in rows:
        writer.writerow([
            r['employee_code'] or '', r['staff_name'], r['department'] or '',
            r['leave_type_name'], PAY_LABEL.get(float(r['pay_rate']), f"{r['pay_rate']}倍"),
            str(r['start_date']), str(r['end_date']),
            float(r['total_days']),
            r['reason'] or '', r['substitute_name'] or '',
            {'approved':'已核准','rejected':'已退回','pending':'待審核'}.get(r['status'], r['status']),
        ])

    from flask import Response
    return Response(
        output.getvalue().encode('utf-8-sig'),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename=leave_{month or year or "all"}.csv'}
    )


# ── Patch existing review functions with LINE notifications ──────

def _patch_reviews_with_notifications():
    """
    This is called after all route functions are defined.
    We monkey-patch the review endpoints to send LINE notifications.
    The actual patching is done inline in the route handlers below
    via the _notify_review_result helper.
    """
    pass

@app.route('/api/punch/requests/<int:rid>', methods=['PUT'])
@login_required
def api_punch_req_review_v2(rid):
    b           = request.get_json(force=True)
    action      = b.get('action')
    reviewed_by = b.get('reviewed_by', '').strip()
    review_note = b.get('review_note', '').strip()
    if action not in ('approve', 'reject'):
        return jsonify({'error': 'invalid action'}), 400
    new_status = 'approved' if action == 'approve' else 'rejected'
    with get_db() as conn:
        row = conn.execute("""
            UPDATE punch_requests
            SET status=%s, reviewed_by=%s, review_note=%s, reviewed_at=NOW()
            WHERE id=%s
            RETURNING *, (SELECT name FROM punch_staff WHERE id=staff_id) as staff_name
        """, (new_status, reviewed_by, review_note, rid)).fetchone()
        if not row: return ('', 404)
        if action == 'approve':
            conn.execute("""
                INSERT INTO punch_records
                  (staff_id, punch_type, punched_at, note, is_manual, manual_by)
                VALUES (%s,%s,%s,%s,TRUE,%s)
            """, (row['staff_id'], row['punch_type'], row['requested_at'],
                  f'補打卡申請 #{rid}:{row["reason"]}', reviewed_by))
            punch_month = str(row['requested_at'])[:7]
            _trigger_salary_regen_for_leave(conn, row['staff_id'], punch_month)
    # LINE notification
    LABEL = {'in':'上班打卡','out':'下班打卡','break_out':'休息開始','break_in':'休息結束'}
    dt_str = row['requested_at'].isoformat()[:16].replace('T',' ')
    extra  = f"{LABEL.get(row['punch_type'],'')} {dt_str}"
    if review_note: extra += f"\n審核意見:{review_note}"
    _notify_review_result(row['staff_id'], '補打卡申請', action, extra)
    return jsonify(punch_req_row(row))


# ═══════════════════════════════════════════════════════════════════
# Dashboard API
# ═══════════════════════════════════════════════════════════════════



# ── Dashboard 擴充 API ────────────────────────────────────────────────────────







# ── 年度扣繳憑單 ────────────────────────────────────────────────────────────

@app.route('/api/export/withholding', methods=['GET'])
@require_module('salary')
def api_export_withholding():
    """年度薪資所得扣繳憑單(所得類別50)"""
    from datetime import date as _dwh
    year   = request.args.get('year', str(_dwh.today().year))
    fmt    = request.args.get('format', 'html')

    fs = _get_finance_settings()
    company_name   = fs.get('company_name', '')
    company_tax_id = fs.get('company_tax_id', '')
    company_address= fs.get('company_address', '')

    with get_db() as conn:
        rows = conn.execute("""
            SELECT ps.id, ps.name, ps.national_id, ps.address,
                   COALESCE(SUM(sr.allowance_total), 0)       AS gross_salary,
                   COALESCE(SUM(sr.income_tax_withheld), 0)   AS tax_withheld,
                   COALESCE(AVG(sr.insured_salary), 0)        AS avg_insured
            FROM salary_records sr
            JOIN punch_staff ps ON ps.id = sr.staff_id
            WHERE sr.month LIKE %s AND sr.status='confirmed'
            GROUP BY ps.id, ps.name, ps.national_id, ps.address
            ORDER BY ps.name
        """, (f'{year}-%',)).fetchall()

    # 計算二代健保補充費
    def supp_nhi(gross, insured):
        base = float(gross) - float(insured) * 12
        return max(0, round(base * 0.0211, 0)) if base > 0 else 0

    data = []
    for i, r in enumerate(rows, 1):
        gross = float(r['gross_salary'])
        insured = float(r['avg_insured'])
        data.append({
            'no':          i,
            'name':        r['name'],
            'national_id': r['national_id'] or '-',
            'address':     r['address'] or '-',
            'gross':       gross,
            'supp_nhi':    supp_nhi(gross, insured),
            'tax':         float(r['tax_withheld']),
        })

    if fmt == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO
        wb = wb2 = openpyxl.Workbook()
        ws = wb.active; ws.title = f'{year}年扣繳憑單'
        hfill = PatternFill('solid', fgColor='0F1C3A')
        thin  = Border(*[Side(style='thin', color='DDDDDD')]*4)
        hdrs  = ['序號','姓名','身分證字號','地址','年度薪資合計','二代健保補充費','扣繳稅額']
        ws.append(hdrs)
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(1, ci); c.font = Font(bold=True, color='FFFFFF', size=10); c.fill = hfill
            c.alignment = Alignment(horizontal='center', vertical='center'); c.border = thin
        ws.column_dimensions['A'].width = 5; ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 14; ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 16; ws.column_dimensions['F'].width = 16; ws.column_dimensions['G'].width = 12
        for d in data:
            ws.append([d['no'], d['name'], d['national_id'], d['address'],
                       d['gross'], d['supp_nhi'], d['tax']])
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        from flask import Response as _FR2
        return _FR2(buf.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=withholding_{year}.xlsx'})

    # HTML printable
    rows_html = ''.join(f"""
      <tr>
        <td style="text-align:center">{d['no']}</td>
        <td>{d['name']}</td>
        <td style="font-family:monospace">{d['national_id']}</td>
        <td style="font-size:11px">{d['address']}</td>
        <td style="text-align:right;font-family:monospace">{d['gross']:,.0f}</td>
        <td style="text-align:right;font-family:monospace">{d['supp_nhi']:,.0f}</td>
        <td style="text-align:right;font-family:monospace">{d['tax']:,.0f}</td>
      </tr>""" for d in data)
    html = f"""<!DOCTYPE html><html lang="zh-TW"><head>
<meta charset="UTF-8"><title>{year}年度薪資扣繳憑單</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Noto Sans TC',sans-serif;font-size:12px;padding:20px;color:#1e2a45}}
h2{{font-size:16px;font-weight:700;margin-bottom:4px}}
.meta{{font-size:11px;color:#666;margin-bottom:16px}}
table{{width:100%;border-collapse:collapse;margin-bottom:20px}}
th{{background:#0f1c3a;color:#fff;padding:7px 10px;font-size:11px;font-weight:600;text-align:left}}
td{{padding:6px 10px;border-bottom:1px solid #eee;font-size:12px}}
tr:nth-child(even){{background:#f8f9fb}}
.note{{font-size:10px;color:#888;border-top:1px solid #ddd;padding-top:8px}}
@media print{{button{{display:none}}}}
</style></head><body>
<button onclick="window.print()" style="margin-bottom:16px;padding:6px 16px;background:#0f1c3a;color:#fff;border:none;border-radius:4px;cursor:pointer;font-size:12px">列印</button>
<h2>{year} 年度薪資所得扣繳憑單(所得類別 50)</h2>
<div class="meta">扣繳義務人:{company_name}　統一編號:{company_tax_id}　地址:{company_address}　製表日期:{_dwh.today().isoformat()}</div>
<table>
<thead><tr><th>#</th><th>員工姓名</th><th>身分證字號</th><th>地址</th><th>年度薪資合計(元)</th><th>二代健保補充費(元)</th><th>扣繳稅額(元)</th></tr></thead>
<tbody>{rows_html}</tbody>
</table>
<div class="note">※ 本報表依薪資紀錄計算,二代健保補充費 = 超出投保薪資部分 × 2.11%.扣繳稅額請依各月薪資記錄人工確認.</div>
</body></html>"""
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


# ── 勞健保 EDI 申報 ─────────────────────────────────────────────────────────

def _get_insurance_settings():
    try:
        with get_db() as conn:
            rows = conn.execute("SELECT setting_key, setting_value FROM insurance_settings").fetchall()
        return {r['setting_key']: r['setting_value'] for r in rows}
    except Exception:
        return {}


def _edi_bytes(val, width, numeric=False):
    """Encode value to fixed-width bytes (Big5 for text, ASCII-padded for numeric)"""
    s = str(val or '')
    if numeric:
        return s.rjust(width, '0').encode('ascii', errors='replace')[:width]
    try:
        b = s.encode('big5', errors='replace')
    except Exception:
        b = s.encode('ascii', errors='replace')
    if len(b) < width:
        b = b + b' ' * (width - len(b))
    return b[:width]


@app.route('/api/insurance/settings', methods=['GET'])
@require_module('salary')
def api_insurance_settings_get():
    return jsonify(_get_insurance_settings())

@app.route('/api/insurance/settings', methods=['PUT'])
@require_module('salary')
def api_insurance_settings_put():
    b = request.get_json(force=True)
    with get_db() as conn:
        for k in ('labor_insurance_no', 'health_insurance_no', 'employer_name', 'employer_id'):
            conn.execute(
                "INSERT INTO insurance_settings VALUES (%s,%s) ON CONFLICT (setting_key) DO UPDATE SET setting_value=EXCLUDED.setting_value",
                (k, str(b.get(k, '')).strip()))
    return jsonify({'ok': True})


def _get_edi_staff(staff_ids_str):
    """Fetch staff rows for EDI, optionally filtered by comma-separated IDs."""
    with get_db() as conn:
        if staff_ids_str:
            ids = [int(x) for x in staff_ids_str.split(',') if x.strip().isdigit()]
            rows = conn.execute(
                f"SELECT * FROM punch_staff WHERE id = ANY(%s) AND active=TRUE ORDER BY name",
                (ids,)).fetchall()
        else:
            rows = conn.execute("SELECT * FROM punch_staff WHERE active=TRUE ORDER BY name").fetchall()
    return rows


@app.route('/api/export/edi/labor-enroll', methods=['GET'])
@require_module('salary')
def api_edi_labor_enroll():
    """勞工保險加退保申報 EDI(Big5 固定寬度格式)"""
    event_type  = request.args.get('event_type', 'in')   # in=加保 out=退保
    staff_ids   = request.args.get('staff_ids', '')
    event_date  = request.args.get('event_date', '')
    cfg         = _get_insurance_settings()
    labor_no    = cfg.get('labor_insurance_no', '').ljust(8)[:8]
    event_code  = b'1' if event_type == 'in' else b'2'
    event_roc   = _roc_date(event_date).encode('ascii')

    lines = []
    for s in _get_edi_staff(staff_ids):
        gender_code = b'1' if (s.get('gender') or '').upper() in ('M', '男') else b'2'
        insured = str(int(float(s.get('insured_salary') or 0))).rjust(6, '0').encode('ascii')
        line = (
            _edi_bytes(labor_no, 8) +
            _edi_bytes(s['name'], 20) +
            _edi_bytes(s.get('national_id', ''), 10) +
            _roc_date(s.get('birth_date')).encode('ascii') +
            event_roc +
            event_code +
            insured +
            gender_code +
            b'00'   # 職業類別(一般)
        )
        lines.append(line)
    content = b'\r\n'.join(lines)
    fname   = f'labor_{"enroll" if event_type=="in" else "exit"}_{event_date or "date"}.edi'
    from flask import Response as _FRe
    return _FRe(content, mimetype='application/octet-stream',
                headers={'Content-Disposition': f'attachment; filename={fname}'})


@app.route('/api/export/edi/labor-salary', methods=['GET'])
@require_module('salary')
def api_edi_labor_salary():
    """勞工保險投保薪資調整申報 EDI"""
    month     = request.args.get('month', '')
    staff_ids = request.args.get('staff_ids', '')
    cfg       = _get_insurance_settings()
    labor_no  = cfg.get('labor_insurance_no', '').ljust(8)[:8]
    if not month:
        from datetime import date as _dm2
        month = _dm2.today().strftime('%Y-%m')
    month_roc = f"{int(month[:4]) - 1911:03d}{month[5:7]}".encode('ascii')

    lines = []
    for s in _get_edi_staff(staff_ids):
        insured = str(int(float(s.get('insured_salary') or 0))).rjust(6, '0').encode('ascii')
        line = (
            _edi_bytes(labor_no, 8) +
            _edi_bytes(s['name'], 20) +
            _edi_bytes(s.get('national_id', ''), 10) +
            insured +
            month_roc
        )
        lines.append(line)
    content = b'\r\n'.join(lines)
    from flask import Response as _FRs
    return _FRs(content, mimetype='application/octet-stream',
                headers={'Content-Disposition': f'attachment; filename=labor_salary_{month}.edi'})


@app.route('/api/export/edi/health-enroll', methods=['GET'])
@require_module('salary')
def api_edi_health_enroll():
    """全民健康保險加退保申報 EDI"""
    event_type = request.args.get('event_type', 'in')
    staff_ids  = request.args.get('staff_ids', '')
    event_date = request.args.get('event_date', '')
    cfg        = _get_insurance_settings()
    health_no  = cfg.get('health_insurance_no', '').ljust(10)[:10]
    event_code = b'1' if event_type == 'in' else b'2'
    event_roc  = _roc_date(event_date).encode('ascii')

    lines = []
    for s in _get_edi_staff(staff_ids):
        gender_code = b'1' if (s.get('gender') or '').upper() in ('M', '男') else b'2'
        insured = str(int(float(s.get('insured_salary') or 0))).rjust(6, '0').encode('ascii')
        line = (
            _edi_bytes(health_no, 10) +
            _edi_bytes(s['name'], 20) +
            _edi_bytes(s.get('national_id', ''), 10) +
            _roc_date(s.get('birth_date')).encode('ascii') +
            event_roc +
            event_code +
            insured +
            gender_code
        )
        lines.append(line)
    content = b'\r\n'.join(lines)
    fname   = f'health_{"enroll" if event_type=="in" else "exit"}_{event_date or "date"}.edi'
    from flask import Response as _FRh
    return _FRh(content, mimetype='application/octet-stream',
                headers={'Content-Disposition': f'attachment; filename={fname}'})


# ── 多店管理 ─────────────────────────────────────────────────────────────────








# ── 排班需求 & 自動排班 ──────────────────────────────────────────────────────







# ─── Entry Point ──────────────────────────────────────────────────────────────

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8010))
    app.run(host='0.0.0.0', port=port, debug=False)

# ═══════════════════════════════════════════════════════════════════
# Feature: Salary PDF (HTML print endpoint)
# ═══════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════
# Feature: Batch Review (批次審核)
# ═══════════════════════════════════════════════════════════════════

@app.route('/api/punch/requests/batch', methods=['POST'])
@login_required
def api_punch_req_batch():
    b      = request.get_json(force=True)
    ids    = [int(i) for i in b.get('ids', [])]
    action = b.get('action')
    by     = b.get('reviewed_by', '管理員')
    note   = b.get('review_note', '')
    if not ids or action not in ('approve', 'reject'):
        return jsonify({'error': '參數錯誤'}), 400
    new_status = 'approved' if action == 'approve' else 'rejected'
    done = 0
    with get_db() as conn:
        for rid in ids:
            row = conn.execute("""
                UPDATE punch_requests SET status=%s, reviewed_by=%s,
                  review_note=%s, reviewed_at=NOW()
                WHERE id=%s AND status='pending' RETURNING *
            """, (new_status, by, note, rid)).fetchone()
            if row:
                if action == 'approve':
                    conn.execute("""
                        INSERT INTO punch_records
                          (staff_id, punch_type, punched_at, note, is_manual, manual_by)
                        VALUES (%s,%s,%s,%s,TRUE,%s)
                    """, (row['staff_id'], row['punch_type'], row['requested_at'],
                          f'補打卡申請#{rid}', by))
                    punch_month = str(row['requested_at'])[:7]
                    _trigger_salary_regen_for_leave(conn, row['staff_id'], punch_month)
                _notify_review_result(row['staff_id'], '補打卡申請', action,
                                      note and f'批次審核意見:{note}' or '')
                done += 1
    return jsonify({'ok': True, 'done': done})








# ═══════════════════════════════════════════════════════════════════
# Feature: Attendance Anomaly Detection (出勤異常)
# ═══════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════
# Feature: Staff Termination (離職流程)
# ═══════════════════════════════════════════════════════════════════

@app.route('/api/punch/staff/<int:sid>/terminate', methods=['POST'])
@login_required
def api_staff_terminate(sid):
    """辦理離職:設定離職日、停用帳號、記錄備註"""
    b = request.get_json(force=True)
    termination_date = b.get('termination_date', '')
    reason           = b.get('reason', '').strip()
    last_month       = b.get('last_salary_month', '')
    note             = b.get('note', '').strip()

    if not termination_date:
        return jsonify({'error': '請填寫離職日期'}), 400

    with get_db() as conn:
        # Ensure column exists
        try:
            conn.execute("ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS termination_date DATE")
            conn.execute("ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS termination_reason TEXT DEFAULT ''")
            conn.execute("ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS termination_note TEXT DEFAULT ''")
        except Exception:
            pass

        row = conn.execute("""
            UPDATE punch_staff SET
              active = FALSE,
              termination_date   = %s,
              termination_reason = %s,
              termination_note   = %s,
              salary_notes = COALESCE(salary_notes,'') || %s
            WHERE id = %s RETURNING *
        """, (termination_date, reason, note,
              f'\n【離職】{termination_date} {reason}',
              sid)).fetchone()
        if not row:
            return ('', 404)
        # 清除離職日之後的排班及所有待審申請
        conn.execute(
            "DELETE FROM shift_assignments WHERE staff_id=%s AND shift_date > %s",
            (sid, termination_date))
        for tbl in ('leave_requests', 'overtime_requests', 'schedule_requests'):
            try:
                conn.execute(
                    f"UPDATE {tbl} SET status='cancelled' WHERE staff_id=%s AND status IN ('pending','modified_pending')",
                    (sid,))
            except Exception:
                pass

    return jsonify({
        'ok': True,
        'staff_id': sid,
        'name': row['name'],
        'termination_date': termination_date,
        'last_salary_month': last_month,
    })


@app.route('/api/punch/staff/<int:sid>/reinstate', methods=['POST'])
@login_required
def api_staff_reinstate(sid):
    """復職(重新啟用帳號)"""
    with get_db() as conn:
        row = conn.execute("""
            UPDATE punch_staff SET active=TRUE,
              termination_date=NULL, termination_reason='', termination_note=''
            WHERE id=%s RETURNING *
        """, (sid,)).fetchone()
    return jsonify(punch_staff_row(row)) if row else ('', 404)


@app.route('/api/punch/staff/terminated', methods=['GET'])
@login_required
def api_staff_terminated_list():
    """離職員工清單"""
    with get_db() as conn:
        # Check if column exists
        try:
            rows = conn.execute("""
                SELECT id, name, employee_code, department, role,
                       hire_date, termination_date, termination_reason
                FROM punch_staff
                WHERE active = FALSE
                ORDER BY termination_date DESC NULLS LAST, name
            """).fetchall()
        except Exception:
            rows = conn.execute(
                "SELECT id, name, employee_code, department, role, hire_date FROM punch_staff WHERE active=FALSE"
            ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        for f in ('hire_date','termination_date'):
            if d.get(f): d[f] = str(d[f])
        result.append(d)
    return jsonify(result)


# ═══════════════════════════════════════════════════════════════════
# Feature: Salary Formula Builder support (公式說明 API)
# ═══════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════
# Finance Module (財務模組)
# ═══════════════════════════════════════════════════════════════════


def init_finance_db():
    migrations = [
        """CREATE TABLE IF NOT EXISTS finance_categories (
            id          SERIAL PRIMARY KEY,
            name        TEXT NOT NULL,
            type        TEXT NOT NULL DEFAULT 'expense',
            color       TEXT DEFAULT '#4a7bda',
            sort_order  INT DEFAULT 0,
            active      BOOLEAN DEFAULT TRUE,
            created_at  TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS finance_records (
            id              SERIAL PRIMARY KEY,
            record_date     DATE NOT NULL,
            category_id     INT REFERENCES finance_categories(id) ON DELETE SET NULL,
            type            TEXT NOT NULL DEFAULT 'expense',
            title           TEXT NOT NULL,
            amount          NUMERIC(14,2) NOT NULL DEFAULT 0,
            tax_amount      NUMERIC(14,2) DEFAULT 0,
            vendor          TEXT DEFAULT '',
            invoice_no      TEXT DEFAULT '',
            note            TEXT DEFAULT '',
            document_id     INT,
            created_by      TEXT DEFAULT '',
            created_at      TIMESTAMPTZ DEFAULT NOW(),
            updated_at      TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS finance_documents (
            id              SERIAL PRIMARY KEY,
            filename        TEXT NOT NULL,
            doc_type        TEXT DEFAULT '',
            ocr_raw         JSONB DEFAULT '{}',
            upload_date     DATE DEFAULT CURRENT_DATE,
            created_at      TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS finance_recurring (
            id              SERIAL PRIMARY KEY,
            title           TEXT NOT NULL,
            type            TEXT NOT NULL DEFAULT 'expense',
            category_id     INT REFERENCES finance_categories(id) ON DELETE SET NULL,
            amount          NUMERIC(14,2) NOT NULL DEFAULT 0,
            tax_amount      NUMERIC(14,2) DEFAULT 0,
            vendor          TEXT DEFAULT '',
            note            TEXT DEFAULT '',
            frequency       TEXT NOT NULL DEFAULT 'monthly',
            day_of_month    INT DEFAULT 1,
            start_date      DATE NOT NULL,
            end_date        DATE,
            last_generated  TEXT DEFAULT '',
            active          BOOLEAN DEFAULT TRUE,
            created_at      TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS bank_statements (
            id                  SERIAL PRIMARY KEY,
            account_name        TEXT DEFAULT '',
            txn_date            DATE NOT NULL,
            amount              NUMERIC(14,2) NOT NULL,
            txn_type            TEXT DEFAULT 'debit',
            description         TEXT DEFAULT '',
            reconciled          BOOLEAN DEFAULT FALSE,
            matched_record_id   INT REFERENCES finance_records(id) ON DELETE SET NULL,
            import_batch        TEXT DEFAULT '',
            created_at          TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS finance_payables (
            id              SERIAL PRIMARY KEY,
            payable_type    TEXT NOT NULL DEFAULT 'payable',
            title           TEXT NOT NULL,
            party_name      TEXT DEFAULT '',
            invoice_no      TEXT DEFAULT '',
            amount          NUMERIC(14,2) NOT NULL DEFAULT 0,
            due_date        DATE,
            status          TEXT NOT NULL DEFAULT 'open',
            paid_date       DATE,
            linked_record_id INT REFERENCES finance_records(id) ON DELETE SET NULL,
            note            TEXT DEFAULT '',
            created_at      TIMESTAMPTZ DEFAULT NOW(),
            updated_at      TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS finance_budgets (
            id              SERIAL PRIMARY KEY,
            year            INT NOT NULL,
            month           INT NOT NULL,
            category_id     INT REFERENCES finance_categories(id) ON DELETE CASCADE,
            budget_amount   NUMERIC(14,2) NOT NULL DEFAULT 0,
            created_at      TIMESTAMPTZ DEFAULT NOW(),
            updated_at      TIMESTAMPTZ DEFAULT NOW(),
            UNIQUE(year, month, category_id)
        )""",
        "ALTER TABLE salary_records ADD COLUMN IF NOT EXISTS finance_synced BOOLEAN DEFAULT FALSE",
    ]
    for sql in migrations:
        try:
            with get_db() as conn:
                conn.execute(sql)
        except Exception as e:
            print(f"[finance_init] {str(e)[:80]}")

    # Seed default categories
    defaults_income = [
        ('餐飲內用收入', 'income', '#2e9e6b', 1),
        ('外帶收入',     'income', '#0ea5e9', 2),
        ('外送收入',     'income', '#8b5cf6', 3),
        ('其他收入',     'income', '#c8a96e', 4),
    ]
    defaults_expense = [
        ('食材成本',   'expense', '#d64242', 10),
        ('薪資支出',   'expense', '#e07b2a', 11),
        ('租金',       'expense', '#8892a4', 12),
        ('水電費',     'expense', '#4a7bda', 13),
        ('設備維修',   'expense', '#e05c8a', 14),
        ('消耗品',     'expense', '#6366f1', 15),
        ('廣告行銷',   'expense', '#f59e0b', 16),
        ('其他支出',   'expense', '#64748b', 17),
    ]
    try:
        with get_db() as conn:
            cnt = conn.execute("SELECT COUNT(*) as c FROM finance_categories").fetchone()['c']
            if cnt == 0:
                for name, ftype, color, sort in (defaults_income + defaults_expense):
                    conn.execute(
                        "INSERT INTO finance_categories (name,type,color,sort_order) VALUES (%s,%s,%s,%s)",
                        (name, ftype, color, sort)
                    )
    except Exception as e:
        print(f"[finance_seed] {e}")

init_finance_db()



# ── Finance Categories ─────────────────────────────────────────

@app.route('/api/finance/categories', methods=['GET'])
@require_module('finance')
def api_finance_categories_list():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM finance_categories ORDER BY sort_order, id").fetchall()
    return jsonify([_finance_cat_row(r) for r in rows])

@app.route('/api/finance/categories', methods=['POST'])
@require_module('finance')
def api_finance_category_create():
    b = request.get_json(force=True)
    if not b.get('name','').strip(): return jsonify({'error': '名稱為必填'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO finance_categories (name,type,color,sort_order,active,statement_section)
            VALUES (%s,%s,%s,%s,%s,%s) RETURNING *
        """, (b['name'].strip(), b.get('type','expense'), b.get('color','#4a7bda'),
              int(b.get('sort_order',0)), bool(b.get('active',True)),
              b.get('statement_section') or ('operating_revenue' if b.get('type')=='income' else 'operating_expense')
             )).fetchone()
    return jsonify(_finance_cat_row(row)), 201

@app.route('/api/finance/categories/<int:cid>', methods=['PUT'])
@require_module('finance')
def api_finance_category_update(cid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE finance_categories SET name=%s,type=%s,color=%s,sort_order=%s,active=%s,statement_section=%s
            WHERE id=%s RETURNING *
        """, (b.get('name','').strip(), b.get('type','expense'), b.get('color','#4a7bda'),
              int(b.get('sort_order',0)), bool(b.get('active',True)),
              b.get('statement_section') or ('operating_revenue' if b.get('type')=='income' else 'operating_expense'),
              cid)).fetchone()
    return jsonify(_finance_cat_row(row)) if row else ('', 404)

@app.route('/api/finance/categories/<int:cid>', methods=['DELETE'])
@require_module('finance')
def api_finance_category_delete(cid):
    with get_db() as conn:
        conn.execute("DELETE FROM finance_categories WHERE id=%s", (cid,))
    return jsonify({'deleted': cid})

# ── Finance Records ────────────────────────────────────────────

@app.route('/api/finance/records', methods=['GET'])
@require_module('finance')
def api_finance_records_list():
    month  = request.args.get('month', '')
    ftype  = request.args.get('type', '')
    cat_id = request.args.get('category_id', '')
    conds, params = ['TRUE'], []
    if month:
        _d_s, _d_e = _month_date_range(month)
        conds.append('fr.record_date >= %s AND fr.record_date < %s'); params.extend([_d_s, _d_e])
    if ftype:
        conds.append("fr.type=%s"); params.append(ftype)
    if cat_id:
        conds.append("fr.category_id=%s"); params.append(int(cat_id))
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT fr.*, fc.name as category_name, fc.color as category_color,
                   fd.filename as doc_filename, fd.ocr_raw as ocr_raw
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            LEFT JOIN finance_documents fd ON fd.id=fr.document_id
            WHERE {' AND '.join(conds)}
            ORDER BY fr.record_date DESC, fr.id DESC
        """, params).fetchall()
    result = []
    for r in rows:
        d = _finance_rec_row(r)
        d['category_name']  = r['category_name']
        d['category_color'] = r['category_color']
        d['doc_filename']   = r['doc_filename']
        d['ocr_raw']        = r['ocr_raw'] if r['ocr_raw'] else None
        result.append(d)
    return jsonify(result)


@app.route('/api/finance/documents', methods=['GET'])
@require_module('finance')
def api_finance_documents_list():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT fd.*,
                   COUNT(fr.id) as linked_count,
                   MAX(fr.title) as linked_title,
                   MAX(fr.id) as linked_record_id
            FROM finance_documents fd
            LEFT JOIN finance_records fr ON fr.document_id = fd.id
            GROUP BY fd.id
            ORDER BY fd.created_at DESC
        """).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        if d.get('upload_date'): d['upload_date'] = str(d['upload_date'])
        if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
        d['linked_count'] = int(d['linked_count'] or 0)
        result.append(d)
    return jsonify(result)

@app.route('/api/finance/records', methods=['POST'])
@require_module('finance')
def api_finance_record_create():
    b = request.get_json(force=True)
    if not b.get('title','').strip(): return jsonify({'error': '標題為必填'}), 400
    if not b.get('record_date'):      return jsonify({'error': '日期為必填'}), 400
    if float(b.get('amount', 0) or 0) <= 0: return jsonify({'error': '金額必須大於0'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO finance_records
              (record_date, category_id, type, title, amount, tax_amount, vendor, invoice_no, note, document_id, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *
        """, (b['record_date'], b.get('category_id') or None, b.get('type','expense'),
              b['title'].strip(), float(b.get('amount',0)), float(b.get('tax_amount',0)),
              b.get('vendor','').strip(), b.get('invoice_no','').strip(),
              b.get('note','').strip(), b.get('document_id') or None,
              session.get('admin_display_name',''))).fetchone()
    return jsonify(_finance_rec_row(row)), 201

@app.route('/api/finance/records/<int:rid>', methods=['PUT'])
@require_module('finance')
def api_finance_record_update(rid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE finance_records SET
              record_date=%s, category_id=%s, type=%s, title=%s, amount=%s,
              tax_amount=%s, vendor=%s, invoice_no=%s, note=%s, updated_at=NOW()
            WHERE id=%s RETURNING *
        """, (b['record_date'], b.get('category_id') or None, b.get('type','expense'),
              b.get('title','').strip(), float(b.get('amount',0)), float(b.get('tax_amount',0)),
              b.get('vendor','').strip(), b.get('invoice_no','').strip(),
              b.get('note','').strip(), rid)).fetchone()
    return jsonify(_finance_rec_row(row)) if row else ('', 404)

@app.route('/api/finance/records/<int:rid>', methods=['DELETE'])
@require_module('finance')
def api_finance_record_delete(rid):
    with get_db() as conn:
        conn.execute("DELETE FROM finance_records WHERE id=%s", (rid,))
    return jsonify({'deleted': rid})

# ── Finance P&L Summary ────────────────────────────────────────

@app.route('/api/finance/summary/<year>/<month>', methods=['GET'])
@require_module('finance')
def api_finance_summary(year, month):
    period = f"{year}-{month.zfill(2)}"
    with get_db() as conn:
        _fin_d_s, _fin_d_e = _month_date_range(period)
        totals = conn.execute("""
            SELECT type, COALESCE(SUM(amount),0) as total
            FROM finance_records
            WHERE record_date >= %s AND record_date < %s
            GROUP BY type
        """, (_fin_d_s, _fin_d_e)).fetchall()
        income  = next((float(r['total']) for r in totals if r['type']=='income'), 0.0)
        expense = next((float(r['total']) for r in totals if r['type']=='expense'), 0.0)

        by_cat = conn.execute("""
            SELECT fc.name, fc.color, fr.type, COALESCE(SUM(fr.amount),0) as total
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            WHERE fr.record_date >= %s AND fr.record_date < %s
            GROUP BY fc.name, fc.color, fr.type
            ORDER BY total DESC
        """, (_fin_d_s, _fin_d_e)).fetchall()

        # Last 6 months trend
        trend = conn.execute("""
            SELECT to_char(record_date,'YYYY-MM') as mon,
                   type, COALESCE(SUM(amount),0) as total
            FROM finance_records
            WHERE record_date >= (DATE_TRUNC('month', %s::date) - INTERVAL '5 months')
              AND record_date <  (DATE_TRUNC('month', %s::date) + INTERVAL '1 month')
            GROUP BY to_char(record_date,'YYYY-MM'), type
            ORDER BY mon
        """, (f"{period}-01", f"{period}-01")).fetchall()

    return jsonify({
        'income':  income,
        'expense': expense,
        'net':     income - expense,
        'by_category': [
            {'name': r['name'] or '未分類', 'color': r['color'] or '#8892a4',
             'type': r['type'], 'total': float(r['total'])}
            for r in by_cat
        ],
        'trend': [
            {'month': r['mon'], 'type': r['type'], 'total': float(r['total'])}
            for r in trend
        ],
    })

# ── Finance OCR ────────────────────────────────────────────────

@app.route('/api/finance/ocr', methods=['POST'])
@require_module('finance')
def api_finance_ocr():
    import anthropic as _ant
    import base64, re as _re

    if not ANTHROPIC_API_KEY:
        return jsonify({'error': '尚未設定 ANTHROPIC_API_KEY 環境變數'}), 500

    file = request.files.get('file')
    if not file:
        return jsonify({'error': '請上傳圖片或 PDF 檔案'}), 400

    raw = file.read()
    media_type = file.content_type or 'image/jpeg'
    # Only image types supported by Claude vision
    if media_type not in ('image/jpeg','image/png','image/gif','image/webp'):
        media_type = 'image/jpeg'

    img_b64 = base64.standard_b64encode(raw).decode()

    client = _ant.Anthropic(api_key=ANTHROPIC_API_KEY)
    try:
        msg = client.messages.create(
            model='claude-sonnet-4-6',
            max_tokens=1024,
            messages=[{
                'role': 'user',
                'content': [
                    {'type': 'image', 'source': {'type': 'base64', 'media_type': media_type, 'data': img_b64}},
                    {'type': 'text', 'text': (
                        '請辨識此文件,以JSON格式回傳以下欄位(找不到的欄位填null):\n'
                        '{"date":"YYYY-MM-DD","vendor":"廠商名稱","invoice_no":"發票或單據號碼",'
                        '"total_amount":含稅總金額數字,"tax_amount":稅額數字,"pre_tax_amount":未稅金額數字,'
                        '"doc_type":"invoice或receipt或expense之一",'
                        '"title":"建議記帳標題(簡短)",'
                        '"items":[{"name":"品項","qty":數量,"unit_price":單價,"amount":小計}],'
                        '"currency":"TWD"}\n只回傳JSON,不要其他文字或markdown.'
                    )}
                ]
            }]
        )
        text = msg.content[0].text.strip()
        text = _re.sub(r'^```json\s*', '', text, flags=_re.MULTILINE)
        text = _re.sub(r'\s*```$', '', text, flags=_re.MULTILINE)
        result = _json.loads(text)
    except _json.JSONDecodeError:
        result = {'raw_text': text, 'error': 'OCR 回傳格式無法解析'}
    except Exception as e:
        return jsonify({'error': f'OCR 失敗:{str(e)}'}), 500

    try:
        with get_db() as conn:
            doc = conn.execute("""
                INSERT INTO finance_documents (filename, doc_type, ocr_raw, upload_date)
                VALUES (%s,%s,%s,CURRENT_DATE) RETURNING id
            """, (file.filename, result.get('doc_type',''), _json.dumps(result))).fetchone()
        result['document_id'] = doc['id']
    except Exception as e:
        print(f"[finance_ocr doc save] {e}")

    return jsonify(result)

# ── Finance Export ─────────────────────────────────────────────

@app.route('/api/finance/export', methods=['GET'])
@require_module('finance')
def api_finance_export():
    import csv, io as _io
    month = request.args.get('month', '')
    conds, params = ['TRUE'], []
    if month:
        _d_s, _d_e = _month_date_range(month)
        conds.append('fr.record_date >= %s AND fr.record_date < %s'); params.extend([_d_s, _d_e])
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT fr.record_date, fr.type, fr.title, fr.amount, fr.tax_amount,
                   fr.vendor, fr.invoice_no, fr.note, fc.name as category_name
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            WHERE {' AND '.join(conds)}
            ORDER BY fr.record_date, fr.id
        """, params).fetchall()
    out = _io.StringIO()
    w = csv.writer(out)
    w.writerow(['日期','類型','類別','標題','金額','稅額','廠商','單據號碼','備註'])
    for r in rows:
        w.writerow([str(r['record_date']), '收入' if r['type']=='income' else '支出',
                    r['category_name'] or '', r['title'], r['amount'], r['tax_amount'] or 0,
                    r['vendor'] or '', r['invoice_no'] or '', r['note'] or ''])
    fname = f"finance_{month or 'all'}.csv"
    return (('\ufeff' + out.getvalue()).encode('utf-8'),
            200, {'Content-Type': 'text/csv; charset=utf-8',
                  'Content-Disposition': f'attachment; filename={fname}'})

# ── Finance Settings & Financial Statements ────────────────────

def init_finance_settings_db():
    migrations = [
        "ALTER TABLE finance_categories ADD COLUMN IF NOT EXISTS statement_section TEXT",
        """CREATE TABLE IF NOT EXISTS finance_settings (
            id            SERIAL PRIMARY KEY,
            setting_key   TEXT UNIQUE NOT NULL,
            setting_value TEXT DEFAULT ''
        )""",
    ]
    for sql in migrations:
        try:
            with get_db() as conn:
                conn.execute(sql)
        except Exception as e:
            print(f"[finance_settings_init] {str(e)[:80]}")

    # Set default statement_section based on type for existing rows with NULL
    section_defaults = {
        '餐飲內用收入': 'operating_revenue',
        '外帶收入':     'operating_revenue',
        '外送收入':     'operating_revenue',
        '其他收入':     'other_revenue',
        '食材成本':     'cogs',
        '薪資支出':     'operating_expense',
        '租金':         'operating_expense',
        '水電費':       'operating_expense',
        '設備維修':     'operating_expense',
        '消耗品':       'operating_expense',
        '廣告行銷':     'operating_expense',
        '其他支出':     'other_expense',
    }
    try:
        with get_db() as conn:
            # Fill named defaults
            for name, sec in section_defaults.items():
                conn.execute(
                    "UPDATE finance_categories SET statement_section=%s WHERE name=%s AND statement_section IS NULL",
                    (sec, name)
                )
            # Remaining NULLs: income -> operating_revenue, expense -> operating_expense
            conn.execute("""
                UPDATE finance_categories
                SET statement_section = CASE WHEN type='income' THEN 'operating_revenue' ELSE 'operating_expense' END
                WHERE statement_section IS NULL
            """)
    except Exception as e:
        print(f"[finance_settings_seed] {e}")

    # Seed settings defaults
    for k, v in [('company_name', ''), ('opening_cash', '0'), ('opening_equity', '0'),
                  ('company_tax_id', ''), ('company_address', '')]:
        try:
            with get_db() as conn:
                conn.execute(
                    "INSERT INTO finance_settings (setting_key, setting_value) VALUES (%s,%s) ON CONFLICT (setting_key) DO NOTHING",
                    (k, v)
                )
        except Exception as e:
            print(f"[finance_settings_default] {e}")

init_finance_settings_db()


def init_insurance_db():
    try:
        with get_db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS insurance_settings (
                    setting_key   TEXT PRIMARY KEY,
                    setting_value TEXT DEFAULT ''
                )
            """)
        for k, v in [('labor_insurance_no', ''), ('health_insurance_no', ''),
                     ('employer_name', ''), ('employer_id', '')]:
            with get_db() as conn:
                conn.execute(
                    "INSERT INTO insurance_settings (setting_key, setting_value) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                    (k, v))
    except Exception as e:
        print(f"[insurance_init] {e}")

init_insurance_db()


# ═══════════════════════════════════════════════════════════════════════════════
# 教育訓練追蹤 (Training & Certificate Tracking)
# ═══════════════════════════════════════════════════════════════════════════════

def init_training_db():
    try:
        with get_db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS training_records (
                    id              SERIAL PRIMARY KEY,
                    staff_id        INT REFERENCES punch_staff(id) ON DELETE CASCADE,
                    course_name     TEXT NOT NULL,
                    category        TEXT NOT NULL DEFAULT 'general',
                    completed_date  DATE,
                    expiry_date     DATE,
                    certificate_no  TEXT DEFAULT '',
                    note            TEXT DEFAULT '',
                    created_at      TIMESTAMPTZ DEFAULT NOW(),
                    updated_at      TIMESTAMPTZ DEFAULT NOW()
                )
            """)
    except Exception as e:
        print(f"[training_init] {e}")

init_training_db()

TRAINING_CATEGORIES = {
    'food_safety':  '食品安全',
    'fire_safety':  '消防安全',
    'first_aid':    '急救訓練',
    'hygiene':      '衛生管理',
    'service':      '服務禮儀',
    'equipment':    '設備操作',
    'general':      '一般訓練',
    'other':        '其他',
}






# ── 薪資計算預覽 (Salary Preview without saving) ───────────────────────────────



@app.route('/api/finance/settings', methods=['GET'])
@require_module('finance')
def api_finance_settings_get():
    with get_db() as conn:
        rows = conn.execute("SELECT setting_key, setting_value FROM finance_settings").fetchall()
    return jsonify({r['setting_key']: r['setting_value'] for r in rows})


@app.route('/api/finance/settings', methods=['POST'])
@require_module('finance')
def api_finance_settings_save():
    data = request.get_json(force=True)
    with get_db() as conn:
        for k, v in data.items():
            conn.execute(
                "INSERT INTO finance_settings (setting_key, setting_value) VALUES (%s,%s) ON CONFLICT (setting_key) DO UPDATE SET setting_value=EXCLUDED.setting_value",
                (k, str(v))
            )
    return jsonify({'ok': True})


def _get_finance_settings():
    try:
        with get_db() as conn:
            rows = conn.execute("SELECT setting_key, setting_value FROM finance_settings").fetchall()
            return {r['setting_key']: r['setting_value'] for r in rows}
    except:
        return {}




def _compute_statements(year, month):
    """Compute all three financial statements for the given year/month."""
    from collections import defaultdict
    period = f"{year}-{str(month).zfill(2)}"
    settings = _get_finance_settings()
    opening_cash   = float(settings.get('opening_cash',   0) or 0)
    opening_equity = float(settings.get('opening_equity', 0) or 0)
    company_name   = settings.get('company_name', '') or '公司名稱'

    with get_db() as conn:
        records = conn.execute("""
            SELECT fr.type, fr.amount,
                   fc.name            AS cat_name,
                   fc.statement_section AS section
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id = fr.category_id
            WHERE fr.record_date >= %s AND fr.record_date < %s
        """, _month_date_range(period)).fetchall()

        # Cumulative net before this month (for balance sheet period-start)
        prev = conn.execute("""
            SELECT
                COALESCE(SUM(CASE WHEN type='income'  THEN amount ELSE 0 END), 0) AS cum_income,
                COALESCE(SUM(CASE WHEN type='expense' THEN amount ELSE 0 END), 0) AS cum_expense
            FROM finance_records
            WHERE record_date < DATE_TRUNC('month', %s::date)
        """, (f"{period}-01",)).fetchone()

    cum_net_before = float(prev['cum_income']) - float(prev['cum_expense'])

    by_section = defaultdict(float)
    by_cat     = defaultdict(float)
    for r in records:
        sec = r['section'] or ('operating_revenue' if r['type'] == 'income' else 'operating_expense')
        by_section[sec]               += float(r['amount'])
        by_cat[(sec, r['cat_name'] or '未分類')] += float(r['amount'])

    operating_revenue = by_section['operating_revenue']
    other_revenue     = by_section['other_revenue']
    cogs              = by_section['cogs']
    operating_expense = by_section['operating_expense']
    other_expense     = by_section['other_expense']

    gross_profit      = operating_revenue - cogs
    operating_income  = gross_profit - operating_expense
    net_income        = operating_income + other_revenue - other_expense
    total_income      = operating_revenue + other_revenue
    total_expense     = cogs + operating_expense + other_expense

    cum_net_total     = cum_net_before + net_income
    cash_balance      = opening_cash + opening_equity + cum_net_total
    # 期初現金亦屬業主投入(權益),須計入權益側,資產負債表才會平衡
    contributed_capital = opening_cash + opening_equity
    total_equity      = contributed_capital + cum_net_total

    def cat_lines(section):
        return [{'name': k[1], 'amount': round(v, 2)}
                for k, v in sorted(by_cat.items()) if k[0] == section]

    return {
        'company_name': company_name,
        'year': int(year), 'month': int(month),
        'roc_year': _roc_year(year),
        'last_day': _month_last_day(year, month),
        'income_statement': {
            'operating_revenue':       round(operating_revenue, 2),
            'operating_revenue_lines': cat_lines('operating_revenue'),
            'other_revenue':           round(other_revenue, 2),
            'other_revenue_lines':     cat_lines('other_revenue'),
            'cogs':                    round(cogs, 2),
            'cogs_lines':              cat_lines('cogs'),
            'gross_profit':            round(gross_profit, 2),
            'operating_expense':       round(operating_expense, 2),
            'operating_expense_lines': cat_lines('operating_expense'),
            'operating_income':        round(operating_income, 2),
            'other_expense':           round(other_expense, 2),
            'other_expense_lines':     cat_lines('other_expense'),
            'net_income':              round(net_income, 2),
        },
        'balance_sheet': {
            'cash':                    round(cash_balance, 2),
            'total_assets':            round(cash_balance, 2),
            'total_liabilities':       0,
            'opening_equity':          round(contributed_capital, 2),
            'retained_earnings':       round(cum_net_total, 2),
            'total_equity':            round(total_equity, 2),
            'total_liabilities_equity': round(total_equity, 2),
        },
        'cash_flow': {
            'operating_inflow':        round(total_income, 2),
            'operating_inflow_lines':  cat_lines('operating_revenue') + cat_lines('other_revenue'),
            'operating_outflow':       round(total_expense, 2),
            'operating_outflow_lines': cat_lines('cogs') + cat_lines('operating_expense') + cat_lines('other_expense'),
            'operating_net':           round(total_income - total_expense, 2),
            'investing_net':           0,
            'financing_net':           0,
            'net_change':              round(total_income - total_expense, 2),
            'opening_cash':            round(opening_cash + opening_equity + cum_net_before, 2),
            'closing_cash':            round(cash_balance, 2),
        },
    }


@app.route('/api/finance/statements/<year>/<month>', methods=['GET'])
@require_module('finance')
def api_finance_statements(year, month):
    try:
        return jsonify(_compute_statements(year, month))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/finance/export/statements/<year>/<month>', methods=['GET'])
@require_module('finance')
def api_finance_export_statements(year, month):
    import openpyxl, io as _io
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    d  = _compute_statements(year, month)
    co = d['company_name']
    ry = d['roc_year']
    m  = d['month']
    ld = d['last_day']
    IS = d['income_statement']
    BS = d['balance_sheet']
    CF = d['cash_flow']

    wb = openpyxl.Workbook()

    NAVY   = '1C3557'
    AMT    = '#,##0'
    FONT   = '標楷體'
    thin   = Side(style='thin')
    medium = Side(style='medium')

    def _border(top=False, bottom=False, dbl=False):
        return Border(
            top    = thin   if top else None,
            bottom = medium if dbl  else (thin if bottom else None),
        )

    def setup_ws(ws, title, date_str):
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 4
        ws.column_dimensions['C'].width = 18
        for row_vals, styles in [
            ([co, '', ''],        {'font': Font(FONT, bold=True, size=14), 'align': 'center', 'merge': True}),
            ([title, '', ''],     {'font': Font(FONT, bold=True, size=13), 'align': 'center', 'merge': True}),
            ([date_str, '', ''],  {'font': Font(FONT, size=11),            'align': 'center', 'merge': True}),
        ]:
            ws.append(row_vals)
            r = ws.max_row
            ws.cell(r, 1).font      = styles['font']
            ws.cell(r, 1).alignment = Alignment(horizontal=styles['align'])
            if styles.get('merge'):
                ws.merge_cells(f'A{r}:C{r}')
        ws.append([])  # blank

        # column headers
        ws.append(['項　　目', '', '金　額(元)'])
        r = ws.max_row
        ws.cell(r, 1).font      = Font(FONT, bold=True, size=11, color='FFFFFF')
        ws.cell(r, 3).font      = Font(FONT, bold=True, size=11, color='FFFFFF')
        ws.cell(r, 1).fill      = PatternFill('solid', fgColor=NAVY)
        ws.cell(r, 3).fill      = PatternFill('solid', fgColor=NAVY)
        ws.cell(r, 2).fill      = PatternFill('solid', fgColor=NAVY)
        ws.cell(r, 1).alignment = Alignment(horizontal='center')
        ws.cell(r, 3).alignment = Alignment(horizontal='right')

    def row(ws, label, amount=None, indent=0, bold=False, subtotal=False, total=False, dbl=False):
        prefix = '　' * indent
        ws.append([prefix + label, '', amount])
        r = ws.max_row
        b = bold or subtotal or total
        ws.cell(r, 1).font      = Font(FONT, bold=b, size=11)
        ws.cell(r, 1).alignment = Alignment(horizontal='left')
        if amount is not None:
            ws.cell(r, 3).font           = Font(FONT, bold=b, size=11)
            ws.cell(r, 3).number_format  = AMT
            ws.cell(r, 3).alignment      = Alignment(horizontal='right')
        if subtotal or total:
            ws.cell(r, 3).border = _border(top=(subtotal or total), bottom=(subtotal or total), dbl=dbl)
        elif amount is None:
            ws.cell(r, 1).font = Font(FONT, bold=True, size=11)

    # ─── 損益表 ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = '損益表'
    setup_ws(ws1, '損益表', f'中華民國{ry}年{m}月份')

    row(ws1, '一、營業收入', bold=True)
    for l in IS['operating_revenue_lines']:
        row(ws1, l['name'], l['amount'], indent=2)
    row(ws1, '營業收入合計', IS['operating_revenue'], indent=1, subtotal=True)
    ws1.append([])

    row(ws1, '二、營業成本', bold=True)
    for l in IS['cogs_lines']:
        row(ws1, l['name'], l['amount'], indent=2)
    row(ws1, '營業成本合計', IS['cogs'], indent=1, subtotal=True)
    ws1.append([])

    row(ws1, '毛　利', IS['gross_profit'], indent=1, total=True)
    ws1.append([])

    row(ws1, '三、營業費用', bold=True)
    for l in IS['operating_expense_lines']:
        row(ws1, l['name'], l['amount'], indent=2)
    row(ws1, '營業費用合計', IS['operating_expense'], indent=1, subtotal=True)
    ws1.append([])

    row(ws1, '營業利益(損失)', IS['operating_income'], indent=1, total=True)
    ws1.append([])

    if IS['other_revenue'] or IS['other_revenue_lines']:
        row(ws1, '四、營業外收入', bold=True)
        for l in IS['other_revenue_lines']:
            row(ws1, l['name'], l['amount'], indent=2)
        row(ws1, '營業外收入合計', IS['other_revenue'], indent=1, subtotal=True)
        ws1.append([])

    if IS['other_expense'] or IS['other_expense_lines']:
        row(ws1, '五、營業外費用', bold=True)
        for l in IS['other_expense_lines']:
            row(ws1, l['name'], l['amount'], indent=2)
        row(ws1, '營業外費用合計', IS['other_expense'], indent=1, subtotal=True)
        ws1.append([])

    row(ws1, '本期淨利(損)', IS['net_income'], bold=True, total=True, dbl=True)

    # ─── 資產負債表 ───────────────────────────────────────────────
    ws2 = wb.create_sheet('資產負債表')
    setup_ws(ws2, '資產負債表', f'中華民國{ry}年{m}月{ld}日')

    row(ws2, '【資　產】', bold=True)
    row(ws2, '流動資產', indent=1, bold=True)
    row(ws2, '現金及約當現金', BS['cash'], indent=2)
    row(ws2, '資產合計', BS['total_assets'], indent=1, total=True)
    ws2.append([])

    row(ws2, '【負　債】', bold=True)
    row(ws2, '流動負債', indent=1, bold=True)
    row(ws2, '應付帳款', 0, indent=2)
    row(ws2, '負債合計', BS['total_liabilities'], indent=1, total=True)
    ws2.append([])

    row(ws2, '【股東權益】', bold=True)
    row(ws2, '資本額', BS['opening_equity'], indent=2)
    row(ws2, '保留盈餘', BS['retained_earnings'], indent=2)
    row(ws2, '股東權益合計', BS['total_equity'], indent=1, total=True)
    ws2.append([])

    row(ws2, '負債及股東權益合計', BS['total_liabilities_equity'], bold=True, total=True, dbl=True)

    # ─── 現金流量表 ───────────────────────────────────────────────
    ws3 = wb.create_sheet('現金流量表')
    setup_ws(ws3, '現金流量表(直接法)', f'中華民國{ry}年{m}月份')

    row(ws3, '一、營業活動之現金流量', bold=True)
    row(ws3, '(一)收現收入', indent=1, bold=True)
    for l in CF['operating_inflow_lines']:
        row(ws3, l['name'], l['amount'], indent=3)
    row(ws3, '收現合計', CF['operating_inflow'], indent=2, subtotal=True)
    ws3.append([])
    row(ws3, '(二)付現費用', indent=1, bold=True)
    for l in CF['operating_outflow_lines']:
        row(ws3, l['name'], -l['amount'], indent=3)
    row(ws3, '付現合計', -CF['operating_outflow'], indent=2, subtotal=True)
    ws3.append([])
    row(ws3, '營業活動淨現金流量', CF['operating_net'], indent=1, total=True)
    ws3.append([])

    row(ws3, '二、投資活動之現金流量', bold=True)
    row(ws3, '投資活動淨現金流量', CF['investing_net'], indent=1, total=True)
    ws3.append([])

    row(ws3, '三、籌資活動之現金流量', bold=True)
    row(ws3, '籌資活動淨現金流量', CF['financing_net'], indent=1, total=True)
    ws3.append([])

    row(ws3, '四、本期現金增減', CF['net_change'], bold=True, total=True)
    row(ws3, '五、期初現金及約當現金', CF['opening_cash'], bold=True)
    row(ws3, '六、期末現金及約當現金', CF['closing_cash'], bold=True, total=True, dbl=True)

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"statements_{year}{str(month).zfill(2)}.xlsx"
    return (buf.read(), 200, {
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Content-Disposition': f'attachment; filename="{fname}"',
    })


# ═══════════════════════════════════════════════════════════════════
# Feature 1: 定期自動分錄 (Recurring Entries)
# ═══════════════════════════════════════════════════════════════════


@app.route('/api/finance/recurring', methods=['GET'])
@require_module('finance')
def api_recurring_list():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT fr.*, fc.name as category_name
            FROM finance_recurring fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            ORDER BY fr.active DESC, fr.id
        """).fetchall()
    result = []
    for r in rows:
        d = _recurring_row(r)
        d['category_name'] = r['category_name']
        result.append(d)
    return jsonify(result)

@app.route('/api/finance/recurring', methods=['POST'])
@require_module('finance')
def api_recurring_create():
    b = request.get_json(force=True)
    if not b.get('title','').strip(): return jsonify({'error': '標題為必填'}), 400
    if not b.get('start_date'):       return jsonify({'error': '開始日期為必填'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO finance_recurring
              (title, type, category_id, amount, tax_amount, vendor, note,
               frequency, day_of_month, start_date, end_date, active)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,TRUE) RETURNING *
        """, (b['title'].strip(), b.get('type','expense'), b.get('category_id') or None,
              float(b.get('amount',0)), float(b.get('tax_amount',0)),
              b.get('vendor','').strip(), b.get('note','').strip(),
              b.get('frequency','monthly'), int(b.get('day_of_month',1) or 1),
              b['start_date'], b.get('end_date') or None)).fetchone()
    return jsonify(_recurring_row(row)), 201

@app.route('/api/finance/recurring/<int:rid>', methods=['PUT'])
@require_module('finance')
def api_recurring_update(rid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE finance_recurring SET
              title=%s, type=%s, category_id=%s, amount=%s, tax_amount=%s,
              vendor=%s, note=%s, frequency=%s, day_of_month=%s,
              start_date=%s, end_date=%s, active=%s
            WHERE id=%s RETURNING *
        """, (b.get('title','').strip(), b.get('type','expense'), b.get('category_id') or None,
              float(b.get('amount',0)), float(b.get('tax_amount',0)),
              b.get('vendor','').strip(), b.get('note','').strip(),
              b.get('frequency','monthly'), int(b.get('day_of_month',1) or 1),
              b.get('start_date'), b.get('end_date') or None,
              bool(b.get('active', True)), rid)).fetchone()
    return jsonify(_recurring_row(row)) if row else ('', 404)

@app.route('/api/finance/recurring/<int:rid>', methods=['DELETE'])
@require_module('finance')
def api_recurring_delete(rid):
    with get_db() as conn:
        conn.execute("DELETE FROM finance_recurring WHERE id=%s", (rid,))
    return jsonify({'deleted': rid})

@app.route('/api/finance/recurring/generate', methods=['POST'])
@require_module('finance')
def api_recurring_generate():
    """為指定月份產生定期分錄(冪等:已產生則跳過)"""
    from datetime import date as _d, timedelta as _td
    import calendar as _cal
    b = request.get_json(force=True)
    month = b.get('month', '')  # YYYY-MM
    if not month:
        from datetime import datetime, timezone, timedelta
        month = datetime.now(timezone(timedelta(hours=8))).strftime('%Y-%m')
    y, m = int(month[:4]), int(month[5:])
    days_in_month = _cal.monthrange(y, m)[1]

    created, skipped = 0, 0
    with get_db() as conn:
        rows = conn.execute("""
            SELECT * FROM finance_recurring
            WHERE active=TRUE
              AND start_date <= %s
              AND (end_date IS NULL OR end_date >= %s)
        """, (f"{month}-28", f"{month}-01")).fetchall()

        for r in rows:
            # Check already generated this month
            if r['last_generated'] == month:
                skipped += 1
                continue
            # Check frequency
            freq = r['frequency']
            start_m = r['start_date'].month
            if freq == 'quarterly' and (m - start_m) % 3 != 0:
                skipped += 1
                continue
            if freq == 'yearly' and m != start_m:
                skipped += 1
                continue
            # Determine record date
            day = min(int(r['day_of_month'] or 1), days_in_month)
            rec_date = _d(y, m, day)
            conn.execute("""
                INSERT INTO finance_records
                  (record_date, category_id, type, title, amount, tax_amount, vendor, note, created_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'auto-recurring')
            """, (rec_date, r['category_id'], r['type'], r['title'],
                  r['amount'], r['tax_amount'] or 0, r['vendor'] or '', r['note'] or ''))
            conn.execute("UPDATE finance_recurring SET last_generated=%s WHERE id=%s",
                         (month, r['id']))
            created += 1

    return jsonify({'created': created, 'skipped': skipped, 'month': month})


# ═══════════════════════════════════════════════════════════════════
# Feature 2: 銀行對帳 (Bank Reconciliation)
# ═══════════════════════════════════════════════════════════════════


@app.route('/api/finance/bank/import', methods=['POST'])
@require_module('finance')
def api_bank_import():
    """匯入銀行對帳單 CSV"""
    import csv, io as _io
    file = request.files.get('file')
    if not file: return jsonify({'error': '請上傳 CSV 檔案'}), 400
    raw = file.read().decode('utf-8-sig', errors='replace')
    account_name = request.form.get('account_name', '').strip() or '銀行帳戶'
    import_batch = _dt.now(TW_TZ).strftime('%Y%m%d%H%M%S')

    reader = csv.reader(_io.StringIO(raw))
    rows_data = [r for r in reader if any(c.strip() for c in r)]
    if not rows_data: return jsonify({'error': 'CSV 無資料'}), 400

    # Auto-detect header row (skip rows where first column is not a date-like string)
    def _is_date(s):
        s = s.strip().replace('/', '-').replace('.', '-')
        for fmt in ('%Y-%m-%d','%Y-%m-%d','%m-%d-%Y','%d-%m-%Y'):
            try: _dt.strptime(s, fmt); return True
            except: pass
        # ROC year: 民國 e.g. 113/01/15
        import re
        if re.match(r'^\d{2,3}[/-]\d{1,2}[/-]\d{1,2}$', s):
            parts = re.split(r'[/-]', s)
            if int(parts[0]) < 200: return True
        return False

    def _parse_date(s):
        import re
        s = s.strip()
        parts = re.split(r'[/\-\.]', s)
        if len(parts) == 3:
            if int(parts[0]) < 200:
                # ROC year
                y2 = int(parts[0]) + 1911
                return f"{y2}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
            if int(parts[0]) > 31:
                return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
            if int(parts[2]) > 31:
                return f"{parts[2]}-{parts[0].zfill(2)}-{parts[1].zfill(2)}"
        return None

    def _parse_amount(s):
        import re
        s = re.sub(r'[,$\s]', '', str(s).strip())
        try: return float(s)
        except: return None

    inserted = 0
    with get_db() as conn:
        for row in rows_data:
            if len(row) < 2: continue
            date_str = _parse_date(row[0])
            if not date_str: continue
            desc = row[1].strip() if len(row) > 1 else ''
            # Format: date, desc, debit, credit  OR  date, desc, amount
            if len(row) >= 4:
                debit  = _parse_amount(row[2])
                credit = _parse_amount(row[3])
                if debit and debit > 0:
                    conn.execute("""INSERT INTO bank_statements
                        (account_name,txn_date,amount,txn_type,description,import_batch)
                        VALUES (%s,%s,%s,'debit',%s,%s)
                    """, (account_name, date_str, debit, desc, import_batch))
                    inserted += 1
                if credit and credit > 0:
                    conn.execute("""INSERT INTO bank_statements
                        (account_name,txn_date,amount,txn_type,description,import_batch)
                        VALUES (%s,%s,%s,'credit',%s,%s)
                    """, (account_name, date_str, credit, desc, import_batch))
                    inserted += 1
            elif len(row) >= 3:
                amt = _parse_amount(row[2])
                if amt is not None and amt != 0:
                    txn_type = 'credit' if amt > 0 else 'debit'
                    conn.execute("""INSERT INTO bank_statements
                        (account_name,txn_date,amount,txn_type,description,import_batch)
                        VALUES (%s,%s,%s,%s,%s,%s)
                    """, (account_name, date_str, abs(amt), txn_type, desc, import_batch))
                    inserted += 1
    return jsonify({'inserted': inserted, 'batch': import_batch})

@app.route('/api/finance/bank/statements', methods=['GET'])
@require_module('finance')
def api_bank_statements():
    month = request.args.get('month', '')
    conds, params = ['TRUE'], []
    if month:
        _d_s, _d_e = _month_date_range(month)
        conds.append("bs.txn_date >= %s AND bs.txn_date < %s"); params += [_d_s, _d_e]
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT bs.*, fr.title as matched_title, fr.amount as matched_amount,
                   fr.record_date as matched_date
            FROM bank_statements bs
            LEFT JOIN finance_records fr ON fr.id=bs.matched_record_id
            WHERE {' AND '.join(conds)}
            ORDER BY bs.txn_date DESC, bs.id DESC
        """, params).fetchall()
    result = []
    for r in rows:
        d = _bank_row(r)
        d['matched_title']  = r['matched_title']
        d['matched_amount'] = float(r['matched_amount']) if r['matched_amount'] else None
        d['matched_date']   = str(r['matched_date']) if r['matched_date'] else None
        result.append(d)
    return jsonify(result)

@app.route('/api/finance/bank/statements/<int:sid>', methods=['DELETE'])
@require_module('finance')
def api_bank_statement_delete(sid):
    with get_db() as conn:
        conn.execute("DELETE FROM bank_statements WHERE id=%s", (sid,))
    return jsonify({'deleted': sid})

@app.route('/api/finance/bank/match', methods=['POST'])
@require_module('finance')
def api_bank_match():
    b = request.get_json(force=True)
    sid = b.get('statement_id')
    rid = b.get('record_id')  # None to unmatch
    with get_db() as conn:
        if rid:
            conn.execute("UPDATE bank_statements SET reconciled=TRUE, matched_record_id=%s WHERE id=%s",
                         (rid, sid))
        else:
            conn.execute("UPDATE bank_statements SET reconciled=FALSE, matched_record_id=NULL WHERE id=%s",
                         (sid,))
    return jsonify({'ok': True})

@app.route('/api/finance/bank/auto-match', methods=['POST'])
@require_module('finance')
def api_bank_auto_match():
    """自動比對:相同金額且日期在 3 天內"""
    b = request.get_json(force=True)
    month = b.get('month', '')
    matched = 0
    _bm_d = _month_date_range(month) if month else None
    with get_db() as conn:
        stmts = conn.execute("""
            SELECT * FROM bank_statements
            WHERE reconciled=FALSE
            """ + ("AND txn_date >= %s AND txn_date < %s" if month else ""),
            (list(_bm_d) if month else [])).fetchall()
        for s in stmts:
            # Find finance record: same amount, date within ±3 days, same type direction
            ftype = 'income' if s['txn_type'] == 'credit' else 'expense'
            rec = conn.execute("""
                SELECT id FROM finance_records
                WHERE type=%s AND amount=%s
                  AND ABS(record_date - %s::date) <= 3
                  AND id NOT IN (
                      SELECT matched_record_id FROM bank_statements
                      WHERE matched_record_id IS NOT NULL
                  )
                ORDER BY ABS(record_date - %s::date), id
                LIMIT 1
            """, (ftype, s['amount'], s['txn_date'], s['txn_date'])).fetchone()
            if rec:
                conn.execute("""UPDATE bank_statements SET reconciled=TRUE, matched_record_id=%s
                                WHERE id=%s""", (rec['id'], s['id']))
                matched += 1
    return jsonify({'matched': matched})

@app.route('/api/finance/bank/summary', methods=['GET'])
@require_module('finance')
def api_bank_summary():
    month = request.args.get('month', '')
    cond = "AND txn_date >= %s AND txn_date < %s" if month else ""
    params = list(_month_date_range(month)) if month else []
    with get_db() as conn:
        r = conn.execute(f"""
            SELECT
              COUNT(*) as total,
              SUM(CASE WHEN reconciled THEN 1 ELSE 0 END) as matched,
              SUM(CASE WHEN txn_type='credit' THEN amount ELSE 0 END) as total_credit,
              SUM(CASE WHEN txn_type='debit'  THEN amount ELSE 0 END) as total_debit,
              SUM(CASE WHEN reconciled AND txn_type='credit' THEN amount ELSE 0 END) as matched_credit,
              SUM(CASE WHEN reconciled AND txn_type='debit'  THEN amount ELSE 0 END) as matched_debit
            FROM bank_statements WHERE TRUE {cond}
        """, params).fetchone()
    d = dict(r)
    for k in d:
        if d[k] is not None: d[k] = float(d[k]) if isinstance(d[k], type(r['total_credit'])) else int(d[k])
    return jsonify(d)


# ═══════════════════════════════════════════════════════════════════
# Feature 3: 稅務申報準備 (Tax Filing Prep - Taiwan VAT 401)
# ═══════════════════════════════════════════════════════════════════

@app.route('/api/finance/tax/<int:year>/<int:period>', methods=['GET'])
@require_module('finance')
def api_finance_tax(year, period):
    """
    period: 1=Jan-Feb, 2=Mar-Apr, 3=May-Jun, 4=Jul-Aug, 5=Sep-Oct, 6=Nov-Dec
    """
    if period < 1 or period > 6:
        return jsonify({'error': '期別需為 1-6'}), 400
    m_start = (period - 1) * 2 + 1
    m_end   = m_start + 1
    months  = [f"{year}-{str(m).zfill(2)}" for m in range(m_start, m_end + 1)]
    _d_s = _month_date_range(f"{year}-{m_start:02d}")[0]
    _d_e = _month_date_range(f"{year}-{m_end:02d}")[1]

    with get_db() as conn:
        rows = conn.execute("""
            SELECT fr.type, fr.amount, fr.tax_amount, fr.title,
                   fr.vendor, fr.invoice_no, fr.record_date,
                   fc.name as category_name
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            WHERE fr.record_date >= %s AND fr.record_date < %s
            ORDER BY fr.record_date, fr.type
        """, (_d_s, _d_e)).fetchall()

    sales_rows    = [r for r in rows if r['type'] == 'income']
    purchase_rows = [r for r in rows if r['type'] == 'expense']

    sales_amount    = sum(float(r['amount'])     for r in sales_rows)
    sales_tax       = sum(float(r['tax_amount'] or 0) for r in sales_rows)
    purchase_amount = sum(float(r['amount'])     for r in purchase_rows)
    purchase_tax    = sum(float(r['tax_amount'] or 0) for r in purchase_rows)
    tax_payable     = round(sales_tax - purchase_tax, 2)

    def _fmt_row(r):
        return {
            'date':     str(r['record_date']),
            'title':    r['title'],
            'vendor':   r['vendor'] or '',
            'invoice_no': r['invoice_no'] or '',
            'amount':   float(r['amount']),
            'tax_amount': float(r['tax_amount'] or 0),
            'category': r['category_name'] or '未分類',
        }

    return jsonify({
        'year': year, 'period': period,
        'roc_year': year - 1911,
        'months': months,
        'sales': {
            'rows':   [_fmt_row(r) for r in sales_rows],
            'amount': round(sales_amount, 2),
            'tax':    round(sales_tax, 2),
        },
        'purchases': {
            'rows':   [_fmt_row(r) for r in purchase_rows],
            'amount': round(purchase_amount, 2),
            'tax':    round(purchase_tax, 2),
        },
        'tax_payable': tax_payable,
        'is_refund':   tax_payable < 0,
    })


# ═══════════════════════════════════════════════════════════════════
# Feature 4: 應收/應付帳款 (AR/AP Tracking)
# ═══════════════════════════════════════════════════════════════════


@app.route('/api/finance/payables', methods=['GET'])
@require_module('finance')
def api_payables_list():
    from datetime import date as _d
    ptype  = request.args.get('type', '')      # receivable / payable
    status = request.args.get('status', '')    # open / paid / overdue
    conds, params = ['TRUE'], []
    if ptype:  conds.append("payable_type=%s"); params.append(ptype)
    if status == 'overdue':
        conds.append("status='open' AND due_date < CURRENT_DATE")
    elif status:
        conds.append("status=%s"); params.append(status)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT *, CURRENT_DATE - due_date AS days_overdue
            FROM finance_payables
            WHERE {' AND '.join(conds)}
            ORDER BY
              CASE WHEN status='open' AND due_date < CURRENT_DATE THEN 0
                   WHEN status='open' THEN 1
                   ELSE 2 END,
              due_date
        """, params).fetchall()
    result = []
    for r in rows:
        d = _payable_row(r)
        d['days_overdue'] = int(r['days_overdue']) if r['days_overdue'] is not None else 0
        # Compute effective status
        if d['status'] == 'open' and d.get('due_date') and str(_d.today()) > d['due_date']:
            d['effective_status'] = 'overdue'
        else:
            d['effective_status'] = d['status']
        result.append(d)
    return jsonify(result)

@app.route('/api/finance/payables', methods=['POST'])
@require_module('finance')
def api_payable_create():
    b = request.get_json(force=True)
    if not b.get('title','').strip(): return jsonify({'error': '標題為必填'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO finance_payables
              (payable_type, title, party_name, invoice_no, amount, due_date, status, note)
            VALUES (%s,%s,%s,%s,%s,%s,'open',%s) RETURNING *
        """, (b.get('payable_type','payable'), b['title'].strip(),
              b.get('party_name','').strip(), b.get('invoice_no','').strip(),
              float(b.get('amount',0)), b.get('due_date') or None,
              b.get('note','').strip())).fetchone()
    return jsonify(_payable_row(row)), 201

@app.route('/api/finance/payables/<int:pid>', methods=['PUT'])
@require_module('finance')
def api_payable_update(pid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE finance_payables SET
              payable_type=%s, title=%s, party_name=%s, invoice_no=%s,
              amount=%s, due_date=%s, status=%s,
              paid_date=%s, note=%s, updated_at=NOW()
            WHERE id=%s RETURNING *
        """, (b.get('payable_type','payable'), b.get('title','').strip(),
              b.get('party_name','').strip(), b.get('invoice_no','').strip(),
              float(b.get('amount',0)), b.get('due_date') or None,
              b.get('status','open'), b.get('paid_date') or None,
              b.get('note','').strip(), pid)).fetchone()
    return jsonify(_payable_row(row)) if row else ('', 404)

@app.route('/api/finance/payables/<int:pid>', methods=['DELETE'])
@require_module('finance')
def api_payable_delete(pid):
    with get_db() as conn:
        conn.execute("DELETE FROM finance_payables WHERE id=%s", (pid,))
    return jsonify({'deleted': pid})

@app.route('/api/finance/payables/aging', methods=['GET'])
@require_module('finance')
def api_payables_aging():
    ptype = request.args.get('type', 'payable')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT *,
              CURRENT_DATE - due_date AS days_overdue
            FROM finance_payables
            WHERE payable_type=%s AND status='open'
        """, (ptype,)).fetchall()
    buckets = {'current': 0, 'd1_30': 0, 'd31_60': 0, 'd61_90': 0, 'd90plus': 0}
    bucket_rows = {'current': [], 'd1_30': [], 'd31_60': [], 'd61_90': [], 'd90plus': []}
    for r in rows:
        do = int(r['days_overdue']) if r['days_overdue'] is not None else 0
        d = _payable_row(r)
        d['days_overdue'] = do
        if do <= 0:    k = 'current'
        elif do <= 30: k = 'd1_30'
        elif do <= 60: k = 'd31_60'
        elif do <= 90: k = 'd61_90'
        else:          k = 'd90plus'
        buckets[k]      += float(r['amount'])
        bucket_rows[k].append(d)
    return jsonify({'buckets': buckets, 'rows': bucket_rows, 'type': ptype})


# ═══════════════════════════════════════════════════════════════════
# Feature 5: 預算管理 (Budget vs Actual)
# ═══════════════════════════════════════════════════════════════════

@app.route('/api/finance/budgets', methods=['GET'])
@require_module('finance')
def api_budgets_list():
    year  = request.args.get('year',  '')
    month = request.args.get('month', '')
    if not year or not month:
        from datetime import datetime, timezone, timedelta
        now = datetime.now(timezone(timedelta(hours=8)))
        year, month = str(now.year), str(now.month)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT fb.*, fc.name as category_name, fc.type as category_type, fc.color
            FROM finance_budgets fb
            JOIN finance_categories fc ON fc.id=fb.category_id
            WHERE fb.year=%s AND fb.month=%s
            ORDER BY fc.type, fc.sort_order
        """, (int(year), int(month))).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d['budget_amount'] = float(d['budget_amount'])
        if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
        if d.get('updated_at'): d['updated_at'] = d['updated_at'].isoformat()
        result.append(d)
    return jsonify(result)

@app.route('/api/finance/budgets', methods=['POST'])
@require_module('finance')
def api_budgets_save():
    """Upsert list of budgets: [{category_id, budget_amount}]"""
    b = request.get_json(force=True)
    year  = int(b.get('year',  0))
    month = int(b.get('month', 0))
    items = b.get('items', [])
    if not year or not month: return jsonify({'error': '年月為必填'}), 400
    with get_db() as conn:
        for it in items:
            cid = it.get('category_id')
            amt = float(it.get('budget_amount', 0))
            if cid is None: continue
            if amt == 0:
                conn.execute("DELETE FROM finance_budgets WHERE year=%s AND month=%s AND category_id=%s",
                             (year, month, cid))
            else:
                conn.execute("""
                    INSERT INTO finance_budgets (year, month, category_id, budget_amount, updated_at)
                    VALUES (%s,%s,%s,%s,NOW())
                    ON CONFLICT (year, month, category_id)
                    DO UPDATE SET budget_amount=EXCLUDED.budget_amount, updated_at=NOW()
                """, (year, month, cid, amt))
    return jsonify({'ok': True})

@app.route('/api/finance/budgets/vs-actual', methods=['GET'])
@require_module('finance')
def api_budgets_vs_actual():
    year  = request.args.get('year',  '')
    month = request.args.get('month', '')
    if not year or not month:
        from datetime import datetime, timezone, timedelta
        now = datetime.now(timezone(timedelta(hours=8)))
        year, month = str(now.year), str(now.month)
    period = f"{year}-{str(month).zfill(2)}"
    with get_db() as conn:
        cats = conn.execute("""
            SELECT id, name, type, color FROM finance_categories WHERE active=TRUE ORDER BY type, sort_order
        """).fetchall()
        budgets = conn.execute("""
            SELECT category_id, budget_amount FROM finance_budgets WHERE year=%s AND month=%s
        """, (int(year), int(month))).fetchall()
        _d_s, _d_e = _month_date_range(period)
        actuals = conn.execute("""
            SELECT category_id, SUM(amount) as total
            FROM finance_records
            WHERE record_date >= %s AND record_date < %s
            GROUP BY category_id
        """, (_d_s, _d_e)).fetchall()
    budget_map = {r['category_id']: float(r['budget_amount']) for r in budgets}
    actual_map = {r['category_id']: float(r['total']) for r in actuals}
    result = []
    for c in cats:
        cid = c['id']
        bgt = budget_map.get(cid, 0)
        act = actual_map.get(cid, 0)
        pct = round(act / bgt * 100, 1) if bgt > 0 else None
        result.append({
            'category_id':   cid,
            'category_name': c['name'],
            'category_type': c['type'],
            'color':         c['color'],
            'budget':        bgt,
            'actual':        act,
            'remaining':     round(bgt - act, 2),
            'pct':           pct,
            'over_budget':   bgt > 0 and act > bgt,
        })
    return jsonify({'year': year, 'month': month, 'items': result})


# ═══════════════════════════════════════════════════════════════════
# Feature 6: 薪資費用連動 (Payroll -> Finance)
# ═══════════════════════════════════════════════════════════════════

@app.route('/api/finance/payroll/status', methods=['GET'])
@require_module('finance')
def api_payroll_status():
    """列出各月薪資是否已同步至財務"""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT month,
                   COUNT(*) as total,
                   SUM(CASE WHEN finance_synced THEN 1 ELSE 0 END) as synced,
                   SUM(net_pay) as total_net_pay
            FROM salary_records
            WHERE status IN ('confirmed','draft')
            GROUP BY month ORDER BY month DESC LIMIT 24
        """).fetchall()
    return jsonify([{
        'month':         r['month'],
        'total':         int(r['total']),
        'synced':        int(r['synced']),
        'total_net_pay': float(r['total_net_pay'] or 0),
        'all_synced':    int(r['synced']) == int(r['total']),
    } for r in rows])

@app.route('/api/finance/payroll/sync', methods=['POST'])
@require_module('finance')
def api_payroll_sync():
    """將指定月份已確認薪資寫入財務記錄"""
    b     = request.get_json(force=True)
    month = b.get('month', '')
    if not month: return jsonify({'error': '請提供月份'}), 400
    # Find or create 薪資支出 category
    with get_db() as conn:
        cat = conn.execute("""
            SELECT id FROM finance_categories WHERE name='薪資支出' AND type='expense' LIMIT 1
        """).fetchone()
        if not cat:
            cat = conn.execute("""
                INSERT INTO finance_categories (name,type,color,sort_order)
                VALUES ('薪資支出','expense','#e07b2a',11) RETURNING *
            """).fetchone()
        cat_id = cat['id']

        records = conn.execute("""
            SELECT sr.*, ps.name as staff_name
            FROM salary_records sr
            JOIN punch_staff ps ON ps.id=sr.staff_id
            WHERE sr.month=%s AND sr.finance_synced=FALSE
        """, (month,)).fetchall()

        if not records:
            return jsonify({'created': 0, 'message': '無需同步的薪資記錄'})

        record_date = f"{month}-{str(28).zfill(2)}"  # Use 28th as payroll date
        created = 0
        for sr in records:
            # Main salary entry
            conn.execute("""
                INSERT INTO finance_records
                  (record_date, category_id, type, title, amount, note, created_by)
                VALUES (%s,%s,'expense',%s,%s,%s,'payroll-sync')
            """, (record_date, cat_id,
                  f"{sr['staff_name']} {month} 薪資",
                  float(sr['net_pay']),
                  f"薪資記錄 #{sr['id']}"))
            conn.execute("UPDATE salary_records SET finance_synced=TRUE WHERE id=%s", (sr['id'],))
            created += 1

    return jsonify({'created': created, 'month': month})


# ── Tax -> Finance sync ──────────────────────────────────────────

@app.route('/api/finance/tax/<int:year>/<int:period>/sync', methods=['POST'])
@require_module('finance')
def api_finance_tax_sync(year, period):
    """將應繳/退稅金額建立為財務分錄,流入損益表"""
    if period < 1 or period > 6:
        return jsonify({'error': '期別需為 1-6'}), 400
    m_start = (period - 1) * 2 + 1
    m_end   = m_start + 1
    months  = [f"{year}-{str(m).zfill(2)}" for m in range(m_start, m_end + 1)]
    roc_year = year - 1911
    _d_s = _month_date_range(f"{year}-{m_start:02d}")[0]
    _d_e = _month_date_range(f"{year}-{m_end:02d}")[1]

    with get_db() as conn:
        rows = conn.execute("""
            SELECT type, SUM(tax_amount) as tax_total
            FROM finance_records
            WHERE record_date >= %s AND record_date < %s
              AND tax_amount IS NOT NULL AND tax_amount <> 0
            GROUP BY type
        """, (_d_s, _d_e)).fetchall()

    sales_tax    = sum(float(r['tax_total']) for r in rows if r['type'] == 'income')
    purchase_tax = sum(float(r['tax_total']) for r in rows if r['type'] == 'expense')
    tax_payable  = round(sales_tax - purchase_tax, 2)

    if tax_payable == 0:
        return jsonify({'created': 0, 'message': '稅額為零,無需建立分錄'})

    # Record date = last day of period's last month
    import calendar as _cal
    record_date = f"{year}-{str(m_end).zfill(2)}-{_cal.monthrange(year, m_end)[1]}"
    note = f"銷項稅 ${round(sales_tax,0):,.0f} − 進項稅 ${round(purchase_tax,0):,.0f} = {'應繳' if tax_payable>0 else '退稅'} ${abs(round(tax_payable,0)):,.0f}"
    period_label = f"民國{roc_year}年第{period}期({months[0]}~{months[-1]})"

    created = 0
    with get_db() as conn:
        # 冪等性:同一期別已有 tax-sync 記錄就不重複建立
        existing = conn.execute("""
            SELECT id FROM finance_records
            WHERE created_by='tax-sync' AND record_date=%s
              AND title LIKE %s LIMIT 1
        """, (record_date, f'%{period_label}%')).fetchone()
        if existing:
            return jsonify({'created': 0, 'message': f'該期別({period_label})已建立過分錄,略過.',
                            'tax_payable': tax_payable})

        if tax_payable > 0:
            cat = conn.execute(
                "SELECT id FROM finance_categories WHERE name='稅費' AND type='expense' LIMIT 1"
            ).fetchone()
            if not cat:
                cat = conn.execute("""
                    INSERT INTO finance_categories (name, type, color, sort_order, statement_section)
                    VALUES ('稅費','expense','#8892a4', 99,'operating_expense') RETURNING *
                """).fetchone()
            conn.execute("""
                INSERT INTO finance_records
                  (record_date, category_id, type, title, amount, tax_amount, note, created_by)
                VALUES (%s,%s,'expense',%s,%s,0,%s,'tax-sync')
            """, (record_date, cat['id'],
                  f"應繳營業稅 {period_label}", tax_payable, note))
            created += 1
        else:
            cat = conn.execute(
                "SELECT id FROM finance_categories WHERE name='其他收入' AND type='income' LIMIT 1"
            ).fetchone()
            if not cat:
                cat = conn.execute("""
                    INSERT INTO finance_categories (name, type, color, sort_order, statement_section)
                    VALUES ('其他收入','income','#c8a96e', 99,'other_revenue') RETURNING *
                """).fetchone()
            conn.execute("""
                INSERT INTO finance_records
                  (record_date, category_id, type, title, amount, tax_amount, note, created_by)
                VALUES (%s,%s,'income',%s,%s,0,%s,'tax-sync')
            """, (record_date, cat['id'],
                  f"營業稅退稅 {period_label}", abs(tax_payable), note))
            created += 1

    return jsonify({'created': created, 'tax_payable': tax_payable, 'record_date': record_date})


# ═══════════════════════════════════════════════════════════════════
# LINE Broadcast Helper
# ═══════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════
# Expense Claims 費用報帳申請
# ═══════════════════════════════════════════════════════════════════

def _init_expense_db():
    sqls = [
        """CREATE TABLE IF NOT EXISTS expense_claims (
            id                SERIAL PRIMARY KEY,
            staff_id          INT REFERENCES punch_staff(id) ON DELETE CASCADE,
            title             TEXT NOT NULL,
            amount            NUMERIC(12,2) NOT NULL DEFAULT 0,
            expense_date      DATE NOT NULL,
            category          TEXT DEFAULT '',
            note              TEXT DEFAULT '',
            status            TEXT NOT NULL DEFAULT 'pending',
            document_id       INT REFERENCES finance_documents(id) ON DELETE SET NULL,
            review_note       TEXT DEFAULT '',
            reviewed_by       TEXT DEFAULT '',
            reviewed_at       TIMESTAMPTZ,
            finance_record_id INT REFERENCES finance_records(id) ON DELETE SET NULL,
            created_at        TIMESTAMPTZ DEFAULT NOW()
        )""",
    ]
    for sql in sqls:
        try:
            with get_db() as conn:
                conn.execute(sql)
        except Exception as e:
            print(f"[expense_init] {e}")

_init_expense_db()




# ── Employee endpoints ──────────────────────────────────────────







# ── Leave: medical certificate upload ───────────────────────────





# ── Admin endpoints ─────────────────────────────────────────────





# ═══════════════════════════════════════════════════════════════════════════
# 績效考核模組
# ═══════════════════════════════════════════════════════════════════════════

def _init_performance_db():
    sqls = [
        """CREATE TABLE IF NOT EXISTS performance_templates (
            id          SERIAL PRIMARY KEY,
            name        TEXT NOT NULL,
            description TEXT DEFAULT '',
            period      TEXT DEFAULT 'quarterly',
            items       JSONB DEFAULT '[]',
            active      BOOLEAN DEFAULT TRUE,
            created_at  TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS performance_reviews (
            id              SERIAL PRIMARY KEY,
            staff_id        INT REFERENCES punch_staff(id) ON DELETE CASCADE,
            template_id     INT REFERENCES performance_templates(id) ON DELETE SET NULL,
            period_label    TEXT NOT NULL,
            scores          JSONB DEFAULT '{}',
            total_score     NUMERIC(6,2) DEFAULT 0,
            max_score       NUMERIC(6,2) DEFAULT 100,
            grade           TEXT DEFAULT '',
            comments        TEXT DEFAULT '',
            reviewer        TEXT DEFAULT '',
            salary_adjusted BOOLEAN DEFAULT FALSE,
            salary_delta    NUMERIC(12,2) DEFAULT 0,
            reviewed_at     TIMESTAMPTZ DEFAULT NOW(),
            created_at      TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS performance_config (
            key        TEXT PRIMARY KEY,
            value      JSONB NOT NULL,
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )""",
    ]
    for sql in sqls:
        try:
            with get_db() as conn:
                conn.execute(sql)
        except Exception as e:
            print(f"[perf_init] {e}")

_init_performance_db()


# ── 考核範本 CRUD ───────────────────────────────────────────────





# ── 考核記錄 CRUD ───────────────────────────────────────────────





# ── 員工查自己的考核 ────────────────────────────────────────────



# ── 評級設定 CRUD ───────────────────────────────────────────────




# ═══════════════════════════════════════════════════════════════════════════
# LINE Bot 雙向查詢擴充
# ═══════════════════════════════════════════════════════════════════════════

def _line_query_leave_balance(staff, user_id):
    """查詢員工本年度假期餘額(含待審核天數)"""
    from datetime import date as _dlb
    year = _dlb.today().year
    try:
        with get_db() as conn:
            rows = conn.execute("""
                SELECT lb.total_days, lb.used_days, lt.name AS type_name, lt.id AS type_id
                FROM leave_balances lb
                JOIN leave_types lt ON lt.id=lb.leave_type_id
                WHERE lb.staff_id=%s AND lb.year=%s
                ORDER BY lt.sort_order
            """, (staff['id'], year)).fetchall()
            pending_map = {
                r['leave_type_id']: float(r['pending_days'] or 0)
                for r in conn.execute("""
                    SELECT leave_type_id, SUM(total_days) AS pending_days
                    FROM leave_requests
                    WHERE staff_id=%s AND status='pending'
                      AND EXTRACT(YEAR FROM start_date)=%s
                    GROUP BY leave_type_id
                """, (staff['id'], year)).fetchall()
            }
    except Exception as e:
        _send_line_punch(user_id, f'查詢失敗:{e}')
        return
    if not rows:
        _send_line_punch(user_id, f'📋 {staff["name"]} {year} 年\n尚無假期餘額記錄,請聯絡管理員.')
        return
    lines = [f'📋 {staff["name"]} {year} 年假期餘額']
    for r in rows:
        total   = float(r['total_days'] or 0)
        used    = float(r['used_days']  or 0)
        pending = pending_map.get(r['type_id'], 0.0)
        remain  = total - used
        bar     = '▓' * int(remain) + '░' * max(0, int(total - remain))
        pending_str = f'  ⏳ 待審核 {pending:.1f} 天\n' if pending > 0 else ''
        lines.append(f'\n【{r["type_name"]}】\n  剩餘 {remain:.1f} 天 / 共 {total:.0f} 天\n{pending_str}  {bar}')
    _send_line_punch(user_id, '\n'.join(lines))


def _line_query_salary(staff, user_id):
    """查詢員工最近一筆薪資記錄"""
    try:
        with get_db() as conn:
            row = conn.execute("""
                SELECT month, net_pay, base_salary, allowance_total, deduction_total, status
                FROM salary_records
                WHERE staff_id=%s
                ORDER BY month DESC LIMIT 1
            """, (staff['id'],)).fetchone()
    except Exception as e:
        _send_line_punch(user_id, f'查詢失敗:{e}')
        return
    if not row:
        _send_line_punch(user_id, f'📊 {staff["name"]}\n尚無薪資記錄.')
        return
    status_map = {'draft':'草稿', 'confirmed':'已確認', 'paid':'已發放'}
    _send_line_punch(user_id,
        f'📊 {staff["name"]} {row["month"]} 薪資\n\n'
        f'底薪:NT$ {float(row["base_salary"] or 0):,.0f}\n'
        f'津貼:NT$ {float(row["allowance_total"] or 0):,.0f}\n'
        f'扣除:NT$ {float(row["deduction_total"] or 0):,.0f}\n'
        f'━━━━━━━━━━━━\n'
        f'實領:NT$ {float(row["net_pay"] or 0):,.0f}\n'
        f'狀態:{status_map.get(row["status"], row["status"])}\n\n'
        f'詳細資訊請至員工系統薪資單查看.')


def _line_submit_leave(staff, user_id, text):
    """
    解析並建立請假申請.
    格式:
      請假                                -> Quick Reply 選假別
      請假 假別                            -> Quick Reply 選日期
      請假 假別 DATE                       -> Quick Reply 選時段
      請假 假別 DATE 全天/上午/下午          -> 送出
      請假 假別 DATE HH:MM                 -> Quick Reply 選結束時間
      請假 假別 DATE HH:MM HH:MM           -> 送出(指定時間)
      請假 假別 DATE1 DATE2               -> 送出(多天)
    """
    import re as _re_lv
    from datetime import date as _dlv, timedelta as _tdlv
    WDAY_LV = ['一', '二', '三', '四', '五', '六', '日']
    TIME_PAT = _re_lv.compile(r'^\d{2}:\d{2}$')
    DATE_PAT = _re_lv.compile(r'^\d{4}-\d{2}-\d{2}$')
    parts = text.strip().split()
    # parts[0] = '請假'

    # Step 1: only "請假" -> Quick Reply with leave types + remaining balance
    if len(parts) == 1:
        year = _dlv.today().year
        with get_db() as conn:
            types = conn.execute(
                "SELECT id, name FROM leave_types WHERE active=TRUE ORDER BY sort_order"
            ).fetchall()
            balances = {
                r['leave_type_id']: (float(r['total_days'] or 0) - float(r['used_days'] or 0))
                for r in conn.execute("""
                    SELECT leave_type_id, total_days, used_days FROM leave_balances
                    WHERE staff_id=%s AND year=%s
                """, (staff['id'], year)).fetchall()
            }
        if not types:
            _send_line_punch(user_id, '目前無可用假別,請聯絡管理員.')
            return
        lines = ['🌿 請假申請\n\n可用假別(剩餘天數):']
        items = []
        for r in types:
            rem = balances.get(r['id'])
            rem_str = f' {rem:.1f}天' if rem is not None else ''
            lines.append(f'• {r["name"]}{rem_str}')
            items.append({'label': f'{r["name"]}{rem_str}', 'text': f'請假 {r["name"]}'})
        lines.append('\n請點下方按鈕選擇假別:')
        _send_line_with_quick_reply(user_id, '\n'.join(lines), items[:13])
        return

    # Step 2: "請假 假別" (no date) -> Quick Reply with date options
    if len(parts) == 2:
        leave_type_name = parts[1]
        today = _dlv.today()
        end7  = today + _tdlv(days=13)
        with get_db() as conn:
            sched7 = _get_staff_scheduled_dates(conn, staff['id'],
                                                today.isoformat(), end7.isoformat()) or set()
        date_items = []
        for i in range(14):
            d = today + _tdlv(days=i)
            if d.weekday() >= 5 and d.isoformat() not in sched7:
                continue
            label = ('今天 ' if i == 0 else '明天 ' if i == 1 else '') + f'{d.strftime("%m/%d")}({WDAY_LV[d.weekday()]})'
            date_items.append({'label': label, 'text': f'請假 {leave_type_name} {d.isoformat()}'})
            if len(date_items) == 6:
                break
        _send_line_with_quick_reply(user_id,
            f'🌿 請假 · {leave_type_name}\n\n請選擇日期,或手動輸入多天:\n'
            f'請假 {leave_type_name} 開始日 結束日',
            date_items)
        return

    # Step 2.5: "請假 假別 DATE" -> Quick Reply: 全天/上午半天/下午半天/指定時間
    if len(parts) == 3 and DATE_PAT.match(parts[2]):
        leave_type_name = parts[1]
        date_str = parts[2]
        items_period = [
            {'label': '全天',     'text': f'請假 {leave_type_name} {date_str} 全天'},
            {'label': '上午半天', 'text': f'請假 {leave_type_name} {date_str} 上午'},
            {'label': '下午半天', 'text': f'請假 {leave_type_name} {date_str} 下午'},
            {'label': '指定時間', 'text': f'請假 {leave_type_name} {date_str} 指定時間'},
        ]
        _send_line_with_quick_reply(user_id,
            f'🌿 請假 · {leave_type_name}\n日期:{date_str}\n\n請選擇時段:',
            items_period)
        return

    # Step 2.6: "請假 假別 DATE 指定時間" -> Quick Reply: 選開始時間
    if len(parts) == 4 and DATE_PAT.match(parts[2]) and parts[3] == '指定時間':
        leave_type_name = parts[1]
        date_str = parts[2]
        start_opts = ['07:00','08:00','09:00','10:00','11:00','12:00','13:00','14:00']
        items = [{'label': t, 'text': f'請假 {leave_type_name} {date_str} {t}'} for t in start_opts]
        _send_line_with_quick_reply(user_id,
            f'🌿 請假 · {leave_type_name}\n日期:{date_str}\n\n請選擇開始時間:', items)
        return

    # Step 2.7: "請假 假別 DATE HH:MM" -> Quick Reply: 選結束時間
    if len(parts) == 4 and DATE_PAT.match(parts[2]) and TIME_PAT.match(parts[3]):
        leave_type_name = parts[1]
        date_str = parts[2]
        start_str = parts[3]
        sh, sm = map(int, start_str.split(':'))
        end_opts = []
        for delta_h in [0.5, 1, 1.5, 2, 2.5, 3, 4, 5, 6, 7, 8]:
            total_m = sh * 60 + sm + int(delta_h * 60)
            if total_m > 23 * 60 + 59:
                break
            eh, em = total_m // 60, total_m % 60
            end_opts.append((f'{eh:02d}:{em:02d}', delta_h))
        items = [{'label': f'至 {t}({d:.1g}h)', 'text': f'請假 {leave_type_name} {date_str} {start_str} {t}'}
                 for t, d in end_opts[:13]]
        _send_line_with_quick_reply(user_id,
            f'🌿 請假 · {leave_type_name}\n日期:{date_str}　開始:{start_str}\n\n請選擇結束時間:', items)
        return

    leave_type_name = parts[1]
    date_str1 = parts[2]

    # --- Time-based leave: 請假 假別 DATE HH:MM HH:MM ---
    start_time_val = ''; end_time_val = ''
    if len(parts) >= 5 and DATE_PAT.match(date_str1) and TIME_PAT.match(parts[3]) and TIME_PAT.match(parts[4]):
        start_str = parts[3]; end_str = parts[4]
        try:
            sh, sm = map(int, start_str.split(':'))
            eh, em = map(int, end_str.split(':'))
            hours = ((eh * 60 + em) - (sh * 60 + sm)) / 60
            if hours <= 0:
                _send_line_punch(user_id, '結束時間必須晚於開始時間,請重新選擇.')
                return
        except ValueError:
            _send_line_punch(user_id, '時間格式錯誤,請使用 HH:MM.')
            return
        # 每8小時 = 1天,四捨五入至0.5天
        days = max(0.5, round(hours / 8 * 2) / 2)
        start_time_val = start_str; end_time_val = end_str
        date_str2 = date_str1
        start_half = False; end_half = False
        reason = '(LINE 請假)'

        with get_db() as conn:
            lt = conn.execute(
                "SELECT * FROM leave_types WHERE active=TRUE AND name=%s", (leave_type_name,)
            ).fetchone()
            if not lt:
                lt = conn.execute(
                    "SELECT * FROM leave_types WHERE active=TRUE AND name ILIKE %s LIMIT 1",
                    (f'%{leave_type_name}%',)
                ).fetchone()
            if not lt:
                avail = conn.execute(
                    "SELECT name FROM leave_types WHERE active=TRUE ORDER BY sort_order"
                ).fetchall()
                _send_line_punch(user_id, f'找不到假別「{leave_type_name}」\n可用:{"、".join(r["name"] for r in avail)}')
                return
            year = date_str1[:4]
            bal = conn.execute("""
                SELECT total_days, used_days FROM leave_balances
                WHERE staff_id=%s AND leave_type_id=%s AND year=%s
            """, (staff['id'], lt['id'], int(year))).fetchone()
            remain = None
            if bal:
                remain = float(bal['total_days'] or 0) - float(bal['used_days'] or 0)
                if remain < days:
                    _send_line_punch(user_id,
                        f'⚠️ {lt["name"]} 餘額不足\n剩餘 {remain:.1f} 天,申請 {days} 天')
                    return
            row = conn.execute("""
                INSERT INTO leave_requests
                  (staff_id, leave_type_id, start_date, end_date, total_days,
                   start_half, end_half, reason, status, start_time, end_time, created_at)
                VALUES (%s,%s,%s,%s,%s,FALSE,FALSE,%s,'pending',%s,%s,NOW()) RETURNING id
            """, (staff['id'], lt['id'], date_str1, date_str1, days,
                  reason, start_time_val, end_time_val)).fetchone()
        bal_str = f'(剩餘 {remain:.1f} 天)' if remain is not None else ''
        _send_line_punch(user_id,
            f'✅ 請假申請已送出\n\n'
            f'假別:{lt["name"]} {bal_str}\n'
            f'日期:{date_str1}\n'
            f'時段:{start_str} ~ {end_str}({hours:.1f} 小時)\n'
            f'天數:{days} 天\n\n'
            f'申請號:#{row["id"]},等待管理員審核.')
        return

    # --- Day-based leave (全天 / 上午 / 下午 / multi-day) ---
    # Detect period token (全天/上午/下午) or second date
    start_half = False; end_half = False
    period_token = None
    if len(parts) > 3:
        tok = parts[3].strip()
        if tok in ('全天', '上午', '下午'):
            period_token = tok
            date_str2 = date_str1
        elif DATE_PAT.match(tok):
            date_str2 = tok
        else:
            date_str2 = date_str1
    else:
        date_str2 = date_str1

    if period_token == '上午':
        start_half = True; end_half = True
    elif period_token == '下午':
        start_half = False; end_half = True

    reason = '(LINE 請假)'

    # Validate dates
    try:
        _dlv.fromisoformat(date_str1)
        _dlv.fromisoformat(date_str2)
    except ValueError:
        _send_line_punch(user_id, f'日期格式錯誤,請使用 YYYY-MM-DD,例:{_dlv.today().isoformat()}')
        return

    # Find leave type (fuzzy: exact or contains)
    with get_db() as conn:
        lt = conn.execute(
            "SELECT * FROM leave_types WHERE active=TRUE AND name=%s", (leave_type_name,)
        ).fetchone()
        if not lt:
            lt = conn.execute(
                "SELECT * FROM leave_types WHERE active=TRUE AND name ILIKE %s LIMIT 1",
                (f'%{leave_type_name}%',)
            ).fetchone()
        if not lt:
            avail = conn.execute(
                "SELECT name FROM leave_types WHERE active=TRUE ORDER BY sort_order"
            ).fetchall()
            names = '、'.join(r['name'] for r in avail)
            _send_line_punch(user_id, f'找不到假別「{leave_type_name}」\n\n可用假別:{names}')
            return

        # Check leave balance
        year = date_str1[:4]
        bal = conn.execute("""
            SELECT total_days, used_days FROM leave_balances
            WHERE staff_id=%s AND leave_type_id=%s AND year=%s
        """, (staff['id'], lt['id'], int(year))).fetchone()

        # Calculate requested days (排班日為準;無排班則排除週六日); half day = 0.5
        sched_lv = _get_staff_scheduled_dates(conn, staff['id'], date_str1, date_str2)
        days = _calc_leave_days(date_str1, date_str2, start_half, end_half,
                                scheduled_dates=sched_lv)

        remain = None
        if bal:
            remain = float(bal['total_days'] or 0) - float(bal['used_days'] or 0)
            if remain < days:
                _send_line_punch(user_id,
                    f'⚠️ {lt["name"]} 餘額不足\n剩餘 {remain:.1f} 天,申請 {days} 天\n\n'
                    f'請至員工系統調整後再申請.')
                return

        # Create leave request
        row = conn.execute("""
            INSERT INTO leave_requests
              (staff_id, leave_type_id, start_date, end_date, total_days,
               start_half, end_half, reason, status, start_time, end_time, created_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'pending','','',NOW()) RETURNING id
        """, (staff['id'], lt['id'], date_str1, date_str2, days,
              start_half, end_half, reason)).fetchone()

    period_label = '(上午半天)' if (start_half and end_half and date_str1 == date_str2) else \
                   '(下午半天)' if (end_half and not start_half and date_str1 == date_str2) else ''
    bal_str = f'(剩餘 {remain:.1f} 天)' if remain is not None else ''
    _send_line_punch(user_id,
        f'✅ 請假申請已送出\n\n'
        f'假別:{lt["name"]} {bal_str}\n'
        f'日期:{date_str1}' + (f' ~ {date_str2}' if date_str2 != date_str1 else '') +
        f'{period_label}\n'
        f'天數:{days} 天\n\n'
        f'申請號:#{row["id"]},等待管理員審核.')


def _line_query_performance(staff, user_id):
    """查詢員工最近一次績效考核"""
    try:
        with get_db() as conn:
            row = conn.execute("""
                SELECT pr.period_label, pr.grade, pr.total_score, pr.max_score,
                       pr.comments, pr.salary_adjusted, pr.salary_delta,
                       pr.reviewed_at, pt.name AS tpl_name
                FROM performance_reviews pr
                LEFT JOIN performance_templates pt ON pt.id=pr.template_id
                WHERE pr.staff_id=%s
                ORDER BY pr.reviewed_at DESC LIMIT 1
            """, (staff['id'],)).fetchone()
    except Exception as e:
        _send_line_punch(user_id, f'查詢失敗:{e}')
        return
    if not row:
        _send_line_punch(user_id, f'{staff["name"]}\n尚無績效考核記錄.')
        return
    grade_label = _grade_labels()
    pct = float(row['total_score']) / float(row['max_score']) * 100 if row['max_score'] else 0
    adj = f"\n薪資調整:NT$ {float(row['salary_delta']):+,.0f}" if row['salary_adjusted'] else ''
    reviewed = str(row['reviewed_at'])[:10] if row['reviewed_at'] else ''
    _send_line_punch(user_id,
        f'{staff["name"]} 最近考核\n\n'
        f'期間:{row["period_label"]}\n'
        f'範本:{row["tpl_name"] or "-"}\n'
        f'得分:{float(row["total_score"]):.1f} / {float(row["max_score"]):.0f}({pct:.0f}%)\n'
        f'評級:{row["grade"]} {grade_label.get(row["grade"],"")}'
        f'{adj}\n'
        + (f'備注:{row["comments"][:60]}\n' if row['comments'] else '')
        + f'考核日:{reviewed}')


def _line_query_monthly_records(staff, user_id, text):
    """查詢員工月出勤記錄與打卡明細.
    格式:出勤紀錄 [YYYY-MM](省略月份則查本月)
    """
    import re as _rem
    from datetime import date as _dm, timezone as _tzm, timedelta as _tdm, datetime as _dtm
    TW = _tzm(_tdm(hours=8))

    # 解析月份
    parts = text.strip().split()
    month = None
    if len(parts) >= 2:
        m = _rem.match(r'^(\d{4})-(\d{1,2})$', parts[1])
        if m:
            month = f"{m.group(1)}-{m.group(2).zfill(2)}"
    if not month:
        month = _dtm.now(TW).strftime('%Y-%m')

    try:
        _lb_ts_s, _lb_ts_e = _month_ts_range(month)
        with get_db() as conn:
            rows = conn.execute("""
                SELECT punch_type, punched_at, is_manual
                FROM punch_records
                WHERE staff_id=%s
                  AND punched_at >= %s AND punched_at < %s
                ORDER BY punched_at ASC
            """, (staff['id'], _lb_ts_s, _lb_ts_e)).fetchall()
    except Exception as e:
        _send_line_punch(user_id, f'查詢失敗:{e}')
        return

    if not rows:
        _send_line_punch(user_id, f'📋 {staff["name"]} {month}\n該月尚無打卡記錄.')
        return

    WDAY = ['一', '二', '三', '四', '五', '六', '日']

    # 依日期分組
    days = {}
    for r in rows:
        pa = r['punched_at']
        if pa.tzinfo is None:
            from datetime import timezone as _utzm
            pa = pa.replace(tzinfo=_utzm.utc)
        pa_tw = pa.astimezone(TW)
        ds = pa_tw.strftime('%Y-%m-%d')
        if ds not in days:
            days[ds] = []
        days[ds].append({'type': r['punch_type'], 'time': pa_tw.strftime('%H:%M'), 'manual': bool(r['is_manual'])})

    total_mins = 0
    anomaly_days = 0
    lines = []

    for ds in sorted(days.keys()):
        recs = days[ds]
        d = _dm.fromisoformat(ds)
        wday = WDAY[d.weekday()]

        clock_in  = next((r['time'] for r in recs if r['type'] == 'in'),  None)
        clock_out = next((r['time'] for r in recs if r['type'] == 'out'), None)
        has_manual = any(r['manual'] for r in recs)

        if clock_in and clock_out:
            ci = _dtm.strptime(f'{ds} {clock_in}',  '%Y-%m-%d %H:%M')
            co = _dtm.strptime(f'{ds} {clock_out}', '%Y-%m-%d %H:%M')
            mins = max(0, int((co - ci).total_seconds() / 60))
            total_mins += mins
            h, m = divmod(mins, 60)
            dur = f'{h}h{m:02d}' if m else f'{h}h'
        elif clock_in:
            dur = '⚠️缺下班'
            anomaly_days += 1
        else:
            dur = '⚠️缺上班'
            anomaly_days += 1

        manual_mark = '【補】' if has_manual else ''
        ci_str = clock_in  or '--:--'
        co_str = clock_out or '--:--'
        lines.append(f'{ds[5:]}({wday}) {ci_str}↑{co_str}↓ {dur}{manual_mark}')

    th, tm = divmod(total_mins, 60)
    total_str = f'{th}h{tm:02d}' if tm else f'{th}h'
    anomaly_str = f'|異常 {anomaly_days} 天' if anomaly_days else ''
    header = (f'📋 {staff["name"]} {month} 出勤\n'
              f'出勤 {len(days)} 天|工時 {total_str}{anomaly_str}\n'
              + '─' * 20)

    # 訊息過長時分批送出(LINE 單則上限約 5000 字)
    full = header + '\n' + '\n'.join(lines)
    if len(full) <= 4500:
        _send_line_punch(user_id, full)
    else:
        _send_line_punch(user_id, header)
        chunk, chunk_len = [], 0
        for line in lines:
            if chunk_len + len(line) + 1 > 1800:
                _send_line_punch(user_id, '\n'.join(chunk))
                chunk, chunk_len = [], 0
            chunk.append(line)
            chunk_len += len(line) + 1
        if chunk:
            _send_line_punch(user_id, '\n'.join(chunk))


def _line_overtime_start(staff, user_id):
    """加班 button -> Quick Reply with date options."""
    from datetime import date as _dot_s, timedelta as _tdot_s
    WDAY_OT = ['一', '二', '三', '四', '五', '六', '日']
    today = _dot_s.today()
    items = []
    for i in range(-1, 5):
        d = today + _tdot_s(days=i)
        label = ('昨天 ' if i == -1 else '今天 ' if i == 0 else '明天 ' if i == 1 else '') + \
                f'{d.strftime("%m/%d")}({WDAY_OT[d.weekday()]})'
        items.append({'label': label, 'text': f'申請加班 {d.isoformat()}'})
    _send_line_with_quick_reply(user_id, '⏰ 加班申請\n\n請選擇加班日期:', items)


def _line_submit_overtime(staff, user_id, text):
    """
    LINE 加班申請流程(幾點到幾點):
      申請加班 DATE           -> Quick Reply 選開始時間
      申請加班 DATE HH:MM     -> Quick Reply 選結束時間
      申請加班 DATE HH:MM HH:MM -> 送出申請
    """
    import re as _re_ot
    from datetime import date as _dot, datetime as _dtt
    parts = text.strip().split(None, 3)

    if len(parts) < 2:
        _line_overtime_start(staff, user_id)
        return

    date_str = parts[1]
    try:
        _dot.fromisoformat(date_str)
    except ValueError:
        _send_line_punch(user_id, f'日期格式錯誤,請使用 YYYY-MM-DD,例:{_dot.today().isoformat()}')
        return

    # Step 2: date only -> select start time
    if len(parts) == 2:
        start_options = ['08:00','09:00','17:00','18:00','19:00','20:00','21:00','22:00']
        items = [{'label': t, 'text': f'申請加班 {date_str} {t}'} for t in start_options]
        _send_line_with_quick_reply(user_id,
            f'⏰ 加班申請 · {date_str}\n\n請選擇開始時間:', items)
        return

    start_str = parts[2]
    if not _re_ot.match(r'^\d{2}:\d{2}$', start_str):
        _send_line_punch(user_id, '時間格式錯誤,請使用 HH:MM,例:18:00')
        return

    # Step 3: date + start time -> select end time
    if len(parts) == 3:
        sh, sm = map(int, start_str.split(':'))
        end_options = []
        for delta_h in [1, 1.5, 2, 2.5, 3, 4, 5, 6]:
            total_m = sh * 60 + sm + int(delta_h * 60)
            eh, em = (total_m // 60) % 24, total_m % 60
            end_options.append(f'{eh:02d}:{em:02d}')
        items = [{'label': f'至 {t}(+{d}h)', 'text': f'申請加班 {date_str} {start_str} {t}'}
                 for t, d in zip(end_options, [1, 1.5, 2, 2.5, 3, 4, 5, 6])]
        _send_line_with_quick_reply(user_id,
            f'⏰ 加班申請 · {date_str} {start_str} 開始\n\n請選擇結束時間:', items)
        return

    # Step 4: date + start + end -> submit
    end_str = parts[3].strip().split()[0]  # take only first token (HH:MM)
    if not _re_ot.match(r'^\d{2}:\d{2}$', end_str):
        _send_line_punch(user_id, '時間格式錯誤,請使用 HH:MM,例:20:00')
        return

    try:
        sh, sm = map(int, start_str.split(':'))
        eh, em = map(int, end_str.split(':'))
        hours = ((eh * 60 + em) - (sh * 60 + sm)) / 60
        if hours <= 0:
            hours += 24  # crosses midnight
        if hours <= 0 or hours > 24:
            raise ValueError
    except ValueError:
        _send_line_punch(user_id, '時間計算錯誤,請重新選擇.')
        return

    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO overtime_requests
              (staff_id, request_date, start_time, end_time, ot_hours, reason, status)
            VALUES (%s, %s, %s, %s, %s, %s, 'pending')
            RETURNING id
        """, (staff['id'], date_str, start_str, end_str, round(hours, 2), '(LINE 加班申請)')).fetchone()

    _send_line_punch(user_id,
        f'✅ 加班申請已送出\n\n'
        f'日期:{date_str}\n'
        f'時段:{start_str} ~ {end_str}({hours:.1f} 小時)\n'
        f'申請編號:#{row["id"]}\n\n'
        '請等候管理員審核,審核結果將通知您.')


def _line_show_help(staff, user_id):
    _send_line_punch(user_id,
        f'哈囉 {staff["name"]}!以下是可用的指令:\n\n'
        '─── 打卡 ───\n'
        '📍 傳送位置 -> 自動打卡\n'
        '💬 上班 / 下班 / 休息 / 回來\n'
        '📋 狀態 -> 今日打卡記錄\n\n'
        '─── 查詢 ───\n'
        '🌿 查餘假 -> 本年假期餘額\n'
        '💰 查薪資 -> 最近薪資單\n'
        '📊 出勤紀錄 -> 本月出勤明細\n'
        '   出勤紀錄 2026-03 -> 指定月份\n'
        '考核 -> 最近績效考核\n\n'
        '─── 申請 ───\n'
        '📝 請假 [假別] [日期] -> 送出請假\n'
        '   範例:請假 特休 2026-04-01\n'
        '⏰ 加班 -> 選擇日期與時段申請加班\n'
        '🗂️ 假別 -> 查看可用假別清單\n\n'
        '─── 其他 ───\n'
        '🔓 解除綁定')


def _line_show_leave_types(staff, user_id):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT name, max_days FROM leave_types WHERE active=TRUE ORDER BY sort_order"
        ).fetchall()
    if not rows:
        _send_line_punch(user_id, '目前無可用假別.'); return
    lines = ['🗂️ 可用假別清單\n']
    for r in rows:
        limit = f'(年限 {r["max_days"]} 天)' if r['max_days'] else ''
        lines.append(f'• {r["name"]} {limit}')
    lines.append('\n申請方式:請假 [假別] [日期]')
    _send_line_punch(user_id, '\n'.join(lines))



# ── Blueprints(模組化路由)──
from blueprints.attendance import bp as attendance_bp
app.register_blueprint(attendance_bp)
from blueprints.dashboard import bp as dashboard_bp
app.register_blueprint(dashboard_bp)
from blueprints.shifts import bp as shifts_bp
app.register_blueprint(shifts_bp)
from blueprints.schedule import bp as schedule_bp
app.register_blueprint(schedule_bp)
from blueprints.salary import bp as salary_bp
app.register_blueprint(salary_bp)
from blueprints.leave import bp as leave_bp
app.register_blueprint(leave_bp)
from blueprints.overtime import bp as overtime_bp
app.register_blueprint(overtime_bp)
from blueprints.stores import bp as stores_bp
app.register_blueprint(stores_bp)
from blueprints.holidays import bp as holidays_bp
app.register_blueprint(holidays_bp)
from blueprints.expense import bp as expense_bp
app.register_blueprint(expense_bp)
from blueprints.performance import bp as performance_bp
app.register_blueprint(performance_bp)
from blueprints.training import bp as training_bp
app.register_blueprint(training_bp)
from blueprints.mobile import bp as mobile_bp
app.register_blueprint(mobile_bp)
from blueprints.webauthn import bp as webauthn_bp
app.register_blueprint(webauthn_bp)
from blueprints.announcements import bp as announcements_bp
app.register_blueprint(announcements_bp)
