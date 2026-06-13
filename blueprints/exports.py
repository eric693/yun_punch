"""exports API blueprint(自 app.py 拆出)。"""
import csv
from services import _get_finance_settings
from flask import Blueprint, request, Response
from db import (
    get_db,
)
from auth import (
    login_required, require_module,
)
from utils import (
    _month_ts_range, _month_date_range,
)
import json as _json
import io

bp = Blueprint('exports', __name__)


@bp.route('/api/export/attendance', methods=['GET'])
@login_required
def api_export_attendance():
    """匯出月度出勤明細 CSV"""
    month    = request.args.get('month', '')
    staff_id = request.args.get('staff_id', '')
    if not month:
        from datetime import date as _de
        month = _de.today().strftime('%Y-%m')

    _ts_s, _ts_e = _month_ts_range(month)
    conds, params = ["pr.punched_at >= %s AND pr.punched_at < %s"], [_ts_s, _ts_e]
    if staff_id:
        conds.append("pr.staff_id=%s"); params.append(int(staff_id))

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT
                ps.employee_code,
                ps.name as staff_name,
                ps.department,
                ps.role,
                (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                pr.punch_type,
                to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei', 'HH24:MI') as punch_time,
                pr.is_manual,
                pr.manual_by,
                pr.gps_distance,
                pr.location_name,
                pr.note
            FROM punch_records pr
            JOIN punch_staff ps ON ps.id = pr.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY ps.name, pr.punched_at
        """, params).fetchall()

    PUNCH_LABEL = {'in':'上班打卡','out':'下班打卡','break_out':'休息開始','break_in':'休息結束'}

    output = io.StringIO()
    output.write('\ufeff')  # UTF-8 BOM for Excel
    writer = csv.writer(output)
    writer.writerow(['員工代碼','姓名','部門','職稱','日期','打卡類型','時間','補打卡','操作人','GPS距離(m)','地點','備註'])

    for r in rows:
        writer.writerow([
            r['employee_code'] or '',
            r['staff_name'],
            r['department']    or '',
            r['role']          or '',
            str(r['work_date']),
            PUNCH_LABEL.get(r['punch_type'], r['punch_type']),
            r['punch_time'],
            '是' if r['is_manual'] else '',
            r['manual_by']     or '',
            r['gps_distance']  if r['gps_distance'] is not None else '',
            r['location_name'] or '',
            r['note']          or '',
        ])

    csv_content = output.getvalue()
    from flask import Response
    return Response(
        csv_content.encode('utf-8-sig'),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename=attendance_{month}.csv'}
    )

@bp.route('/api/export/attendance-summary', methods=['GET'])
@login_required
def api_export_attendance_summary():
    """匯出月度出勤摘要 CSV(每人每天工時)"""
    month = request.args.get('month', '')
    if not month:
        from datetime import date as _df
        month = _df.today().strftime('%Y-%m')

    _ts_s, _ts_e = _month_ts_range(month)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT
                ps.employee_code,
                ps.name,
                ps.department,
                ps.role,
                (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                MIN(CASE WHEN pr.punch_type='in'  THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_in,
                MAX(CASE WHEN pr.punch_type='out' THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_out,
                MIN(CASE WHEN pr.punch_type='in'  THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as ci_ts,
                MAX(CASE WHEN pr.punch_type='out' THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as co_ts,
                BOOL_OR(pr.is_manual) as has_manual,
                COUNT(*) as punch_count
            FROM punch_records pr
            JOIN punch_staff ps ON ps.id = pr.staff_id
            WHERE pr.punched_at >= %s AND pr.punched_at < %s
            GROUP BY ps.employee_code, ps.name, ps.department, ps.role,
                     (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY ps.name, (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
        """, (_ts_s, _ts_e)).fetchall()

    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output)
    writer.writerow(['員工代碼','姓名','部門','職稱','日期','上班','下班','工時(h)','打卡次數','含補打'])

    for r in rows:
        dur_h = ''
        if r['ci_ts'] and r['co_ts']:
            from datetime import datetime as _dtx
            try:
                ci = r['ci_ts'] if hasattr(r['ci_ts'], 'timestamp') else _dtx.fromisoformat(str(r['ci_ts']))
                co = r['co_ts'] if hasattr(r['co_ts'], 'timestamp') else _dtx.fromisoformat(str(r['co_ts']))
                dur_h = round((co - ci).total_seconds() / 3600, 2)
            except Exception:
                pass
        writer.writerow([
            r['employee_code'] or '',
            r['name'], r['department'] or '', r['role'] or '',
            str(r['work_date']),
            r['clock_in'] or '', r['clock_out'] or '',
            dur_h,
            r['punch_count'],
            '是' if r['has_manual'] else '',
        ])

    from flask import Response
    return Response(
        output.getvalue().encode('utf-8-sig'),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename=attendance_summary_{month}.csv'}
    )

@bp.route('/api/export/salary', methods=['GET'])
@login_required
def api_export_salary():
    """匯出月度薪資明細 CSV"""
    month = request.args.get('month', '')
    if not month:
        from datetime import date as _dg
        month = _dg.today().strftime('%Y-%m')

    with get_db() as conn:
        rows = conn.execute("""
            SELECT sr.*, ps.name as staff_name, ps.employee_code,
                   ps.department, ps.role, ps.salary_type
            FROM salary_records sr
            JOIN punch_staff ps ON ps.id = sr.staff_id
            WHERE sr.month = %s
            ORDER BY ps.name
        """, (month,)).fetchall()

    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output)
    writer.writerow([
        '員工代碼','姓名','部門','職稱','薪資制度',
        '工作日','出勤天數','請假天數','無薪假天數',
        '津貼合計','扣除合計','加班費','實領金額','狀態','備註'
    ])

    for r in rows:
        items = r['items'] if isinstance(r['items'], list) else _json.loads(r['items'] or '[]')
        sal_type = r['salary_type'] or 'monthly'
        writer.writerow([
            r['employee_code'] or '', r['staff_name'],
            r['department'] or '', r['role'] or '',
            '時薪制' if sal_type == 'hourly' else '月薪制',
            float(r['work_days'] or 0), float(r['actual_days'] or 0),
            float(r['leave_days'] or 0), float(r['unpaid_days'] or 0),
            float(r['allowance_total'] or 0), float(r['deduction_total'] or 0),
            float(r['ot_pay'] or 0), float(r['net_pay'] or 0),
            '已確認' if r['status'] == 'confirmed' else '草稿',
            r['note'] or '',
        ])

    from flask import Response
    return Response(
        output.getvalue().encode('utf-8-sig'),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename=salary_{month}.csv'}
    )

@bp.route('/api/export/leave', methods=['GET'])
@login_required
def api_export_leave():
    """匯出請假記錄 CSV"""
    month    = request.args.get('month', '')
    year     = request.args.get('year',  '')
    staff_id = request.args.get('staff_id', '')

    conds, params = ['lr.status=%s'], ['approved']
    if month:
        _d_s, _d_e = _month_date_range(month)
        conds.append('lr.start_date >= %s AND lr.start_date < %s'); params.extend([_d_s, _d_e])
    if year:  conds.append("EXTRACT(YEAR FROM lr.start_date)=%s"); params.append(int(year))
    if staff_id: conds.append("lr.staff_id=%s"); params.append(int(staff_id))

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT lr.*, ps.name as staff_name, ps.employee_code,
                   ps.department, lt.name as leave_type_name, lt.pay_rate
            FROM leave_requests lr
            JOIN punch_staff ps ON ps.id = lr.staff_id
            JOIN leave_types  lt ON lt.id = lr.leave_type_id
            WHERE {' AND '.join(conds)}
            ORDER BY lr.start_date, ps.name
        """, params).fetchall()

    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output)
    writer.writerow(['員工代碼','姓名','部門','假別','薪資倍率','開始日期','結束日期','天數','原因','代理人','狀態'])

    PAY_LABEL = {1.0:'全薪', 0.5:'半薪', 0.0:'無薪'}
    for r in rows:
        writer.writerow([
            r['employee_code'] or '', r['staff_name'], r['department'] or '',
            r['leave_type_name'], PAY_LABEL.get(float(r['pay_rate']), f"{r['pay_rate']}倍"),
            str(r['start_date']), str(r['end_date']),
            float(r['total_days']),
            r['reason'] or '', r['substitute_name'] or '',
            {'approved':'已核准','rejected':'已退回','pending':'待審核'}.get(r['status'], r['status']),
        ])

    from flask import Response
    return Response(
        output.getvalue().encode('utf-8-sig'),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename=leave_{month or year or "all"}.csv'}
    )

@bp.route('/api/export/withholding', methods=['GET'])
@require_module('salary')
def api_export_withholding():
    """年度薪資所得扣繳憑單(所得類別50)"""
    from datetime import date as _dwh
    year   = request.args.get('year', str(_dwh.today().year))
    fmt    = request.args.get('format', 'html')

    fs = _get_finance_settings()
    company_name   = fs.get('company_name', '')
    company_tax_id = fs.get('company_tax_id', '')
    company_address= fs.get('company_address', '')

    with get_db() as conn:
        rows = conn.execute("""
            SELECT ps.id, ps.name, ps.national_id, ps.address,
                   COALESCE(SUM(sr.allowance_total), 0)       AS gross_salary,
                   COALESCE(SUM(sr.income_tax_withheld), 0)   AS tax_withheld,
                   COALESCE(AVG(sr.insured_salary), 0)        AS avg_insured
            FROM salary_records sr
            JOIN punch_staff ps ON ps.id = sr.staff_id
            WHERE sr.month LIKE %s AND sr.status='confirmed'
            GROUP BY ps.id, ps.name, ps.national_id, ps.address
            ORDER BY ps.name
        """, (f'{year}-%',)).fetchall()

    # 計算二代健保補充費
    def supp_nhi(gross, insured):
        base = float(gross) - float(insured) * 12
        return max(0, round(base * 0.0211, 0)) if base > 0 else 0

    data = []
    for i, r in enumerate(rows, 1):
        gross = float(r['gross_salary'])
        insured = float(r['avg_insured'])
        data.append({
            'no':          i,
            'name':        r['name'],
            'national_id': r['national_id'] or '-',
            'address':     r['address'] or '-',
            'gross':       gross,
            'supp_nhi':    supp_nhi(gross, insured),
            'tax':         float(r['tax_withheld']),
        })

    if fmt == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO
        wb = wb2 = openpyxl.Workbook()
        ws = wb.active; ws.title = f'{year}年扣繳憑單'
        hfill = PatternFill('solid', fgColor='0F1C3A')
        thin  = Border(*[Side(style='thin', color='DDDDDD')]*4)
        hdrs  = ['序號','姓名','身分證字號','地址','年度薪資合計','二代健保補充費','扣繳稅額']
        ws.append(hdrs)
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(1, ci); c.font = Font(bold=True, color='FFFFFF', size=10); c.fill = hfill
            c.alignment = Alignment(horizontal='center', vertical='center'); c.border = thin
        ws.column_dimensions['A'].width = 5; ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 14; ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 16; ws.column_dimensions['F'].width = 16; ws.column_dimensions['G'].width = 12
        for d in data:
            ws.append([d['no'], d['name'], d['national_id'], d['address'],
                       d['gross'], d['supp_nhi'], d['tax']])
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        from flask import Response as _FR2
        return _FR2(buf.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=withholding_{year}.xlsx'})

    # HTML printable
    rows_html = ''.join(f"""
      <tr>
        <td style="text-align:center">{d['no']}</td>
        <td>{d['name']}</td>
        <td style="font-family:monospace">{d['national_id']}</td>
        <td style="font-size:11px">{d['address']}</td>
        <td style="text-align:right;font-family:monospace">{d['gross']:,.0f}</td>
        <td style="text-align:right;font-family:monospace">{d['supp_nhi']:,.0f}</td>
        <td style="text-align:right;font-family:monospace">{d['tax']:,.0f}</td>
      </tr>""" for d in data)
    html = f"""<!DOCTYPE html><html lang="zh-TW"><head>
<meta charset="UTF-8"><title>{year}年度薪資扣繳憑單</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Noto Sans TC',sans-serif;font-size:12px;padding:20px;color:#1e2a45}}
h2{{font-size:16px;font-weight:700;margin-bottom:4px}}
.meta{{font-size:11px;color:#666;margin-bottom:16px}}
table{{width:100%;border-collapse:collapse;margin-bottom:20px}}
th{{background:#0f1c3a;color:#fff;padding:7px 10px;font-size:11px;font-weight:600;text-align:left}}
td{{padding:6px 10px;border-bottom:1px solid #eee;font-size:12px}}
tr:nth-child(even){{background:#f8f9fb}}
.note{{font-size:10px;color:#888;border-top:1px solid #ddd;padding-top:8px}}
@media print{{button{{display:none}}}}
</style></head><body>
<button onclick="window.print()" style="margin-bottom:16px;padding:6px 16px;background:#0f1c3a;color:#fff;border:none;border-radius:4px;cursor:pointer;font-size:12px">列印</button>
<h2>{year} 年度薪資所得扣繳憑單(所得類別 50)</h2>
<div class="meta">扣繳義務人:{company_name}　統一編號:{company_tax_id}　地址:{company_address}　製表日期:{_dwh.today().isoformat()}</div>
<table>
<thead><tr><th>#</th><th>員工姓名</th><th>身分證字號</th><th>地址</th><th>年度薪資合計(元)</th><th>二代健保補充費(元)</th><th>扣繳稅額(元)</th></tr></thead>
<tbody>{rows_html}</tbody>
</table>
<div class="note">※ 本報表依薪資紀錄計算,二代健保補充費 = 超出投保薪資部分 × 2.11%.扣繳稅額請依各月薪資記錄人工確認.</div>
</body></html>"""
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}
