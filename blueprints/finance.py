"""finance API blueprint(自 app.py 拆出)。"""
from flask import Blueprint, request, jsonify, session
from db import (
    get_db,
)
from auth import (
    require_module,
)
from config import (
    TW_TZ, ANTHROPIC_API_KEY,
)
from services import _get_finance_settings
from utils import (
    _month_date_range, _roc_year, _month_last_day,
)
from serializers import (
    _finance_cat_row, _finance_rec_row, _recurring_row, _bank_row, _payable_row,
)
from datetime import datetime as _dt
import json as _json
import datetime

bp = Blueprint('finance', __name__)


@bp.route('/api/finance/categories', methods=['GET'])
@require_module('finance')
def api_finance_categories_list():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM finance_categories ORDER BY sort_order, id").fetchall()
    return jsonify([_finance_cat_row(r) for r in rows])

@bp.route('/api/finance/categories', methods=['POST'])
@require_module('finance')
def api_finance_category_create():
    b = request.get_json(force=True)
    if not b.get('name','').strip(): return jsonify({'error': '名稱為必填'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO finance_categories (name,type,color,sort_order,active,statement_section)
            VALUES (%s,%s,%s,%s,%s,%s) RETURNING *
        """, (b['name'].strip(), b.get('type','expense'), b.get('color','#4a7bda'),
              int(b.get('sort_order',0)), bool(b.get('active',True)),
              b.get('statement_section') or ('operating_revenue' if b.get('type')=='income' else 'operating_expense')
             )).fetchone()
    return jsonify(_finance_cat_row(row)), 201

@bp.route('/api/finance/categories/<int:cid>', methods=['PUT'])
@require_module('finance')
def api_finance_category_update(cid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE finance_categories SET name=%s,type=%s,color=%s,sort_order=%s,active=%s,statement_section=%s
            WHERE id=%s RETURNING *
        """, (b.get('name','').strip(), b.get('type','expense'), b.get('color','#4a7bda'),
              int(b.get('sort_order',0)), bool(b.get('active',True)),
              b.get('statement_section') or ('operating_revenue' if b.get('type')=='income' else 'operating_expense'),
              cid)).fetchone()
    return jsonify(_finance_cat_row(row)) if row else ('', 404)

@bp.route('/api/finance/categories/<int:cid>', methods=['DELETE'])
@require_module('finance')
def api_finance_category_delete(cid):
    with get_db() as conn:
        conn.execute("DELETE FROM finance_categories WHERE id=%s", (cid,))
    return jsonify({'deleted': cid})

@bp.route('/api/finance/records', methods=['GET'])
@require_module('finance')
def api_finance_records_list():
    month  = request.args.get('month', '')
    ftype  = request.args.get('type', '')
    cat_id = request.args.get('category_id', '')
    conds, params = ['TRUE'], []
    if month:
        _d_s, _d_e = _month_date_range(month)
        conds.append('fr.record_date >= %s AND fr.record_date < %s'); params.extend([_d_s, _d_e])
    if ftype:
        conds.append("fr.type=%s"); params.append(ftype)
    if cat_id:
        conds.append("fr.category_id=%s"); params.append(int(cat_id))
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT fr.*, fc.name as category_name, fc.color as category_color,
                   fd.filename as doc_filename, fd.ocr_raw as ocr_raw
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            LEFT JOIN finance_documents fd ON fd.id=fr.document_id
            WHERE {' AND '.join(conds)}
            ORDER BY fr.record_date DESC, fr.id DESC
        """, params).fetchall()
    result = []
    for r in rows:
        d = _finance_rec_row(r)
        d['category_name']  = r['category_name']
        d['category_color'] = r['category_color']
        d['doc_filename']   = r['doc_filename']
        d['ocr_raw']        = r['ocr_raw'] if r['ocr_raw'] else None
        result.append(d)
    return jsonify(result)

@bp.route('/api/finance/documents', methods=['GET'])
@require_module('finance')
def api_finance_documents_list():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT fd.*,
                   COUNT(fr.id) as linked_count,
                   MAX(fr.title) as linked_title,
                   MAX(fr.id) as linked_record_id
            FROM finance_documents fd
            LEFT JOIN finance_records fr ON fr.document_id = fd.id
            GROUP BY fd.id
            ORDER BY fd.created_at DESC
        """).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        if d.get('upload_date'): d['upload_date'] = str(d['upload_date'])
        if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
        d['linked_count'] = int(d['linked_count'] or 0)
        result.append(d)
    return jsonify(result)

@bp.route('/api/finance/records', methods=['POST'])
@require_module('finance')
def api_finance_record_create():
    b = request.get_json(force=True)
    if not b.get('title','').strip(): return jsonify({'error': '標題為必填'}), 400
    if not b.get('record_date'):      return jsonify({'error': '日期為必填'}), 400
    if float(b.get('amount', 0) or 0) <= 0: return jsonify({'error': '金額必須大於0'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO finance_records
              (record_date, category_id, type, title, amount, tax_amount, vendor, invoice_no, note, document_id, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *
        """, (b['record_date'], b.get('category_id') or None, b.get('type','expense'),
              b['title'].strip(), float(b.get('amount',0)), float(b.get('tax_amount',0)),
              b.get('vendor','').strip(), b.get('invoice_no','').strip(),
              b.get('note','').strip(), b.get('document_id') or None,
              session.get('admin_display_name',''))).fetchone()
    return jsonify(_finance_rec_row(row)), 201

@bp.route('/api/finance/records/<int:rid>', methods=['PUT'])
@require_module('finance')
def api_finance_record_update(rid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE finance_records SET
              record_date=%s, category_id=%s, type=%s, title=%s, amount=%s,
              tax_amount=%s, vendor=%s, invoice_no=%s, note=%s, updated_at=NOW()
            WHERE id=%s RETURNING *
        """, (b['record_date'], b.get('category_id') or None, b.get('type','expense'),
              b.get('title','').strip(), float(b.get('amount',0)), float(b.get('tax_amount',0)),
              b.get('vendor','').strip(), b.get('invoice_no','').strip(),
              b.get('note','').strip(), rid)).fetchone()
    return jsonify(_finance_rec_row(row)) if row else ('', 404)

@bp.route('/api/finance/records/<int:rid>', methods=['DELETE'])
@require_module('finance')
def api_finance_record_delete(rid):
    with get_db() as conn:
        conn.execute("DELETE FROM finance_records WHERE id=%s", (rid,))
    return jsonify({'deleted': rid})

@bp.route('/api/finance/summary/<year>/<month>', methods=['GET'])
@require_module('finance')
def api_finance_summary(year, month):
    period = f"{year}-{month.zfill(2)}"
    with get_db() as conn:
        _fin_d_s, _fin_d_e = _month_date_range(period)
        totals = conn.execute("""
            SELECT type, COALESCE(SUM(amount),0) as total
            FROM finance_records
            WHERE record_date >= %s AND record_date < %s
            GROUP BY type
        """, (_fin_d_s, _fin_d_e)).fetchall()
        income  = next((float(r['total']) for r in totals if r['type']=='income'), 0.0)
        expense = next((float(r['total']) for r in totals if r['type']=='expense'), 0.0)

        by_cat = conn.execute("""
            SELECT fc.name, fc.color, fr.type, COALESCE(SUM(fr.amount),0) as total
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            WHERE fr.record_date >= %s AND fr.record_date < %s
            GROUP BY fc.name, fc.color, fr.type
            ORDER BY total DESC
        """, (_fin_d_s, _fin_d_e)).fetchall()

        # Last 6 months trend
        trend = conn.execute("""
            SELECT to_char(record_date,'YYYY-MM') as mon,
                   type, COALESCE(SUM(amount),0) as total
            FROM finance_records
            WHERE record_date >= (DATE_TRUNC('month', %s::date) - INTERVAL '5 months')
              AND record_date <  (DATE_TRUNC('month', %s::date) + INTERVAL '1 month')
            GROUP BY to_char(record_date,'YYYY-MM'), type
            ORDER BY mon
        """, (f"{period}-01", f"{period}-01")).fetchall()

    return jsonify({
        'income':  income,
        'expense': expense,
        'net':     income - expense,
        'by_category': [
            {'name': r['name'] or '未分類', 'color': r['color'] or '#8892a4',
             'type': r['type'], 'total': float(r['total'])}
            for r in by_cat
        ],
        'trend': [
            {'month': r['mon'], 'type': r['type'], 'total': float(r['total'])}
            for r in trend
        ],
    })

@bp.route('/api/finance/ocr', methods=['POST'])
@require_module('finance')
def api_finance_ocr():
    import anthropic as _ant
    import base64, re as _re

    if not ANTHROPIC_API_KEY:
        return jsonify({'error': '尚未設定 ANTHROPIC_API_KEY 環境變數'}), 500

    file = request.files.get('file')
    if not file:
        return jsonify({'error': '請上傳圖片或 PDF 檔案'}), 400

    raw = file.read()
    media_type = file.content_type or 'image/jpeg'
    # Only image types supported by Claude vision
    if media_type not in ('image/jpeg','image/png','image/gif','image/webp'):
        media_type = 'image/jpeg'

    img_b64 = base64.standard_b64encode(raw).decode()

    client = _ant.Anthropic(api_key=ANTHROPIC_API_KEY)
    try:
        msg = client.messages.create(
            model='claude-sonnet-4-6',
            max_tokens=1024,
            messages=[{
                'role': 'user',
                'content': [
                    {'type': 'image', 'source': {'type': 'base64', 'media_type': media_type, 'data': img_b64}},
                    {'type': 'text', 'text': (
                        '請辨識此文件,以JSON格式回傳以下欄位(找不到的欄位填null):\n'
                        '{"date":"YYYY-MM-DD","vendor":"廠商名稱","invoice_no":"發票或單據號碼",'
                        '"total_amount":含稅總金額數字,"tax_amount":稅額數字,"pre_tax_amount":未稅金額數字,'
                        '"doc_type":"invoice或receipt或expense之一",'
                        '"title":"建議記帳標題(簡短)",'
                        '"items":[{"name":"品項","qty":數量,"unit_price":單價,"amount":小計}],'
                        '"currency":"TWD"}\n只回傳JSON,不要其他文字或markdown.'
                    )}
                ]
            }]
        )
        text = msg.content[0].text.strip()
        text = _re.sub(r'^```json\s*', '', text, flags=_re.MULTILINE)
        text = _re.sub(r'\s*```$', '', text, flags=_re.MULTILINE)
        result = _json.loads(text)
    except _json.JSONDecodeError:
        result = {'raw_text': text, 'error': 'OCR 回傳格式無法解析'}
    except Exception as e:
        return jsonify({'error': f'OCR 失敗:{str(e)}'}), 500

    try:
        with get_db() as conn:
            doc = conn.execute("""
                INSERT INTO finance_documents (filename, doc_type, ocr_raw, upload_date)
                VALUES (%s,%s,%s,CURRENT_DATE) RETURNING id
            """, (file.filename, result.get('doc_type',''), _json.dumps(result))).fetchone()
        result['document_id'] = doc['id']
    except Exception as e:
        print(f"[finance_ocr doc save] {e}")

    return jsonify(result)

@bp.route('/api/finance/export', methods=['GET'])
@require_module('finance')
def api_finance_export():
    import csv, io as _io
    month = request.args.get('month', '')
    conds, params = ['TRUE'], []
    if month:
        _d_s, _d_e = _month_date_range(month)
        conds.append('fr.record_date >= %s AND fr.record_date < %s'); params.extend([_d_s, _d_e])
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT fr.record_date, fr.type, fr.title, fr.amount, fr.tax_amount,
                   fr.vendor, fr.invoice_no, fr.note, fc.name as category_name
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            WHERE {' AND '.join(conds)}
            ORDER BY fr.record_date, fr.id
        """, params).fetchall()
    out = _io.StringIO()
    w = csv.writer(out)
    w.writerow(['日期','類型','類別','標題','金額','稅額','廠商','單據號碼','備註'])
    for r in rows:
        w.writerow([str(r['record_date']), '收入' if r['type']=='income' else '支出',
                    r['category_name'] or '', r['title'], r['amount'], r['tax_amount'] or 0,
                    r['vendor'] or '', r['invoice_no'] or '', r['note'] or ''])
    fname = f"finance_{month or 'all'}.csv"
    return (('\ufeff' + out.getvalue()).encode('utf-8'),
            200, {'Content-Type': 'text/csv; charset=utf-8',
                  'Content-Disposition': f'attachment; filename={fname}'})

@bp.route('/api/finance/settings', methods=['GET'])
@require_module('finance')
def api_finance_settings_get():
    with get_db() as conn:
        rows = conn.execute("SELECT setting_key, setting_value FROM finance_settings").fetchall()
    return jsonify({r['setting_key']: r['setting_value'] for r in rows})

@bp.route('/api/finance/settings', methods=['POST'])
@require_module('finance')
def api_finance_settings_save():
    data = request.get_json(force=True)
    with get_db() as conn:
        for k, v in data.items():
            conn.execute(
                "INSERT INTO finance_settings (setting_key, setting_value) VALUES (%s,%s) ON CONFLICT (setting_key) DO UPDATE SET setting_value=EXCLUDED.setting_value",
                (k, str(v))
            )
    return jsonify({'ok': True})

def _compute_statements(year, month):
    """Compute all three financial statements for the given year/month."""
    from collections import defaultdict
    period = f"{year}-{str(month).zfill(2)}"
    settings = _get_finance_settings()
    opening_cash   = float(settings.get('opening_cash',   0) or 0)
    opening_equity = float(settings.get('opening_equity', 0) or 0)
    company_name   = settings.get('company_name', '') or '公司名稱'

    with get_db() as conn:
        records = conn.execute("""
            SELECT fr.type, fr.amount,
                   fc.name            AS cat_name,
                   fc.statement_section AS section
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id = fr.category_id
            WHERE fr.record_date >= %s AND fr.record_date < %s
        """, _month_date_range(period)).fetchall()

        # Cumulative net before this month (for balance sheet period-start)
        prev = conn.execute("""
            SELECT
                COALESCE(SUM(CASE WHEN type='income'  THEN amount ELSE 0 END), 0) AS cum_income,
                COALESCE(SUM(CASE WHEN type='expense' THEN amount ELSE 0 END), 0) AS cum_expense
            FROM finance_records
            WHERE record_date < DATE_TRUNC('month', %s::date)
        """, (f"{period}-01",)).fetchone()

    cum_net_before = float(prev['cum_income']) - float(prev['cum_expense'])

    by_section = defaultdict(float)
    by_cat     = defaultdict(float)
    for r in records:
        sec = r['section'] or ('operating_revenue' if r['type'] == 'income' else 'operating_expense')
        by_section[sec]               += float(r['amount'])
        by_cat[(sec, r['cat_name'] or '未分類')] += float(r['amount'])

    operating_revenue = by_section['operating_revenue']
    other_revenue     = by_section['other_revenue']
    cogs              = by_section['cogs']
    operating_expense = by_section['operating_expense']
    other_expense     = by_section['other_expense']

    gross_profit      = operating_revenue - cogs
    operating_income  = gross_profit - operating_expense
    net_income        = operating_income + other_revenue - other_expense
    total_income      = operating_revenue + other_revenue
    total_expense     = cogs + operating_expense + other_expense

    cum_net_total     = cum_net_before + net_income
    cash_balance      = opening_cash + opening_equity + cum_net_total
    # 期初現金亦屬業主投入(權益),須計入權益側,資產負債表才會平衡
    contributed_capital = opening_cash + opening_equity
    total_equity      = contributed_capital + cum_net_total

    def cat_lines(section):
        return [{'name': k[1], 'amount': round(v, 2)}
                for k, v in sorted(by_cat.items()) if k[0] == section]

    return {
        'company_name': company_name,
        'year': int(year), 'month': int(month),
        'roc_year': _roc_year(year),
        'last_day': _month_last_day(year, month),
        'income_statement': {
            'operating_revenue':       round(operating_revenue, 2),
            'operating_revenue_lines': cat_lines('operating_revenue'),
            'other_revenue':           round(other_revenue, 2),
            'other_revenue_lines':     cat_lines('other_revenue'),
            'cogs':                    round(cogs, 2),
            'cogs_lines':              cat_lines('cogs'),
            'gross_profit':            round(gross_profit, 2),
            'operating_expense':       round(operating_expense, 2),
            'operating_expense_lines': cat_lines('operating_expense'),
            'operating_income':        round(operating_income, 2),
            'other_expense':           round(other_expense, 2),
            'other_expense_lines':     cat_lines('other_expense'),
            'net_income':              round(net_income, 2),
        },
        'balance_sheet': {
            'cash':                    round(cash_balance, 2),
            'total_assets':            round(cash_balance, 2),
            'total_liabilities':       0,
            'opening_equity':          round(contributed_capital, 2),
            'retained_earnings':       round(cum_net_total, 2),
            'total_equity':            round(total_equity, 2),
            'total_liabilities_equity': round(total_equity, 2),
        },
        'cash_flow': {
            'operating_inflow':        round(total_income, 2),
            'operating_inflow_lines':  cat_lines('operating_revenue') + cat_lines('other_revenue'),
            'operating_outflow':       round(total_expense, 2),
            'operating_outflow_lines': cat_lines('cogs') + cat_lines('operating_expense') + cat_lines('other_expense'),
            'operating_net':           round(total_income - total_expense, 2),
            'investing_net':           0,
            'financing_net':           0,
            'net_change':              round(total_income - total_expense, 2),
            'opening_cash':            round(opening_cash + opening_equity + cum_net_before, 2),
            'closing_cash':            round(cash_balance, 2),
        },
    }

@bp.route('/api/finance/statements/<year>/<month>', methods=['GET'])
@require_module('finance')
def api_finance_statements(year, month):
    try:
        return jsonify(_compute_statements(year, month))
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@bp.route('/api/finance/export/statements/<year>/<month>', methods=['GET'])
@require_module('finance')
def api_finance_export_statements(year, month):
    import openpyxl, io as _io
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    d  = _compute_statements(year, month)
    co = d['company_name']
    ry = d['roc_year']
    m  = d['month']
    ld = d['last_day']
    IS = d['income_statement']
    BS = d['balance_sheet']
    CF = d['cash_flow']

    wb = openpyxl.Workbook()

    NAVY   = '1C3557'
    AMT    = '#,##0'
    FONT   = '標楷體'
    thin   = Side(style='thin')
    medium = Side(style='medium')

    def _border(top=False, bottom=False, dbl=False):
        return Border(
            top    = thin   if top else None,
            bottom = medium if dbl  else (thin if bottom else None),
        )

    def setup_ws(ws, title, date_str):
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 4
        ws.column_dimensions['C'].width = 18
        for row_vals, styles in [
            ([co, '', ''],        {'font': Font(FONT, bold=True, size=14), 'align': 'center', 'merge': True}),
            ([title, '', ''],     {'font': Font(FONT, bold=True, size=13), 'align': 'center', 'merge': True}),
            ([date_str, '', ''],  {'font': Font(FONT, size=11),            'align': 'center', 'merge': True}),
        ]:
            ws.append(row_vals)
            r = ws.max_row
            ws.cell(r, 1).font      = styles['font']
            ws.cell(r, 1).alignment = Alignment(horizontal=styles['align'])
            if styles.get('merge'):
                ws.merge_cells(f'A{r}:C{r}')
        ws.append([])  # blank

        # column headers
        ws.append(['項　　目', '', '金　額(元)'])
        r = ws.max_row
        ws.cell(r, 1).font      = Font(FONT, bold=True, size=11, color='FFFFFF')
        ws.cell(r, 3).font      = Font(FONT, bold=True, size=11, color='FFFFFF')
        ws.cell(r, 1).fill      = PatternFill('solid', fgColor=NAVY)
        ws.cell(r, 3).fill      = PatternFill('solid', fgColor=NAVY)
        ws.cell(r, 2).fill      = PatternFill('solid', fgColor=NAVY)
        ws.cell(r, 1).alignment = Alignment(horizontal='center')
        ws.cell(r, 3).alignment = Alignment(horizontal='right')

    def row(ws, label, amount=None, indent=0, bold=False, subtotal=False, total=False, dbl=False):
        prefix = '　' * indent
        ws.append([prefix + label, '', amount])
        r = ws.max_row
        b = bold or subtotal or total
        ws.cell(r, 1).font      = Font(FONT, bold=b, size=11)
        ws.cell(r, 1).alignment = Alignment(horizontal='left')
        if amount is not None:
            ws.cell(r, 3).font           = Font(FONT, bold=b, size=11)
            ws.cell(r, 3).number_format  = AMT
            ws.cell(r, 3).alignment      = Alignment(horizontal='right')
        if subtotal or total:
            ws.cell(r, 3).border = _border(top=(subtotal or total), bottom=(subtotal or total), dbl=dbl)
        elif amount is None:
            ws.cell(r, 1).font = Font(FONT, bold=True, size=11)

    # ─── 損益表 ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = '損益表'
    setup_ws(ws1, '損益表', f'中華民國{ry}年{m}月份')

    row(ws1, '一、營業收入', bold=True)
    for l in IS['operating_revenue_lines']:
        row(ws1, l['name'], l['amount'], indent=2)
    row(ws1, '營業收入合計', IS['operating_revenue'], indent=1, subtotal=True)
    ws1.append([])

    row(ws1, '二、營業成本', bold=True)
    for l in IS['cogs_lines']:
        row(ws1, l['name'], l['amount'], indent=2)
    row(ws1, '營業成本合計', IS['cogs'], indent=1, subtotal=True)
    ws1.append([])

    row(ws1, '毛　利', IS['gross_profit'], indent=1, total=True)
    ws1.append([])

    row(ws1, '三、營業費用', bold=True)
    for l in IS['operating_expense_lines']:
        row(ws1, l['name'], l['amount'], indent=2)
    row(ws1, '營業費用合計', IS['operating_expense'], indent=1, subtotal=True)
    ws1.append([])

    row(ws1, '營業利益(損失)', IS['operating_income'], indent=1, total=True)
    ws1.append([])

    if IS['other_revenue'] or IS['other_revenue_lines']:
        row(ws1, '四、營業外收入', bold=True)
        for l in IS['other_revenue_lines']:
            row(ws1, l['name'], l['amount'], indent=2)
        row(ws1, '營業外收入合計', IS['other_revenue'], indent=1, subtotal=True)
        ws1.append([])

    if IS['other_expense'] or IS['other_expense_lines']:
        row(ws1, '五、營業外費用', bold=True)
        for l in IS['other_expense_lines']:
            row(ws1, l['name'], l['amount'], indent=2)
        row(ws1, '營業外費用合計', IS['other_expense'], indent=1, subtotal=True)
        ws1.append([])

    row(ws1, '本期淨利(損)', IS['net_income'], bold=True, total=True, dbl=True)

    # ─── 資產負債表 ───────────────────────────────────────────────
    ws2 = wb.create_sheet('資產負債表')
    setup_ws(ws2, '資產負債表', f'中華民國{ry}年{m}月{ld}日')

    row(ws2, '【資　產】', bold=True)
    row(ws2, '流動資產', indent=1, bold=True)
    row(ws2, '現金及約當現金', BS['cash'], indent=2)
    row(ws2, '資產合計', BS['total_assets'], indent=1, total=True)
    ws2.append([])

    row(ws2, '【負　債】', bold=True)
    row(ws2, '流動負債', indent=1, bold=True)
    row(ws2, '應付帳款', 0, indent=2)
    row(ws2, '負債合計', BS['total_liabilities'], indent=1, total=True)
    ws2.append([])

    row(ws2, '【股東權益】', bold=True)
    row(ws2, '資本額', BS['opening_equity'], indent=2)
    row(ws2, '保留盈餘', BS['retained_earnings'], indent=2)
    row(ws2, '股東權益合計', BS['total_equity'], indent=1, total=True)
    ws2.append([])

    row(ws2, '負債及股東權益合計', BS['total_liabilities_equity'], bold=True, total=True, dbl=True)

    # ─── 現金流量表 ───────────────────────────────────────────────
    ws3 = wb.create_sheet('現金流量表')
    setup_ws(ws3, '現金流量表(直接法)', f'中華民國{ry}年{m}月份')

    row(ws3, '一、營業活動之現金流量', bold=True)
    row(ws3, '(一)收現收入', indent=1, bold=True)
    for l in CF['operating_inflow_lines']:
        row(ws3, l['name'], l['amount'], indent=3)
    row(ws3, '收現合計', CF['operating_inflow'], indent=2, subtotal=True)
    ws3.append([])
    row(ws3, '(二)付現費用', indent=1, bold=True)
    for l in CF['operating_outflow_lines']:
        row(ws3, l['name'], -l['amount'], indent=3)
    row(ws3, '付現合計', -CF['operating_outflow'], indent=2, subtotal=True)
    ws3.append([])
    row(ws3, '營業活動淨現金流量', CF['operating_net'], indent=1, total=True)
    ws3.append([])

    row(ws3, '二、投資活動之現金流量', bold=True)
    row(ws3, '投資活動淨現金流量', CF['investing_net'], indent=1, total=True)
    ws3.append([])

    row(ws3, '三、籌資活動之現金流量', bold=True)
    row(ws3, '籌資活動淨現金流量', CF['financing_net'], indent=1, total=True)
    ws3.append([])

    row(ws3, '四、本期現金增減', CF['net_change'], bold=True, total=True)
    row(ws3, '五、期初現金及約當現金', CF['opening_cash'], bold=True)
    row(ws3, '六、期末現金及約當現金', CF['closing_cash'], bold=True, total=True, dbl=True)

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"statements_{year}{str(month).zfill(2)}.xlsx"
    return (buf.read(), 200, {
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Content-Disposition': f'attachment; filename="{fname}"',
    })

@bp.route('/api/finance/recurring', methods=['GET'])
@require_module('finance')
def api_recurring_list():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT fr.*, fc.name as category_name
            FROM finance_recurring fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            ORDER BY fr.active DESC, fr.id
        """).fetchall()
    result = []
    for r in rows:
        d = _recurring_row(r)
        d['category_name'] = r['category_name']
        result.append(d)
    return jsonify(result)

@bp.route('/api/finance/recurring', methods=['POST'])
@require_module('finance')
def api_recurring_create():
    b = request.get_json(force=True)
    if not b.get('title','').strip(): return jsonify({'error': '標題為必填'}), 400
    if not b.get('start_date'):       return jsonify({'error': '開始日期為必填'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO finance_recurring
              (title, type, category_id, amount, tax_amount, vendor, note,
               frequency, day_of_month, start_date, end_date, active)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,TRUE) RETURNING *
        """, (b['title'].strip(), b.get('type','expense'), b.get('category_id') or None,
              float(b.get('amount',0)), float(b.get('tax_amount',0)),
              b.get('vendor','').strip(), b.get('note','').strip(),
              b.get('frequency','monthly'), int(b.get('day_of_month',1) or 1),
              b['start_date'], b.get('end_date') or None)).fetchone()
    return jsonify(_recurring_row(row)), 201

@bp.route('/api/finance/recurring/<int:rid>', methods=['PUT'])
@require_module('finance')
def api_recurring_update(rid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE finance_recurring SET
              title=%s, type=%s, category_id=%s, amount=%s, tax_amount=%s,
              vendor=%s, note=%s, frequency=%s, day_of_month=%s,
              start_date=%s, end_date=%s, active=%s
            WHERE id=%s RETURNING *
        """, (b.get('title','').strip(), b.get('type','expense'), b.get('category_id') or None,
              float(b.get('amount',0)), float(b.get('tax_amount',0)),
              b.get('vendor','').strip(), b.get('note','').strip(),
              b.get('frequency','monthly'), int(b.get('day_of_month',1) or 1),
              b.get('start_date'), b.get('end_date') or None,
              bool(b.get('active', True)), rid)).fetchone()
    return jsonify(_recurring_row(row)) if row else ('', 404)

@bp.route('/api/finance/recurring/<int:rid>', methods=['DELETE'])
@require_module('finance')
def api_recurring_delete(rid):
    with get_db() as conn:
        conn.execute("DELETE FROM finance_recurring WHERE id=%s", (rid,))
    return jsonify({'deleted': rid})

@bp.route('/api/finance/recurring/generate', methods=['POST'])
@require_module('finance')
def api_recurring_generate():
    """為指定月份產生定期分錄(冪等:已產生則跳過)"""
    from datetime import date as _d, timedelta as _td
    import calendar as _cal
    b = request.get_json(force=True)
    month = b.get('month', '')  # YYYY-MM
    if not month:
        from datetime import datetime, timezone, timedelta
        month = datetime.now(timezone(timedelta(hours=8))).strftime('%Y-%m')
    y, m = int(month[:4]), int(month[5:])
    days_in_month = _cal.monthrange(y, m)[1]

    created, skipped = 0, 0
    with get_db() as conn:
        rows = conn.execute("""
            SELECT * FROM finance_recurring
            WHERE active=TRUE
              AND start_date <= %s
              AND (end_date IS NULL OR end_date >= %s)
        """, (f"{month}-28", f"{month}-01")).fetchall()

        for r in rows:
            # Check already generated this month
            if r['last_generated'] == month:
                skipped += 1
                continue
            # Check frequency
            freq = r['frequency']
            start_m = r['start_date'].month
            if freq == 'quarterly' and (m - start_m) % 3 != 0:
                skipped += 1
                continue
            if freq == 'yearly' and m != start_m:
                skipped += 1
                continue
            # Determine record date
            day = min(int(r['day_of_month'] or 1), days_in_month)
            rec_date = _d(y, m, day)
            conn.execute("""
                INSERT INTO finance_records
                  (record_date, category_id, type, title, amount, tax_amount, vendor, note, created_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'auto-recurring')
            """, (rec_date, r['category_id'], r['type'], r['title'],
                  r['amount'], r['tax_amount'] or 0, r['vendor'] or '', r['note'] or ''))
            conn.execute("UPDATE finance_recurring SET last_generated=%s WHERE id=%s",
                         (month, r['id']))
            created += 1

    return jsonify({'created': created, 'skipped': skipped, 'month': month})

@bp.route('/api/finance/bank/import', methods=['POST'])
@require_module('finance')
def api_bank_import():
    """匯入銀行對帳單 CSV"""
    import csv, io as _io
    file = request.files.get('file')
    if not file: return jsonify({'error': '請上傳 CSV 檔案'}), 400
    raw = file.read().decode('utf-8-sig', errors='replace')
    account_name = request.form.get('account_name', '').strip() or '銀行帳戶'
    import_batch = _dt.now(TW_TZ).strftime('%Y%m%d%H%M%S')

    reader = csv.reader(_io.StringIO(raw))
    rows_data = [r for r in reader if any(c.strip() for c in r)]
    if not rows_data: return jsonify({'error': 'CSV 無資料'}), 400

    # Auto-detect header row (skip rows where first column is not a date-like string)
    def _is_date(s):
        s = s.strip().replace('/', '-').replace('.', '-')
        for fmt in ('%Y-%m-%d','%Y-%m-%d','%m-%d-%Y','%d-%m-%Y'):
            try: _dt.strptime(s, fmt); return True
            except: pass
        # ROC year: 民國 e.g. 113/01/15
        import re
        if re.match(r'^\d{2,3}[/-]\d{1,2}[/-]\d{1,2}$', s):
            parts = re.split(r'[/-]', s)
            if int(parts[0]) < 200: return True
        return False

    def _parse_date(s):
        import re
        s = s.strip()
        parts = re.split(r'[/\-\.]', s)
        if len(parts) == 3:
            if int(parts[0]) < 200:
                # ROC year
                y2 = int(parts[0]) + 1911
                return f"{y2}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
            if int(parts[0]) > 31:
                return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
            if int(parts[2]) > 31:
                return f"{parts[2]}-{parts[0].zfill(2)}-{parts[1].zfill(2)}"
        return None

    def _parse_amount(s):
        import re
        s = re.sub(r'[,$\s]', '', str(s).strip())
        try: return float(s)
        except: return None

    inserted = 0
    with get_db() as conn:
        for row in rows_data:
            if len(row) < 2: continue
            date_str = _parse_date(row[0])
            if not date_str: continue
            desc = row[1].strip() if len(row) > 1 else ''
            # Format: date, desc, debit, credit  OR  date, desc, amount
            if len(row) >= 4:
                debit  = _parse_amount(row[2])
                credit = _parse_amount(row[3])
                if debit and debit > 0:
                    conn.execute("""INSERT INTO bank_statements
                        (account_name,txn_date,amount,txn_type,description,import_batch)
                        VALUES (%s,%s,%s,'debit',%s,%s)
                    """, (account_name, date_str, debit, desc, import_batch))
                    inserted += 1
                if credit and credit > 0:
                    conn.execute("""INSERT INTO bank_statements
                        (account_name,txn_date,amount,txn_type,description,import_batch)
                        VALUES (%s,%s,%s,'credit',%s,%s)
                    """, (account_name, date_str, credit, desc, import_batch))
                    inserted += 1
            elif len(row) >= 3:
                amt = _parse_amount(row[2])
                if amt is not None and amt != 0:
                    txn_type = 'credit' if amt > 0 else 'debit'
                    conn.execute("""INSERT INTO bank_statements
                        (account_name,txn_date,amount,txn_type,description,import_batch)
                        VALUES (%s,%s,%s,%s,%s,%s)
                    """, (account_name, date_str, abs(amt), txn_type, desc, import_batch))
                    inserted += 1
    return jsonify({'inserted': inserted, 'batch': import_batch})

@bp.route('/api/finance/bank/statements', methods=['GET'])
@require_module('finance')
def api_bank_statements():
    month = request.args.get('month', '')
    conds, params = ['TRUE'], []
    if month:
        _d_s, _d_e = _month_date_range(month)
        conds.append("bs.txn_date >= %s AND bs.txn_date < %s"); params += [_d_s, _d_e]
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT bs.*, fr.title as matched_title, fr.amount as matched_amount,
                   fr.record_date as matched_date
            FROM bank_statements bs
            LEFT JOIN finance_records fr ON fr.id=bs.matched_record_id
            WHERE {' AND '.join(conds)}
            ORDER BY bs.txn_date DESC, bs.id DESC
        """, params).fetchall()
    result = []
    for r in rows:
        d = _bank_row(r)
        d['matched_title']  = r['matched_title']
        d['matched_amount'] = float(r['matched_amount']) if r['matched_amount'] else None
        d['matched_date']   = str(r['matched_date']) if r['matched_date'] else None
        result.append(d)
    return jsonify(result)

@bp.route('/api/finance/bank/statements/<int:sid>', methods=['DELETE'])
@require_module('finance')
def api_bank_statement_delete(sid):
    with get_db() as conn:
        conn.execute("DELETE FROM bank_statements WHERE id=%s", (sid,))
    return jsonify({'deleted': sid})

@bp.route('/api/finance/bank/match', methods=['POST'])
@require_module('finance')
def api_bank_match():
    b = request.get_json(force=True)
    sid = b.get('statement_id')
    rid = b.get('record_id')  # None to unmatch
    with get_db() as conn:
        if rid:
            conn.execute("UPDATE bank_statements SET reconciled=TRUE, matched_record_id=%s WHERE id=%s",
                         (rid, sid))
        else:
            conn.execute("UPDATE bank_statements SET reconciled=FALSE, matched_record_id=NULL WHERE id=%s",
                         (sid,))
    return jsonify({'ok': True})

@bp.route('/api/finance/bank/auto-match', methods=['POST'])
@require_module('finance')
def api_bank_auto_match():
    """自動比對:相同金額且日期在 3 天內"""
    b = request.get_json(force=True)
    month = b.get('month', '')
    matched = 0
    _bm_d = _month_date_range(month) if month else None
    with get_db() as conn:
        stmts = conn.execute("""
            SELECT * FROM bank_statements
            WHERE reconciled=FALSE
            """ + ("AND txn_date >= %s AND txn_date < %s" if month else ""),
            (list(_bm_d) if month else [])).fetchall()
        for s in stmts:
            # Find finance record: same amount, date within ±3 days, same type direction
            ftype = 'income' if s['txn_type'] == 'credit' else 'expense'
            rec = conn.execute("""
                SELECT id FROM finance_records
                WHERE type=%s AND amount=%s
                  AND ABS(record_date - %s::date) <= 3
                  AND id NOT IN (
                      SELECT matched_record_id FROM bank_statements
                      WHERE matched_record_id IS NOT NULL
                  )
                ORDER BY ABS(record_date - %s::date), id
                LIMIT 1
            """, (ftype, s['amount'], s['txn_date'], s['txn_date'])).fetchone()
            if rec:
                conn.execute("""UPDATE bank_statements SET reconciled=TRUE, matched_record_id=%s
                                WHERE id=%s""", (rec['id'], s['id']))
                matched += 1
    return jsonify({'matched': matched})

@bp.route('/api/finance/bank/summary', methods=['GET'])
@require_module('finance')
def api_bank_summary():
    month = request.args.get('month', '')
    cond = "AND txn_date >= %s AND txn_date < %s" if month else ""
    params = list(_month_date_range(month)) if month else []
    with get_db() as conn:
        r = conn.execute(f"""
            SELECT
              COUNT(*) as total,
              SUM(CASE WHEN reconciled THEN 1 ELSE 0 END) as matched,
              SUM(CASE WHEN txn_type='credit' THEN amount ELSE 0 END) as total_credit,
              SUM(CASE WHEN txn_type='debit'  THEN amount ELSE 0 END) as total_debit,
              SUM(CASE WHEN reconciled AND txn_type='credit' THEN amount ELSE 0 END) as matched_credit,
              SUM(CASE WHEN reconciled AND txn_type='debit'  THEN amount ELSE 0 END) as matched_debit
            FROM bank_statements WHERE TRUE {cond}
        """, params).fetchone()
    d = dict(r)
    for k in d:
        if d[k] is not None: d[k] = float(d[k]) if isinstance(d[k], type(r['total_credit'])) else int(d[k])
    return jsonify(d)

@bp.route('/api/finance/tax/<int:year>/<int:period>', methods=['GET'])
@require_module('finance')
def api_finance_tax(year, period):
    """
    period: 1=Jan-Feb, 2=Mar-Apr, 3=May-Jun, 4=Jul-Aug, 5=Sep-Oct, 6=Nov-Dec
    """
    if period < 1 or period > 6:
        return jsonify({'error': '期別需為 1-6'}), 400
    m_start = (period - 1) * 2 + 1
    m_end   = m_start + 1
    months  = [f"{year}-{str(m).zfill(2)}" for m in range(m_start, m_end + 1)]
    _d_s = _month_date_range(f"{year}-{m_start:02d}")[0]
    _d_e = _month_date_range(f"{year}-{m_end:02d}")[1]

    with get_db() as conn:
        rows = conn.execute("""
            SELECT fr.type, fr.amount, fr.tax_amount, fr.title,
                   fr.vendor, fr.invoice_no, fr.record_date,
                   fc.name as category_name
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            WHERE fr.record_date >= %s AND fr.record_date < %s
            ORDER BY fr.record_date, fr.type
        """, (_d_s, _d_e)).fetchall()

    sales_rows    = [r for r in rows if r['type'] == 'income']
    purchase_rows = [r for r in rows if r['type'] == 'expense']

    sales_amount    = sum(float(r['amount'])     for r in sales_rows)
    sales_tax       = sum(float(r['tax_amount'] or 0) for r in sales_rows)
    purchase_amount = sum(float(r['amount'])     for r in purchase_rows)
    purchase_tax    = sum(float(r['tax_amount'] or 0) for r in purchase_rows)
    tax_payable     = round(sales_tax - purchase_tax, 2)

    def _fmt_row(r):
        return {
            'date':     str(r['record_date']),
            'title':    r['title'],
            'vendor':   r['vendor'] or '',
            'invoice_no': r['invoice_no'] or '',
            'amount':   float(r['amount']),
            'tax_amount': float(r['tax_amount'] or 0),
            'category': r['category_name'] or '未分類',
        }

    return jsonify({
        'year': year, 'period': period,
        'roc_year': year - 1911,
        'months': months,
        'sales': {
            'rows':   [_fmt_row(r) for r in sales_rows],
            'amount': round(sales_amount, 2),
            'tax':    round(sales_tax, 2),
        },
        'purchases': {
            'rows':   [_fmt_row(r) for r in purchase_rows],
            'amount': round(purchase_amount, 2),
            'tax':    round(purchase_tax, 2),
        },
        'tax_payable': tax_payable,
        'is_refund':   tax_payable < 0,
    })

@bp.route('/api/finance/payables', methods=['GET'])
@require_module('finance')
def api_payables_list():
    from datetime import date as _d
    ptype  = request.args.get('type', '')      # receivable / payable
    status = request.args.get('status', '')    # open / paid / overdue
    conds, params = ['TRUE'], []
    if ptype:  conds.append("payable_type=%s"); params.append(ptype)
    if status == 'overdue':
        conds.append("status='open' AND due_date < CURRENT_DATE")
    elif status:
        conds.append("status=%s"); params.append(status)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT *, CURRENT_DATE - due_date AS days_overdue
            FROM finance_payables
            WHERE {' AND '.join(conds)}
            ORDER BY
              CASE WHEN status='open' AND due_date < CURRENT_DATE THEN 0
                   WHEN status='open' THEN 1
                   ELSE 2 END,
              due_date
        """, params).fetchall()
    result = []
    for r in rows:
        d = _payable_row(r)
        d['days_overdue'] = int(r['days_overdue']) if r['days_overdue'] is not None else 0
        # Compute effective status
        if d['status'] == 'open' and d.get('due_date') and str(_d.today()) > d['due_date']:
            d['effective_status'] = 'overdue'
        else:
            d['effective_status'] = d['status']
        result.append(d)
    return jsonify(result)

@bp.route('/api/finance/payables', methods=['POST'])
@require_module('finance')
def api_payable_create():
    b = request.get_json(force=True)
    if not b.get('title','').strip(): return jsonify({'error': '標題為必填'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO finance_payables
              (payable_type, title, party_name, invoice_no, amount, due_date, status, note)
            VALUES (%s,%s,%s,%s,%s,%s,'open',%s) RETURNING *
        """, (b.get('payable_type','payable'), b['title'].strip(),
              b.get('party_name','').strip(), b.get('invoice_no','').strip(),
              float(b.get('amount',0)), b.get('due_date') or None,
              b.get('note','').strip())).fetchone()
    return jsonify(_payable_row(row)), 201

@bp.route('/api/finance/payables/<int:pid>', methods=['PUT'])
@require_module('finance')
def api_payable_update(pid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE finance_payables SET
              payable_type=%s, title=%s, party_name=%s, invoice_no=%s,
              amount=%s, due_date=%s, status=%s,
              paid_date=%s, note=%s, updated_at=NOW()
            WHERE id=%s RETURNING *
        """, (b.get('payable_type','payable'), b.get('title','').strip(),
              b.get('party_name','').strip(), b.get('invoice_no','').strip(),
              float(b.get('amount',0)), b.get('due_date') or None,
              b.get('status','open'), b.get('paid_date') or None,
              b.get('note','').strip(), pid)).fetchone()
    return jsonify(_payable_row(row)) if row else ('', 404)

@bp.route('/api/finance/payables/<int:pid>', methods=['DELETE'])
@require_module('finance')
def api_payable_delete(pid):
    with get_db() as conn:
        conn.execute("DELETE FROM finance_payables WHERE id=%s", (pid,))
    return jsonify({'deleted': pid})

@bp.route('/api/finance/payables/aging', methods=['GET'])
@require_module('finance')
def api_payables_aging():
    ptype = request.args.get('type', 'payable')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT *,
              CURRENT_DATE - due_date AS days_overdue
            FROM finance_payables
            WHERE payable_type=%s AND status='open'
        """, (ptype,)).fetchall()
    buckets = {'current': 0, 'd1_30': 0, 'd31_60': 0, 'd61_90': 0, 'd90plus': 0}
    bucket_rows = {'current': [], 'd1_30': [], 'd31_60': [], 'd61_90': [], 'd90plus': []}
    for r in rows:
        do = int(r['days_overdue']) if r['days_overdue'] is not None else 0
        d = _payable_row(r)
        d['days_overdue'] = do
        if do <= 0:    k = 'current'
        elif do <= 30: k = 'd1_30'
        elif do <= 60: k = 'd31_60'
        elif do <= 90: k = 'd61_90'
        else:          k = 'd90plus'
        buckets[k]      += float(r['amount'])
        bucket_rows[k].append(d)
    return jsonify({'buckets': buckets, 'rows': bucket_rows, 'type': ptype})

@bp.route('/api/finance/budgets', methods=['GET'])
@require_module('finance')
def api_budgets_list():
    year  = request.args.get('year',  '')
    month = request.args.get('month', '')
    if not year or not month:
        from datetime import datetime, timezone, timedelta
        now = datetime.now(timezone(timedelta(hours=8)))
        year, month = str(now.year), str(now.month)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT fb.*, fc.name as category_name, fc.type as category_type, fc.color
            FROM finance_budgets fb
            JOIN finance_categories fc ON fc.id=fb.category_id
            WHERE fb.year=%s AND fb.month=%s
            ORDER BY fc.type, fc.sort_order
        """, (int(year), int(month))).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d['budget_amount'] = float(d['budget_amount'])
        if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
        if d.get('updated_at'): d['updated_at'] = d['updated_at'].isoformat()
        result.append(d)
    return jsonify(result)

@bp.route('/api/finance/budgets', methods=['POST'])
@require_module('finance')
def api_budgets_save():
    """Upsert list of budgets: [{category_id, budget_amount}]"""
    b = request.get_json(force=True)
    year  = int(b.get('year',  0))
    month = int(b.get('month', 0))
    items = b.get('items', [])
    if not year or not month: return jsonify({'error': '年月為必填'}), 400
    with get_db() as conn:
        for it in items:
            cid = it.get('category_id')
            amt = float(it.get('budget_amount', 0))
            if cid is None: continue
            if amt == 0:
                conn.execute("DELETE FROM finance_budgets WHERE year=%s AND month=%s AND category_id=%s",
                             (year, month, cid))
            else:
                conn.execute("""
                    INSERT INTO finance_budgets (year, month, category_id, budget_amount, updated_at)
                    VALUES (%s,%s,%s,%s,NOW())
                    ON CONFLICT (year, month, category_id)
                    DO UPDATE SET budget_amount=EXCLUDED.budget_amount, updated_at=NOW()
                """, (year, month, cid, amt))
    return jsonify({'ok': True})

@bp.route('/api/finance/budgets/vs-actual', methods=['GET'])
@require_module('finance')
def api_budgets_vs_actual():
    year  = request.args.get('year',  '')
    month = request.args.get('month', '')
    if not year or not month:
        from datetime import datetime, timezone, timedelta
        now = datetime.now(timezone(timedelta(hours=8)))
        year, month = str(now.year), str(now.month)
    period = f"{year}-{str(month).zfill(2)}"
    with get_db() as conn:
        cats = conn.execute("""
            SELECT id, name, type, color FROM finance_categories WHERE active=TRUE ORDER BY type, sort_order
        """).fetchall()
        budgets = conn.execute("""
            SELECT category_id, budget_amount FROM finance_budgets WHERE year=%s AND month=%s
        """, (int(year), int(month))).fetchall()
        _d_s, _d_e = _month_date_range(period)
        actuals = conn.execute("""
            SELECT category_id, SUM(amount) as total
            FROM finance_records
            WHERE record_date >= %s AND record_date < %s
            GROUP BY category_id
        """, (_d_s, _d_e)).fetchall()
    budget_map = {r['category_id']: float(r['budget_amount']) for r in budgets}
    actual_map = {r['category_id']: float(r['total']) for r in actuals}
    result = []
    for c in cats:
        cid = c['id']
        bgt = budget_map.get(cid, 0)
        act = actual_map.get(cid, 0)
        pct = round(act / bgt * 100, 1) if bgt > 0 else None
        result.append({
            'category_id':   cid,
            'category_name': c['name'],
            'category_type': c['type'],
            'color':         c['color'],
            'budget':        bgt,
            'actual':        act,
            'remaining':     round(bgt - act, 2),
            'pct':           pct,
            'over_budget':   bgt > 0 and act > bgt,
        })
    return jsonify({'year': year, 'month': month, 'items': result})

@bp.route('/api/finance/payroll/status', methods=['GET'])
@require_module('finance')
def api_payroll_status():
    """列出各月薪資是否已同步至財務"""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT month,
                   COUNT(*) as total,
                   SUM(CASE WHEN finance_synced THEN 1 ELSE 0 END) as synced,
                   SUM(net_pay) as total_net_pay
            FROM salary_records
            WHERE status IN ('confirmed','draft')
            GROUP BY month ORDER BY month DESC LIMIT 24
        """).fetchall()
    return jsonify([{
        'month':         r['month'],
        'total':         int(r['total']),
        'synced':        int(r['synced']),
        'total_net_pay': float(r['total_net_pay'] or 0),
        'all_synced':    int(r['synced']) == int(r['total']),
    } for r in rows])

@bp.route('/api/finance/payroll/sync', methods=['POST'])
@require_module('finance')
def api_payroll_sync():
    """將指定月份已確認薪資寫入財務記錄"""
    b     = request.get_json(force=True)
    month = b.get('month', '')
    if not month: return jsonify({'error': '請提供月份'}), 400
    # Find or create 薪資支出 category
    with get_db() as conn:
        cat = conn.execute("""
            SELECT id FROM finance_categories WHERE name='薪資支出' AND type='expense' LIMIT 1
        """).fetchone()
        if not cat:
            cat = conn.execute("""
                INSERT INTO finance_categories (name,type,color,sort_order)
                VALUES ('薪資支出','expense','#e07b2a',11) RETURNING *
            """).fetchone()
        cat_id = cat['id']

        records = conn.execute("""
            SELECT sr.*, ps.name as staff_name
            FROM salary_records sr
            JOIN punch_staff ps ON ps.id=sr.staff_id
            WHERE sr.month=%s AND sr.finance_synced=FALSE
        """, (month,)).fetchall()

        if not records:
            return jsonify({'created': 0, 'message': '無需同步的薪資記錄'})

        record_date = f"{month}-{str(28).zfill(2)}"  # Use 28th as payroll date
        created = 0
        for sr in records:
            # Main salary entry
            conn.execute("""
                INSERT INTO finance_records
                  (record_date, category_id, type, title, amount, note, created_by)
                VALUES (%s,%s,'expense',%s,%s,%s,'payroll-sync')
            """, (record_date, cat_id,
                  f"{sr['staff_name']} {month} 薪資",
                  float(sr['net_pay']),
                  f"薪資記錄 #{sr['id']}"))
            conn.execute("UPDATE salary_records SET finance_synced=TRUE WHERE id=%s", (sr['id'],))
            created += 1

    return jsonify({'created': created, 'month': month})

@bp.route('/api/finance/tax/<int:year>/<int:period>/sync', methods=['POST'])
@require_module('finance')
def api_finance_tax_sync(year, period):
    """將應繳/退稅金額建立為財務分錄,流入損益表"""
    if period < 1 or period > 6:
        return jsonify({'error': '期別需為 1-6'}), 400
    m_start = (period - 1) * 2 + 1
    m_end   = m_start + 1
    months  = [f"{year}-{str(m).zfill(2)}" for m in range(m_start, m_end + 1)]
    roc_year = year - 1911
    _d_s = _month_date_range(f"{year}-{m_start:02d}")[0]
    _d_e = _month_date_range(f"{year}-{m_end:02d}")[1]

    with get_db() as conn:
        rows = conn.execute("""
            SELECT type, SUM(tax_amount) as tax_total
            FROM finance_records
            WHERE record_date >= %s AND record_date < %s
              AND tax_amount IS NOT NULL AND tax_amount <> 0
            GROUP BY type
        """, (_d_s, _d_e)).fetchall()

    sales_tax    = sum(float(r['tax_total']) for r in rows if r['type'] == 'income')
    purchase_tax = sum(float(r['tax_total']) for r in rows if r['type'] == 'expense')
    tax_payable  = round(sales_tax - purchase_tax, 2)

    if tax_payable == 0:
        return jsonify({'created': 0, 'message': '稅額為零,無需建立分錄'})

    # Record date = last day of period's last month
    import calendar as _cal
    record_date = f"{year}-{str(m_end).zfill(2)}-{_cal.monthrange(year, m_end)[1]}"
    note = f"銷項稅 ${round(sales_tax,0):,.0f} − 進項稅 ${round(purchase_tax,0):,.0f} = {'應繳' if tax_payable>0 else '退稅'} ${abs(round(tax_payable,0)):,.0f}"
    period_label = f"民國{roc_year}年第{period}期({months[0]}~{months[-1]})"

    created = 0
    with get_db() as conn:
        # 冪等性:同一期別已有 tax-sync 記錄就不重複建立
        existing = conn.execute("""
            SELECT id FROM finance_records
            WHERE created_by='tax-sync' AND record_date=%s
              AND title LIKE %s LIMIT 1
        """, (record_date, f'%{period_label}%')).fetchone()
        if existing:
            return jsonify({'created': 0, 'message': f'該期別({period_label})已建立過分錄,略過.',
                            'tax_payable': tax_payable})

        if tax_payable > 0:
            cat = conn.execute(
                "SELECT id FROM finance_categories WHERE name='稅費' AND type='expense' LIMIT 1"
            ).fetchone()
            if not cat:
                cat = conn.execute("""
                    INSERT INTO finance_categories (name, type, color, sort_order, statement_section)
                    VALUES ('稅費','expense','#8892a4', 99,'operating_expense') RETURNING *
                """).fetchone()
            conn.execute("""
                INSERT INTO finance_records
                  (record_date, category_id, type, title, amount, tax_amount, note, created_by)
                VALUES (%s,%s,'expense',%s,%s,0,%s,'tax-sync')
            """, (record_date, cat['id'],
                  f"應繳營業稅 {period_label}", tax_payable, note))
            created += 1
        else:
            cat = conn.execute(
                "SELECT id FROM finance_categories WHERE name='其他收入' AND type='income' LIMIT 1"
            ).fetchone()
            if not cat:
                cat = conn.execute("""
                    INSERT INTO finance_categories (name, type, color, sort_order, statement_section)
                    VALUES ('其他收入','income','#c8a96e', 99,'other_revenue') RETURNING *
                """).fetchone()
            conn.execute("""
                INSERT INTO finance_records
                  (record_date, category_id, type, title, amount, tax_amount, note, created_by)
                VALUES (%s,%s,'income',%s,%s,0,%s,'tax-sync')
            """, (record_date, cat['id'],
                  f"營業稅退稅 {period_label}", abs(tax_payable), note))
            created += 1

    return jsonify({'created': created, 'tax_payable': tax_payable, 'record_date': record_date})
