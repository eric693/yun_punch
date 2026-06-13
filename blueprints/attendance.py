"""attendance API blueprint(自 app.py 拆出)。"""
from flask import Blueprint, request, jsonify
from db import (
    get_db,
)
from auth import (
    login_required,
)
from config import (
    TW_TZ,
)
from utils import (
    _month_ts_range, _month_date_range,
)
from services import (
    _is_holiday,
)
from datetime import datetime as _dt
from datetime import timedelta as _td
from datetime import timezone as _tz

bp = Blueprint('attendance', __name__)


@bp.route('/api/attendance/monthly-stats', methods=['GET'])
@login_required
def api_attendance_monthly_stats():
    """
    月出勤統計報表(每位員工匯總)
    回傳:出勤天數、總工時、遲到次數、缺打卡次數、平均工時
    """
    month = request.args.get('month') or _dt.now(TW_TZ).strftime('%Y-%m')
    _ts_s, _ts_e = _month_ts_range(month)
    _d_s, _d_e   = _month_date_range(month)
    with get_db() as conn:
        # 每人每日打卡彙整
        rows = [dict(r) for r in conn.execute("""
            SELECT ps.id as staff_id, ps.name as staff_name,
                   ps.department, ps.role,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   MIN(CASE WHEN pr.punch_type='in'  THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as clock_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as clock_out,
                   BOOL_OR(pr.punch_type='in')  as has_in,
                   BOOL_OR(pr.punch_type='out') as has_out
            FROM punch_records pr
            JOIN punch_staff ps ON ps.id = pr.staff_id AND ps.active = TRUE
            WHERE pr.punched_at >= %s AND pr.punched_at < %s
            GROUP BY ps.id, ps.name, ps.department, ps.role,
                     (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY ps.name, work_date
        """, (_ts_s, _ts_e)).fetchall()]

        # 跨日班次合併:day N 有上班無下班 + day N+1 有下班無上班 -> 歸入 day N
        from datetime import date as _dcs, timedelta as _tdcs
        _stat_map = {(r['staff_id'], str(r['work_date'])): r for r in rows}
        _stat_skip = set()
        for r in rows:
            sid = r['staff_id']
            ds  = str(r['work_date'])
            if r['has_in'] and not r['has_out']:
                next_ds = (_dcs.fromisoformat(ds) + _tdcs(days=1)).isoformat()
                nr = _stat_map.get((sid, next_ds))
                if nr and nr['has_out'] and not nr['has_in']:
                    r['clock_out'] = nr['clock_out']
                    r['has_out']   = True
                    _stat_skip.add((sid, next_ds))
        rows = [r for r in rows if (r['staff_id'], str(r['work_date'])) not in _stat_skip]

        # 班別指派(用於遲到判斷)
        shift_rows = conn.execute("""
            SELECT sa.staff_id, sa.shift_date, st.start_time, st.end_time
            FROM shift_assignments sa
            JOIN shift_types st ON st.id = sa.shift_type_id
            WHERE sa.shift_date >= %s AND sa.shift_date < %s
        """, (_d_s, _d_e)).fetchall()
        shift_map = {(r['staff_id'], str(r['shift_date'])): r for r in shift_rows}

    from collections import defaultdict
    stats = defaultdict(lambda: {
        'staff_id': None, 'staff_name': '', 'department': '', 'role': '',
        'days_worked': 0, 'total_minutes': 0,
        'late_count': 0, 'early_count': 0, 'missing_in_count': 0, 'missing_out_count': 0,
        'anomaly_dates': [],
    })

    for r in rows:
        sid  = r['staff_id']
        ds   = str(r['work_date'])
        s    = stats[sid]
        s['staff_id']   = sid
        s['staff_name'] = r['staff_name']
        s['department'] = r['department'] or ''
        s['role']       = r['role']       or ''

        has_in  = bool(r['has_in'])
        has_out = bool(r['has_out'])

        if has_in or has_out:
            s['days_worked'] += 1

        if r['clock_in'] and r['clock_out']:
            diff = (r['clock_out'] - r['clock_in']).total_seconds() / 60
            if diff > 0:
                s['total_minutes'] += int(diff)

        # 缺打卡
        if has_in and not has_out:
            s['missing_out_count'] += 1
            s['anomaly_dates'].append({'date': ds, 'type': 'missing_out', 'label': '缺下班卡'})
        if not has_in and has_out:
            s['missing_in_count'] += 1
            s['anomaly_dates'].append({'date': ds, 'type': 'missing_in', 'label': '缺上班卡'})

        # 遲到(比對班別)
        if has_in and r['clock_in']:
            shift = shift_map.get((sid, ds))
            if shift and shift['start_time']:
                try:
                    sh, sm = map(int, str(shift['start_time'])[:5].split(':'))
                    ci_local = r['clock_in']
                    ih, im   = ci_local.hour, ci_local.minute
                    late_mins = (ih * 60 + im) - (sh * 60 + sm)
                    if late_mins > 10:
                        s['late_count'] += 1
                        s['anomaly_dates'].append({'date': ds, 'type': 'late',
                                                   'label': f'遲到 {late_mins} 分鐘'})
                except Exception:
                    pass

        # 早退(比對班別)
        if has_out and r['clock_out']:
            shift = shift_map.get((sid, ds))
            if shift and shift['end_time']:
                try:
                    eh, em = map(int, str(shift['end_time'])[:5].split(':'))
                    co_local = r['clock_out']
                    oh, om   = co_local.hour, co_local.minute
                    early_mins = (eh * 60 + em) - (oh * 60 + om)
                    if early_mins > 15:
                        s['early_count'] += 1
                        s['anomaly_dates'].append({'date': ds, 'type': 'early',
                                                   'label': f'早退 {early_mins} 分鐘'})
                except Exception:
                    pass

    result = []
    for s in sorted(stats.values(), key=lambda x: (x['department'], x['staff_name'])):
        h   = s['total_minutes'] // 60
        m   = s['total_minutes'] % 60
        avg = round(s['total_minutes'] / s['days_worked'] / 60, 1) if s['days_worked'] else 0
        s['total_hours']   = round(s['total_minutes'] / 60, 1)
        s['avg_hours_day'] = avg
        s['total_hm']      = f"{h}h {m:02d}m"
        result.append(s)
    return jsonify({'month': month, 'stats': result})

@bp.route('/api/attendance/anomaly-report', methods=['GET'])
@login_required
def api_anomaly_report_excel():
    """匯出出勤異常報告 Excel(缺打卡、遲到、早退)"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    import calendar as _cal
    from datetime import datetime as _dtx, timedelta as _tdx

    month = request.args.get('month', '') or _dt.now(TW_TZ).strftime('%Y-%m')
    try:
        y, mo = int(month[:4]), int(month[5:7])
    except Exception:
        return jsonify({'error': '月份格式錯誤'}), 400

    TW_OFF = _tdx(hours=8)
    _ts_s, _ts_e = _month_ts_range(month)
    _d_s, _d_e   = _month_date_range(month)

    with get_db() as conn:
        punch_rows = conn.execute("""
            SELECT ps.id as staff_id, ps.name as staff_name,
                   ps.department,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   MIN(CASE WHEN pr.punch_type='in'  THEN (pr.punched_at AT TIME ZONE 'Asia/Taipei') END) as clock_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN (pr.punched_at AT TIME ZONE 'Asia/Taipei') END) as clock_out,
                   BOOL_OR(pr.punch_type='in')  as has_in,
                   BOOL_OR(pr.punch_type='out') as has_out
            FROM punch_records pr
            JOIN punch_staff ps ON ps.id=pr.staff_id AND ps.active=TRUE
            WHERE pr.punched_at >= %s AND pr.punched_at < %s
            GROUP BY ps.id, ps.name, ps.department,
                     (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY work_date, ps.name
        """, (_ts_s, _ts_e)).fetchall()

        shift_rows = conn.execute("""
            SELECT sa.staff_id, sa.shift_date,
                   st.start_time::text as start_time,
                   st.end_time::text   as end_time
            FROM shift_assignments sa
            JOIN shift_types st ON st.id=sa.shift_type_id
            WHERE sa.shift_date >= %s AND sa.shift_date < %s
        """, (_d_s, _d_e)).fetchall()

        y_int = int(month[:4]); mo_int = int(month[5:7])
        first_day = f"{y_int}-{mo_int:02d}-01"
        days_in   = _cal.monthrange(y_int, mo_int)[1]
        last_day  = f"{y_int}-{mo_int:02d}-{days_in:02d}"
        leave_rows = conn.execute("""
            SELECT staff_id, start_date, end_date
            FROM leave_requests
            WHERE status='approved'
              AND start_date <= %s AND end_date >= %s
        """, (last_day, first_day)).fetchall()

    # Build lookup maps
    shift_map = {(r['staff_id'], str(r['shift_date'])): r for r in shift_rows}
    leave_set = set()
    from datetime import date as _dax, timedelta as _tdax
    for lr in leave_rows:
        s = lr['start_date']; e = lr['end_date']
        cur = s
        while cur <= e:
            leave_set.add((lr['staff_id'], str(cur)))
            cur = _dax.fromisoformat(str(cur)) + _tdax(days=1)
            cur = cur if isinstance(cur, _dax) else cur.date()

    today = _dax.today()

    # Build anomaly rows
    anomalies = []
    for r in punch_rows:
        ds = str(r['work_date'])
        sid = r['staff_id']
        shift = shift_map.get((sid, ds))

        anomaly_type = ''; detail = ''
        late_min = 0; early_min = 0

        if not r['has_in'] and r['has_out']:
            anomaly_type = '缺上班打卡'; detail = f"僅有下班 {str(r['clock_out'])[11:16]}"
        elif r['has_in'] and not r['has_out']:
            if _dax.fromisoformat(ds) < today:
                anomaly_type = '缺下班打卡'; detail = f"上班 {str(r['clock_in'])[11:16]} 無下班"
        elif r['has_in'] and r['has_out'] and shift:
            ci_t = str(r['clock_in'])[11:16]
            co_t = str(r['clock_out'])[11:16]
            sh_s = str(shift['start_time'])[:5]
            sh_e = str(shift['end_time'])[:5]
            try:
                ci_m = int(ci_t[:2])*60 + int(ci_t[3:5])
                sh_s_m = int(sh_s[:2])*60 + int(sh_s[3:5])
                if ci_m - sh_s_m > 10:
                    late_min = ci_m - sh_s_m
                    anomaly_type = '遲到'; detail = f"應 {sh_s},實際 {ci_t}(+{late_min}分)"
            except Exception:
                pass
            if not anomaly_type:
                try:
                    co_m = int(co_t[:2])*60 + int(co_t[3:5])
                    sh_e_m = int(sh_e[:2])*60 + int(sh_e[3:5])
                    if sh_e_m - co_m > 15:
                        early_min = sh_e_m - co_m
                        anomaly_type = '早退'; detail = f"應 {sh_e},實際 {co_t}(-{early_min}分)"
                except Exception:
                    pass

        if anomaly_type:
            anomalies.append({
                'staff_name':  r['staff_name'],
                'department':  r['department'] or '',
                'date':        ds,
                'shift_start': str(shift['start_time'])[:5] if shift else '-',
                'shift_end':   str(shift['end_time'])[:5]   if shift else '-',
                'clock_in':    str(r['clock_in'])[11:16]  if r['clock_in']  else '-',
                'clock_out':   str(r['clock_out'])[11:16] if r['clock_out'] else '-',
                'anomaly_type': anomaly_type,
                'detail':       detail,
            })

    # Build Excel
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = f'{month} 異常明細'

    thin = Border(
        left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin',  color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'),
    )
    header_fill   = PatternFill('solid', fgColor='0F1C3A')
    warn_fill     = PatternFill('solid', fgColor='FFF3CD')
    err_fill      = PatternFill('solid', fgColor='FDECEA')
    center_align  = Alignment(horizontal='center', vertical='center')

    headers = ['員工姓名','部門','日期','應上班','應下班','實際上班','實際下班','異常類型','說明']
    col_w   = [12, 10, 12, 8, 8, 8, 8, 12, 30]
    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color='FFFFFF', name='Noto Sans TC', size=11)
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    for ri, a in enumerate(anomalies, 2):
        row_fill = err_fill if a['anomaly_type'] in ('缺上班打卡','缺下班打卡') else warn_fill
        vals = [a['staff_name'], a['department'], a['date'],
                a['shift_start'], a['shift_end'],
                a['clock_in'], a['clock_out'],
                a['anomaly_type'], a['detail']]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill      = row_fill
            cell.alignment = center_align if ci != 9 else Alignment(vertical='center')
            cell.border    = thin

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    # Summary sheet
    ws2 = wb.create_sheet('摘要')
    ws2.append(['統計', '數量'])
    ws2.append(['異常總筆數', len(anomalies)])
    by_type = {}
    for a in anomalies:
        by_type[a['anomaly_type']] = by_type.get(a['anomaly_type'], 0) + 1
    for t, c in sorted(by_type.items(), key=lambda x: -x[1]):
        ws2.append([t, c])

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import Response as _FR
    return _FR(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=anomaly_{month}.xlsx'}
    )

@bp.route('/api/attendance/anomalies', methods=['GET'])
@login_required
def api_attendance_anomalies():
    """
    偵測出勤異常:
    - 忘記打下班卡(有上班無下班)
    - 只有下班無上班
    - 遲到(上班時間晚於班別開始時間)
    """
    from datetime import date as _da, datetime as _dta, timezone as _tz, timedelta as _td
    TW    = _tz(_td(hours=8))
    today = _dta.now(TW).date()
    # Check last 7 days
    date_from = today - _td(days=7)

    with get_db() as conn:
        # 取得最近7天打卡記錄(按人、按天)
        rows = conn.execute("""
            SELECT ps.id as staff_id, ps.name, ps.role, ps.department,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   array_agg(pr.punch_type ORDER BY pr.punched_at) as types,
                   MIN(CASE WHEN pr.punch_type='in'  THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as first_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as last_out
            FROM punch_records pr
            JOIN punch_staff ps ON ps.id = pr.staff_id
            WHERE (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date BETWEEN %s AND %s
              AND ps.active = TRUE
            GROUP BY ps.id, ps.name, ps.role, ps.department,
                     (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY work_date DESC, ps.name
        """, (date_from, today)).fetchall()

        # 取得班別指派(用於遲到/早退判斷)
        shift_rows = conn.execute("""
            SELECT sa.staff_id, sa.shift_date, st.start_time, st.end_time, st.name as shift_name
            FROM shift_assignments sa
            JOIN shift_types st ON st.id = sa.shift_type_id
            WHERE sa.shift_date BETWEEN %s AND %s
        """, (date_from, today)).fetchall()
        shift_map = {(r['staff_id'], str(r['shift_date'])): r for r in shift_rows}

        # 今日應出勤但未出勤(排除請假)
        all_staff = conn.execute(
            "SELECT id, name, role, department FROM punch_staff WHERE active=TRUE"
        ).fetchall()
        today_punched_ids = {r['staff_id'] for r in rows if str(r['work_date']) == str(today)}
        on_leave_today_ids = set()
        leave_today = conn.execute("""
            SELECT DISTINCT staff_id FROM leave_requests
            WHERE status='approved' AND start_date <= %s AND end_date >= %s
        """, (today, today)).fetchall()
        for r in leave_today:
            on_leave_today_ids.add(r['staff_id'])

        # 今日有排班的員工(用於判斷誰「應出勤」);以及今日是否為假日/週末
        scheduled_today_rows = conn.execute(
            "SELECT DISTINCT staff_id FROM shift_assignments WHERE shift_date=%s", (today,)
        ).fetchall()
        scheduled_today_ids = {r['staff_id'] for r in scheduled_today_rows}
        today_is_holiday    = _is_holiday(conn, str(today))
        today_is_weekend    = today.weekday() >= 5

    anomalies = []

    # 跨日班別合併:N 日有上班無下班 + N+1 日有下班無上班 -> 不重複報缺卡
    from datetime import date as _da_x, timedelta as _td_x
    _row_map = {(r['staff_id'], str(r['work_date'])): r for r in rows}
    _skip_missing_out = set()
    _skip_missing_in  = set()
    for r in rows:
        _t = list(r['types']) if r['types'] else []
        if 'in' in _t and 'out' not in _t:
            _nd = (_da_x.fromisoformat(str(r['work_date'])) + _td_x(days=1)).isoformat()
            _nr = _row_map.get((r['staff_id'], _nd))
            if _nr:
                _nt = list(_nr['types']) if _nr['types'] else []
                if 'out' in _nt and 'in' not in _nt:
                    _skip_missing_out.add((r['staff_id'], str(r['work_date'])))
                    _skip_missing_in.add((r['staff_id'], _nd))

    # 1. 近7天:有上班但無下班卡
    for r in rows:
        types = list(r['types']) if r['types'] else []
        has_in  = 'in'  in types
        has_out = 'out' in types
        ds = str(r['work_date'])

        if has_in and not has_out and ds != str(today) \
                and (r['staff_id'], ds) not in _skip_missing_out:
            # 昨天或更早沒打下班卡(今天的可能還沒下班)
            anomalies.append({
                'type':       'missing_out',
                'label':      '忘記下班打卡',
                'severity':   'warning',
                'staff_id':   r['staff_id'],
                'name':       r['name'],
                'role':       r['role'] or '',
                'department': r['department'] or '',
                'date':       ds,
                'detail':     f"上班 {r['first_in']},無下班記錄",
            })

        if not has_in and has_out and (r['staff_id'], ds) not in _skip_missing_in:
            anomalies.append({
                'type':       'missing_in',
                'label':      '忘記上班打卡',
                'severity':   'warning',
                'staff_id':   r['staff_id'],
                'name':       r['name'],
                'role':       r['role'] or '',
                'department': r['department'] or '',
                'date':       ds,
                'detail':     f"下班 {r['last_out']},無上班記錄",
            })

        # 遲到判斷(有班別指派)
        if has_in and r['first_in']:
            shift = shift_map.get((r['staff_id'], ds))
            if shift and shift['start_time']:
                try:
                    sh, sm = map(int, str(shift['start_time'])[:5].split(':'))
                    ih, im = map(int, r['first_in'].split(':'))
                    late_mins = (ih * 60 + im) - (sh * 60 + sm)
                    if late_mins > 10:  # 超過10分鐘算遲到
                        anomalies.append({
                            'type':       'late',
                            'label':      '遲到',
                            'severity':   'warning',
                            'staff_id':   r['staff_id'],
                            'name':       r['name'],
                            'role':       r['role'] or '',
                            'department': r['department'] or '',
                            'date':       ds,
                            'detail':     f"應 {shift['start_time'][:5]} 上班,實際 {r['first_in']}(晚 {late_mins} 分鐘)",
                        })
                except Exception:
                    pass

        # 早退判斷(有班別指派);跨日合併進來的下班卡屬前一日班別,不在此判斷
        if has_out and r['last_out'] and ds != str(today) \
                and (r['staff_id'], ds) not in _skip_missing_in:
            shift = shift_map.get((r['staff_id'], ds))
            if shift and shift['end_time']:
                try:
                    eh, em = map(int, str(shift['end_time'])[:5].split(':'))
                    oh, om = map(int, r['last_out'].split(':'))
                    early_mins = (eh * 60 + em) - (oh * 60 + om)
                    if early_mins > 15:  # 超過15分鐘算早退
                        anomalies.append({
                            'type':       'early',
                            'label':      '早退',
                            'severity':   'warning',
                            'staff_id':   r['staff_id'],
                            'name':       r['name'],
                            'role':       r['role'] or '',
                            'department': r['department'] or '',
                            'date':       ds,
                            'detail':     f"應 {shift['end_time'][:5]} 下班,實際 {r['last_out']}(早 {early_mins} 分鐘)",
                        })
                except Exception:
                    pass

    # 2. 今日未出勤(不含請假);只報「今日應出勤」者
    #    有排班 -> 以排班為準;當日完全無人排班(或未使用排班) -> 以平日(非週末非假日)為準
    for s in all_staff:
        if s['id'] in today_punched_ids or s['id'] in on_leave_today_ids:
            continue
        if scheduled_today_ids:
            expected_today = s['id'] in scheduled_today_ids
        else:
            expected_today = (not today_is_weekend) and (not today_is_holiday)
        if expected_today:
            anomalies.append({
                'type':       'absent',
                'label':      '今日未出勤',
                'severity':   'error',
                'staff_id':   s['id'],
                'name':       s['name'],
                'role':       s['role'] or '',
                'department': s['department'] or '',
                'date':       str(today),
                'detail':     '今日尚無打卡記錄且未請假',
            })

    # Sort: error > warning > info, then by date desc
    sev_order = {'error': 0, 'warning': 1, 'info': 2}
    anomalies.sort(key=lambda x: (sev_order.get(x['severity'], 9), x['date']))
    return jsonify({'anomalies': anomalies, 'count': len(anomalies), 'checked_from': str(date_from)})
